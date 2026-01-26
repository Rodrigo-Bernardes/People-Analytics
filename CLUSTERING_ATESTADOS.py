# -*- coding: utf-8 -*-
"""
PROJETO DE CLUSTERING - ANÁLISE DE ATESTADOS MÉDICOS DAS ULs
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.cluster.hierarchy as sch
import scipy.stats as stats
from scipy.stats import zscore
from scipy.spatial.distance import pdist
from sklearn.cluster import AgglomerativeClustering, KMeans
from sklearn.metrics import silhouette_score, silhouette_samples
import pingouin as pg
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import base64
import os
import webbrowser
import logging
import sys
import gc
from datetime import datetime

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

print("✓ Bibliotecas importadas com sucesso.\n")

# ==============================================================================
# 1. CARREGAMENTO, PRÉ-PROCESSAMENTO E FORMATAÇÃO DE COLUNAS
# ==============================================================================

logging.info("="*80)
logging.info("1. CARREGAMENTO, PRÉ-PROCESSAMENTO E FORMATAÇÃO DE COLUNAS")
logging.info("="*80)

try:
    path_arquivo = r'X:\Gestao_de_Pessoas\Analytics\10 - Relatórios\10.4 - HC e Atestados Médicos\Controle_HC e Atestados.xlsb'
    dados_hc = pd.read_excel(path_arquivo, sheet_name='HC', engine='pyxlsb')
    
    # ========== FORMATAÇÃO DE COLUNAS ==========
    
    # Dicionário com todas as colunas a formatar
    colunas_formato = {
        # DATAS (origem Excel: 1899-12-30)
        'nascimento': 'data',
        'data_nasc_conjuge': 'data',
        'data_admissao': 'data',
        'data_rescisao': 'data',
        'ultimo_reajuste_individual': 'data',
        'ultimo_reajuste_coletivo': 'data',
        'data_inicio_ferias': 'data',
        'data_fim_ferias': 'data',
        'exp_45dias': 'data',
        'exp_90dias': 'data',
        'data_inicio_atestado': 'data',
        # NÚMEROS (valores monetários e inteiros)
        'salario_admissao': 'numero',
        'salario_atual': 'numero',
        'salario_total': 'numero',
        'saldo_fgts': 'numero',
        'horas_nao_trabalhadas': 'numero',
        'custo_afastamento': 'numero',
        'primeiro_atestado': 'numero',
        'dias_trabalhado': 'numero',
    }
    
    # Função para converter datas do Excel
    def converter_data_excel(valor):
        """Converte serial date do Excel para datetime"""
        if pd.isna(valor):
            return pd.NaT
        if isinstance(valor, (int, float)):
            # Excel serial date começa em 1899-12-30
            return pd.to_datetime(valor, unit='D', origin='1899-12-30', errors='coerce')
        return pd.to_datetime(valor, errors='coerce')
    
    # Aplicar formatação de DATAS
    for coluna in colunas_formato:
        if coluna in dados_hc.columns:
            if colunas_formato[coluna] == 'data':
                # Converter para datetime
                dados_hc[coluna] = dados_hc[coluna].apply(converter_data_excel)
                logging.info(f"✓ Coluna '{coluna}' convertida para DATA (formato: YYYY-MM-DD)")
    
    # Aplicar formatação de NÚMEROS
    for coluna in colunas_formato:
        if coluna in dados_hc.columns:
            if colunas_formato[coluna] == 'numero':
                # Converter para float, depois para int se apropriado
                dados_hc[coluna] = pd.to_numeric(dados_hc[coluna], errors='coerce')
                # Se for inteiro (dias, primeiro_atestado), converter para int
                if coluna in ['dias_trabalhado', 'primeiro_atestado']:
                    dados_hc[coluna] = dados_hc[coluna].fillna(0).astype(int)
                logging.info(f"✓ Coluna '{coluna}' convertida para NÚMERO")
    
    # Filtrar dados
    dados_hc = dados_hc.loc[(dados_hc['ano_atestado'] >= 2026) & (dados_hc['unidade'] == 'UL')].copy()
    logging.info(f"✓ Base de dados carregada com {dados_hc.shape[0]} registros.")
    
    # Selecionar variáveis para clustering
    hc_cluster = dados_hc[['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']].copy()
    
    # Remover valores faltantes
    hc_cluster = hc_cluster.dropna()
    logging.info(f"✓ Dados após limpeza: {hc_cluster.shape[0]} registros.")
    
    # Codificar logo
    caminho_imagem = 'Logo AFPESP.png'
    try:
        with open(caminho_imagem, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
        img_src_base64 = f"data:image/png;base64,{encoded_string}"
    except:
        img_src_base64 = ""
    
    logging.info("✓ SEÇÃO 1 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 1: {e}")
    import traceback
    traceback.print_exc()
    sys.exit("Script encerrado.")

# ==============================================================================
# 2. ANÁLISE EXPLORATÓRIA
# ==============================================================================

logging.info("="*80)
logging.info("2. ANÁLISE EXPLORATÓRIA DOS DADOS")
logging.info("="*80)

try:
    # Estatísticas descritivas
    estat_descritiva = hc_cluster.describe().round(2)
    tabela_estat = estat_descritiva.to_html(classes='tabela-estatistica', border=0)
    
    # Matriz de correlação
    corr = hc_cluster.corr()
    fig_corr = go.Figure(data=go.Heatmap(
        x=corr.columns,
        y=corr.index,
        z=np.array(corr),
        text=corr.values,
        texttemplate='%{text:.2f}',
        colorscale='Blues'
    ))
    fig_corr.update_layout(height=500, width=600, title='Matriz de Correlação')
    grafico_corr = fig_corr.to_html(full_html=False, include_plotlyjs='cdn')
    
    # Boxplots
    fig_boxplot = make_subplots(
        rows=2, cols=3,
        subplot_titles=('Dias Afastado', 'Idade', 'Primeiro Atestado', 'Salário', 'Dias Trabalhado', ''),
        specs=[[{'type': 'box'}, {'type': 'box'}, {'type': 'box'}],
               [{'type': 'box'}, {'type': 'box'}, {'type': 'box'}]]
    )
    
    fig_boxplot.add_trace(go.Box(y=hc_cluster['dias_afastado'], name='Dias Afastado', marker_color='#0070C0'), row=1, col=1)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['idade'], name='Idade', marker_color='#0070C0'), row=1, col=2)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['primeiro_atestado'], name='Primeiro Atestado', marker_color='#0070C0'), row=1, col=3)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['salario_total'], name='Salário', marker_color='#0070C0'), row=2, col=1)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['dias_trabalhado'], name='Dias Trabalhado', marker_color='#0070C0'), row=2, col=2)
    
    fig_boxplot.update_layout(height=600, showlegend=False, title_text="Boxplots - Variáveis Numéricas")
    boxplot_html = fig_boxplot.to_html(full_html=False, include_plotlyjs=False)
    
    logging.info("✓ SEÇÃO 2 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 2: {e}")
    import traceback
    traceback.print_exc()
    tabela_estat = ""
    grafico_corr = ""
    boxplot_html = ""

# ==============================================================================
# 3. PADRONIZAÇÃO E DETECÇÃO DE OUTLIERS
# ==============================================================================

logging.info("="*80)
logging.info("3. PADRONIZAÇÃO E DETECÇÃO DE OUTLIERS")
logging.info("="*80)

try:
    # Aplicar Z-Score
    hc_pad = hc_cluster.apply(zscore, ddof=1)
    
    # Detectar outliers
    Q1 = hc_pad.quantile(0.25)
    Q3 = hc_pad.quantile(0.75)
    IQR = Q3 - Q1
    lim_inferior = Q1 - 1.5 * IQR
    lim_superior = Q3 + 1.5 * IQR
    
    outliers = (hc_pad < lim_inferior) | (hc_pad > lim_superior)
    num_outliers = outliers.sum()
    
    logging.info(f"✓ Outliers detectados por variável:")
    for col in num_outliers.index:
        logging.info(f"  - {col}: {num_outliers[col]}")
    
    logging.info("✓ SEÇÃO 3 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 3: {e}")
    import traceback
    traceback.print_exc()

# ==============================================================================
# 4. MÉTODO ELBOW
# ==============================================================================

logging.info("="*80)
logging.info("4. MÉTODO ELBOW - IDENTIFICAÇÃO DO Nº DE CLUSTERS")
logging.info("="*80)

try:
    elbow = []
    K = range(1, 10)
    for k in K:
        kmeans_elbow = KMeans(n_clusters=k, init='random', random_state=100, n_init=10).fit(hc_pad)
        elbow.append(kmeans_elbow.inertia_)
    
    fig_elbow = px.line(
        x=list(K), 
        y=elbow, 
        title="Método Elbow - Número de Clusters",
        markers=True,
        line_shape='linear'
    )
    fig_elbow.update_layout(
        xaxis_title="Número de Clusters (K)",
        yaxis_title="Erro Quadrático Interno (WCSS)",
        height=400,
        template='plotly_white'
    )
    grafico_elbow = fig_elbow.to_html(full_html=False, include_plotlyjs=False)
    
    logging.info("✓ SEÇÃO 4 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 4: {e}")
    import traceback
    traceback.print_exc()
    grafico_elbow = ""

# ==============================================================================
# 5. K-MEANS CLUSTERING
# ==============================================================================

logging.info("="*80)
logging.info("5. K-MEANS CLUSTERING")
logging.info("="*80)

try:
    # Aplicar K-Means com 3 clusters
    kmeans_hc = KMeans(n_clusters=3, init='random', random_state=100, n_init=10).fit(hc_pad)
    
    # Adicionar clusters aos dados
    hc_pad['cluster_kmeans'] = kmeans_hc.labels_
    dados_hc['cluster_kmeans'] = kmeans_hc.labels_
    
    # Converter para string
    dados_hc['cluster_kmeans'] = dados_hc['cluster_kmeans'].astype(str)
    hc_pad['cluster_kmeans'] = hc_pad['cluster_kmeans'].astype(str)
    
    logging.info(f"✓ K-Means aplicado com 3 clusters")
    logging.info(f"  Distribuição dos clusters: {np.bincount(kmeans_hc.labels_)}")
    
    # Centroides
    variaveis_cent = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
    cent_finais = pd.DataFrame(kmeans_hc.cluster_centers_, columns=variaveis_cent)
    cent_finais.index.name = 'Cluster'
    tabela_centroides = cent_finais.round(2).to_html(classes='tabela-estatistica', border=0)
    
    logging.info("✓ SEÇÃO 5 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 5: {e}")
    import traceback
    traceback.print_exc()
    tabela_centroides = ""

# === Centroids em z-score ===
variaveis_cent = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
centroides_z = pd.DataFrame(kmeans_hc.cluster_centers_, columns=variaveis_cent)
centroides_z.index.name = 'Cluster'

# === Centroids na escala original ===
means_orig = hc_cluster[variaveis_cent].mean()
stds_orig = hc_cluster[variaveis_cent].std(ddof=1)
centroides_orig = centroides_z.copy()
for c in variaveis_cent:
    centroides_orig[c] = centroides_z[c] * stds_orig[c] + means_orig[c]
centroides_orig.index.name = 'Cluster'

# Tabelas HTML
tabela_centroides_z = centroides_z.round(2).to_html(classes='tabela-estatistica', border=0)
tabela_centroides_orig = centroides_orig.round(2).to_html(classes='tabela-estatistica', border=0)

# === Radar (Spider) ===
eixos = variaveis_cent
fig_radar = go.Figure()
for i in range(centroides_z.shape[0]):
    r_vals = centroides_z.loc[i, eixos].tolist()
    r_vals.append(r_vals[0])
    theta = eixos + [eixos[0]]
    fig_radar.add_trace(go.Scatterpolar(
        r=r_vals,
        theta=theta,
        fill='toself',
        name=f'Cluster {i}',
        opacity=0.6
    ))

fig_radar.update_layout(
    title='Radar dos Centroids (z-score)',
    polar=dict(radialaxis=dict(visible=True, tickformat=".1f")),
    showlegend=True,
    height=500,
    template='plotly_white'
)
grafico_radar = fig_radar.to_html(full_html=False, include_plotlyjs=False)

# === GRÁFICOS 3D ===
logging.info("="*80)
logging.info("6. GRÁFICOS 3D DOS CLUSTERS")
logging.info("="*80)

try:
    cores_personalizadas = ['#440154', '#31688e', '#35b779']
    
    logging.info("Gerando gráfico 3D - 1...")
    fig_3d_1 = px.scatter_3d(
        dados_hc,
        x='idade',
        y='salario_total',
        z='dias_trabalhado',
        color='cluster_kmeans',
        color_discrete_sequence=cores_personalizadas,
        title='Clusters: Idade × Salário × Dias Trabalhado'
    )
    fig_3d_1.update_layout(height=600, showlegend=True)
    
    fig_3d_1.add_trace(go.Scatter3d(
        x=centroides_orig['idade'],
        y=centroides_orig['salario_total'],
        z=centroides_orig['dias_trabalhado'],
        mode='markers+text',
        marker=dict(size=10, color=['#440154', '#31688e', '#35b779'], symbol='diamond'),
        text=[f'C{i}' for i in centroides_orig.index],
        textposition='top center',
        name='Centroides'
    ))
    
    grafico_3d_1 = fig_3d_1.to_html(full_html=False, include_plotlyjs=False)
    logging.info("✓ Gráfico 3D - 1 gerado com sucesso")
    
    logging.info("Gerando gráfico 3D - 2...")
    fig_3d_2 = px.scatter_3d(
        dados_hc,
        x='dias_afastado',
        y='salario_total',
        z='primeiro_atestado',
        color='cluster_kmeans',
        color_discrete_sequence=cores_personalizadas,
        title='Clusters: Dias Afastado × Salário × Primeiro Atestado'
    )
    fig_3d_2.update_layout(height=600, showlegend=True)
    
    fig_3d_2.add_trace(go.Scatter3d(
        x=centroides_orig['dias_afastado'],
        y=centroides_orig['salario_total'],
        z=centroides_orig['primeiro_atestado'],
        mode='markers+text',
        marker=dict(size=10, color=['#440154', '#31688e', '#35b779'], symbol='diamond'),
        text=[f'C{i}' for i in centroides_orig.index],
        textposition='top center',
        name='Centroides'
    ))
    
    grafico_3d_2 = fig_3d_2.to_html(full_html=False, include_plotlyjs=False)
    logging.info("✓ Gráfico 3D - 2 gerado com sucesso")
    
    logging.info("Gerando gráfico 3D - 3...")
    fig_3d_3 = px.scatter_3d(
        dados_hc,
        x='dias_afastado',
        y='idade',
        z='primeiro_atestado',
        color='cluster_kmeans',
        color_discrete_sequence=cores_personalizadas,
        title='Clusters: Dias Afastado × Idade × Primeiro Atestado'
    )
    fig_3d_3.update_layout(height=600, showlegend=True)
    
    fig_3d_3.add_trace(go.Scatter3d(
        x=centroides_orig['dias_afastado'],
        y=centroides_orig['idade'],
        z=centroides_orig['primeiro_atestado'],
        mode='markers+text',
        marker=dict(size=10, color=['#440154', '#31688e', '#35b779'], symbol='diamond'),
        text=[f'C{i}' for i in centroides_orig.index],
        textposition='top center',
        name='Centroides'
    ))
    
    grafico_3d_3 = fig_3d_3.to_html(full_html=False, include_plotlyjs=False)
    logging.info("✓ Gráfico 3D - 3 gerado com sucesso")
    logging.info("✓ SEÇÃO 6 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 6: {e}")
    import traceback
    traceback.print_exc()
    grafico_3d_1 = "<p>Erro ao gerar gráfico 3D - 1</p>"
    grafico_3d_2 = "<p>Erro ao gerar gráfico 3D - 2</p>"
    grafico_3d_3 = "<p>Erro ao gerar gráfico 3D - 3</p>"

# ==============================================================================
# 7. ANÁLISE ANOVA
# ==============================================================================

logging.info("="*80)
logging.info("7. ANÁLISE ANOVA E VALIDAÇÃO DAS VARIÁVEIS")
logging.info("="*80)

try:
    variaveis_anova = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
    lista_resultados_anova = []

    for var in variaveis_anova:
        # Executa ANOVA
        resultado = pg.anova(dv=var, between='cluster_kmeans', data=hc_pad, detailed=True)
        
        # Extrai métricas
        f_value = resultado['F'].values[0]
        p_value = resultado['p-unc'].values[0]
        eta_sq = resultado['np2'].values[0] # Tamanho do efeito (Eta-squared parcial)
        
        # Verifica significância (p < 0.05)
        significativo = "✅ Sim" if p_value < 0.05 else "❌ Não"
        
        lista_resultados_anova.append({
            'Variável': var,
            'F-Score': f_value,
            'P-Value': p_value,
            'Eta² (Tamanho Efeito)': eta_sq,
            'Estat. Significativo?': significativo
        })
        
        logging.info(f"✓ ANOVA {var}: p={p_value:.4f} | {significativo}")

    # Cria DataFrame Resumo
    df_anova_resumo = pd.DataFrame(lista_resultados_anova)
    
    # Ordena pelo F-Score (quanto maior, mais a variável separa os grupos)
    df_anova_resumo = df_anova_resumo.sort_values(by='F-Score', ascending=False)

    # Gera HTML da tabela ANOVA
    # Formatando p-value e F-score para leitura amigável
    formatters = {
        'F-Score': '{:.2f}'.format,
        'P-Value': '{:.4f}'.format,
        'Eta² (Tamanho Efeito)': '{:.3f}'.format
    }
    tabela_anova = df_anova_resumo.to_html(
        classes='tabela-estatistica', 
        index=False, 
        border=0,
        formatters=formatters,
        escape=False # Permite renderizar os emojis
    )

    logging.info("✓ Tabela ANOVA gerada com sucesso.")
    logging.info("✓ SEÇÃO 7 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 7: {e}")
    import traceback
    traceback.print_exc()
    tabela_anova = "<p>Erro ao gerar tabela ANOVA</p>"

# ==============================================================================
# 8. CARACTERÍSTICAS DOS CLUSTERS
# ==============================================================================

logging.info("="*80)
logging.info("8. CARACTERÍSTICAS DOS CLUSTERS")
logging.info("="*80)

try:
    ordered_vars = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
    df_violin = dados_hc[ordered_vars + ['cluster_kmeans']].dropna().copy()
    
    # Médias por cluster
    medias_cluster = dados_hc.groupby('cluster_kmeans')[variaveis_cent].mean().round(2)
    tabela_medias = medias_cluster.to_html(classes='tabela-estatistica', border=0)
    
    # Gráfico de Salário
    medias_salario = medias_cluster[['salario_total']].reset_index()
    medias_salario.columns = ['cluster_kmeans', 'Salário']
    
    fig_salario = px.bar(
        medias_salario,
        x='cluster_kmeans',
        y='Salário',
        color='cluster_kmeans',
        color_discrete_sequence=['#440154', '#31688e', '#35b779'],
        title='Média de Salário por Cluster',
        labels={'cluster_kmeans': 'Cluster', 'Salário': 'Salário (R$)'}
    )
    fig_salario.update_layout(height=400, showlegend=False, xaxis_title='Cluster', yaxis_title='Salário (R$)')
    grafico_salario = fig_salario.to_html(full_html=False, include_plotlyjs=False)
    
    # Gráfico de Dias
    medias_dias = medias_cluster[['dias_afastado', 'primeiro_atestado', 'dias_trabalhado']].reset_index()
    medias_dias_pivot = medias_dias.set_index('cluster_kmeans')
    medias_dias_pivot.columns = ['Dias Afastado', 'Dias até 1º Atestado', 'Dias Trabalhado']
    
    fig_dias = go.Figure(data=go.Heatmap(
        z=medias_dias_pivot.values,
        x=medias_dias_pivot.columns,
        y=[f'Cluster {i}' for i in medias_dias_pivot.index],
        text=medias_dias_pivot.values.round(1),
        texttemplate='%{text:.1f}',
        colorscale='Viridis',
        hovertemplate='%{y}<br>%{x}<br>Média: %{z:.2f}<extra></extra>'
    ))
    fig_dias.update_layout(title='Médias de Dias por Cluster (Heatmap)', height=350, xaxis_title='Variável', yaxis_title='Cluster', coloraxis_colorbar=dict(title='Dias'))
    grafico_dias = fig_dias.to_html(full_html=False, include_plotlyjs=False)
    
    # Gráfico de Idade
    medias_idade = medias_cluster[['idade']].reset_index()
    medias_idade.columns = ['cluster_kmeans', 'Idade']
    
    fig_idade = px.bar(
        medias_idade,
        x='cluster_kmeans',
        y='Idade',
        color='cluster_kmeans',
        color_discrete_sequence=['#440154', '#31688e', '#35b779'],
        title='Média de Idade por Cluster',
        labels={'cluster_kmeans': 'Cluster', 'Idade': 'Idade (anos)'}
    )
    fig_idade.update_layout(height=400, showlegend=False, xaxis_title='Cluster', yaxis_title='Idade (anos)')
    grafico_idade = fig_idade.to_html(full_html=False, include_plotlyjs=False)
    
    # Gráfico de Sexo
    if 'sexo' in dados_hc.columns:
        dist_sexo = pd.crosstab(dados_hc['cluster_kmeans'], dados_hc['sexo'])
        fig_sexo = px.bar(
            dist_sexo.reset_index().melt(id_vars='cluster_kmeans', var_name='Sexo', value_name='Contagem'),
            x='cluster_kmeans',
            y='Contagem',
            color='Sexo',
            color_discrete_sequence=['#F48FB1', '#64B5F6'],
            barmode='stack',
            title='Distribuição de Gênero por Cluster',
            labels={'cluster_kmeans': 'Cluster', 'Contagem': 'Contagem'}
        )
        fig_sexo.update_layout(height=400, xaxis_title='Cluster', yaxis_title='Contagem')
        grafico_sexo = fig_sexo.to_html(full_html=False, include_plotlyjs=False)
    else:
        grafico_sexo = ""
    
    # Gráficos de Violino
    cores_violin = ['#440154', '#31688e', '#35b779']
    
    def criar_grafico_violino(df, var, titulo, label_y):
        """Cria gráfico de violino com tratamento de outliers"""
        serie = df[[var, 'cluster_kmeans']].dropna().copy()
        q1, q99 = serie[var].quantile([0.01, 0.99])
        pad = (q99 - q1) * 0.05
        y_low = float(q1 - pad)
        y_high = float(q99 + pad)
        
        if not np.isfinite(y_low) or not np.isfinite(y_high) or y_low >= y_high:
            y_low, y_high = None, None
        
        fig = px.violin(
            serie,
            x='cluster_kmeans',
            y=var,
            color='cluster_kmeans',
            box=True,
            points=False,
            color_discrete_sequence=cores_violin,
            title=titulo,
            labels={'cluster_kmeans': 'Cluster', var: label_y}
        )
        
        fig.update_yaxes(title_text=label_y, range=[y_low, y_high] if (y_low is not None and y_high is not None) else None)
        fig.update_layout(height=450, showlegend=False, template='plotly_white', xaxis_title='Cluster')
        
        return fig.to_html(full_html=False, include_plotlyjs=False)
    
    grafico_violin_dias = criar_grafico_violino(df_violin, 'dias_afastado', 'Distribuição: Dias Afastado', 'Dias')
    grafico_violin_idade = criar_grafico_violino(df_violin, 'idade', 'Distribuição: Idade', 'Idade (anos)')
    grafico_violin_primeiro = criar_grafico_violino(df_violin, 'primeiro_atestado', 'Distribuição: Dias até 1º Atestado', 'Dias')
    grafico_violin_tempo = criar_grafico_violino(df_violin, 'dias_trabalhado', 'Distribuição: Dias Trabalhado', 'Dias')
    
    var = 'salario_total'
    serie = df_violin[[var, 'cluster_kmeans']].dropna().copy()
    q1, q99 = serie[var].quantile([0.01, 0.99])
    pad = (q99 - q1) * 0.05
    y_low = float(q1 - pad)
    y_high = float(q99 + pad)
    if not np.isfinite(y_low) or not np.isfinite(y_high) or y_low >= y_high:
        y_low, y_high = None, None
    
    fig_v_sal = px.violin(
        serie,
        x='cluster_kmeans',
        y=var,
        color='cluster_kmeans',
        box=True,
        points=False,
        color_discrete_sequence=cores_violin,
        title='Distribuição: Salário',
        labels={'cluster_kmeans': 'Cluster', var: 'Salário (R$)'}
    )
    
    fig_v_sal.update_yaxes(title_text='Salário (R$)', range=[y_low, y_high] if (y_low is not None and y_high is not None) else None)
    fig_v_sal.update_layout(height=450, showlegend=False, template='plotly_white', xaxis_title='Cluster')
    grafico_violin_salario = fig_v_sal.to_html(full_html=False, include_plotlyjs=False)
    
    # Silhouette
    labels_int = kmeans_hc.labels_.astype(int)
    s_vals = silhouette_samples(hc_pad[variaveis_cent], labels_int)
    
    # CORREÇÃO AQUI: Forçar cluster_kmeans para string para garantir cores discretas
    df_sil = pd.DataFrame({
        'cluster_kmeans': labels_int.astype(str), 
        'silhouette': s_vals
    })
    
    # Ordenar para garantir consistência na legenda
    df_sil = df_sil.sort_values('cluster_kmeans')
    
    fig_sil = px.histogram(
        df_sil,
        x='silhouette',
        color='cluster_kmeans',
        nbins=30,
        barmode='overlay',
        opacity=0.6,
        color_discrete_sequence=['#440154', '#31688e', '#35b779'],
        title=f'Distribuição do Silhouette por Cluster (score médio = {silhouette_score(hc_pad[variaveis_cent], labels_int):.3f})'
    )
    fig_sil.update_layout(height=400, xaxis_title='Silhouette', yaxis_title='Frequência')
    grafico_silhouette = fig_sil.to_html(full_html=False, include_plotlyjs=False)
    
    logging.info("✓ SEÇÃO 8 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 8: {e}")
    import traceback
    traceback.print_exc()
    tabela_medias = ""
    grafico_salario = ""
    grafico_dias = ""
    grafico_idade = ""
    grafico_sexo = ""
    grafico_violin_dias = ""
    grafico_violin_idade = ""
    grafico_violin_primeiro = ""
    grafico_violin_tempo = ""
    grafico_violin_salario = ""
    grafico_silhouette = ""

# ==============================================================================
# 9. GRÁFICO DE CORRIDA
# ==============================================================================

logging.info("="*80)
logging.info("9. GRÁFICO DE CORRIDA - DIAS AFASTADO ACUMULADO")
logging.info("="*80)

try:
    if 'empresa_resumo' in dados_hc.columns and 'data_inicio_atestado' in dados_hc.columns:
        colunas_corrida = ['dias_afastado', 'empresa_resumo', 'data_inicio_atestado']
        dados_corrida = dados_hc[colunas_corrida].copy()
        
        # data_inicio_atestado já deve estar em datetime após formatação
        dados_corrida = dados_corrida.dropna(subset=['data_inicio_atestado'])
        
        if len(dados_corrida) > 0:
            todas_empresas = dados_corrida['empresa_resumo'].unique()
            todas_datas = pd.date_range(
                start=dados_corrida['data_inicio_atestado'].min(),
                end=dados_corrida['data_inicio_atestado'].max()
            )
            
            combinacoes = pd.MultiIndex.from_product(
                [todas_empresas, todas_datas],
                names=['empresa_resumo', 'data_inicio_atestado']
            ).to_frame(index=False)
            
            dados_corrida_completo = pd.merge(combinacoes, dados_corrida, on=['empresa_resumo', 'data_inicio_atestado'], how='left')
            dados_corrida_completo['dias_afastado'] = dados_corrida_completo['dias_afastado'].fillna(0)
            dados_corrida_completo = dados_corrida_completo.sort_values(by=['empresa_resumo', 'data_inicio_atestado'])
            dados_corrida_completo['Dias Acumulado'] = dados_corrida_completo.groupby('empresa_resumo')['dias_afastado'].cumsum()
            dados_corrida_completo = dados_corrida_completo.rename(columns={'empresa_resumo': 'Empresa'})
            
            fig_corrida = px.bar(
                dados_corrida_completo,
                x='Dias Acumulado',
                y='Empresa',
                orientation='h',
                color='Empresa',
                text='Dias Acumulado',
                animation_frame='data_inicio_atestado',
                animation_group='Empresa',
                title='Dias de Afastamento Acumulados por Empresa'
            )
            fig_corrida.update_layout(height=600, margin=dict(l=150, r=40, t=60, b=40))
            grafico_corrida = fig_corrida.to_html(full_html=False, include_plotlyjs=False)
        else:
            grafico_corrida = "<p>Sem dados para gráfico de corrida</p>"
    else:
        grafico_corrida = "<p>Colunas necessárias não encontradas</p>"
    
    logging.info("✓ SEÇÃO 9 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 9: {e}")
    import traceback
    traceback.print_exc()
    grafico_corrida = ""

# ==============================================================================
# 10. EXPORTAR DADOS FORMATADOS E GERAR RELATÓRIO
# ==============================================================================

logging.info("="*80)
logging.info("10. EXPORTAR DADOS E GERAR RELATÓRIO")
logging.info("="*80)

try:
    # Salvar dados em Excel
    arquivo_excel = 'Clusters_Atestados_UL.xlsx'
    
    # Usar ExcelWriter para controlar formato
    with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
        dados_hc.to_excel(writer, sheet_name='Dados', index=False)
        
        # Após escrever, formatar
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
    
    # Recarregar e formatar
    wb = load_workbook(arquivo_excel)
    ws = wb['Dados']
    
    # Dicionário de formatação
    formato_colunas = {
        'nascimento': 'mm/dd/yyyy',
        'data_nasc_conjuge': 'mm/dd/yyyy',
        'data_admissao': 'mm/dd/yyyy',
        'data_rescisao': 'mm/dd/yyyy',
        'ultimo_reajuste_individual': 'mm/dd/yyyy',
        'ultimo_reajuste_coletivo': 'mm/dd/yyyy',
        'data_inicio_ferias': 'mm/dd/yyyy',
        'data_fim_ferias': 'mm/dd/yyyy',
        'exp_45dias': 'mm/dd/yyyy',
        'exp_90dias': 'mm/dd/yyyy',
        'data_inicio_atestado': 'mm/dd/yyyy',
        'salario_admissao': '#,##0.00',
        'salario_atual': '#,##0.00',
        'salario_total': '#,##0.00',
        'saldo_fgts': '#,##0.00',
        'horas_nao_trabalhadas': '#,##0.00',
        'custo_afastamento': '#,##0.00',
        'primeiro_atestado': '0',
        'dias_trabalhado': '0',
    }
    
    # Aplicar formatação
    for col_num, col_title in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        col_letter = col_title[0].column_letter
        header_value = col_title[0].value
        
        if header_value in formato_colunas:
            fmt = formato_colunas[header_value]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.number_format = fmt
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(arquivo_excel)
    logging.info(f"✓ Arquivo Excel formatado salvo: {arquivo_excel}")
    
    # Gerar HTML
    html_final = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Análise de Clustering - Atestados Médicos (UL)</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f5f5;
            color: #333;
        }}
        
        .header {{
            background: linear-gradient(135deg, #0070C0 0%, #003d7a 100%);
            color: white;
            padding: 40px 20px;
            text-align: center;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        
        .header img {{
            max-width: 150px;
            margin-bottom: 20px;
        }}
        
        .header h1 {{
            font-size: 32px;
            margin-bottom: 10px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
        }}
        
        .tabs {{
            display: flex;
            gap: 10px;
            margin-bottom: 30px;
            border-bottom: 2px solid #ddd;
            flex-wrap: wrap;
            justify-content: space-between;
            align-items: center;
        }}
        
        .tabs-buttons {{
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }}
        
        .tab-button {{
            padding: 12px 24px;
            background-color: #f0f0f0;
            border: none;
            cursor: pointer;
            font-size: 16px;
            border-radius: 5px 5px 0 0;
            transition: all 0.3s;
            font-weight: 500;
        }}
        
        .tab-button.active {{
            background-color: #0070C0;
            color: white;
        }}
        
        .tab-button:hover {{
            background-color: #0070C0;
            color: white;
        }}
        
        .download-button {{
            padding: 12px 24px;
            background-color: #28a745;
            color: white;
            border: none;
            cursor: pointer;
            font-size: 16px;
            border-radius: 5px;
            transition: all 0.3s;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        
        .download-button:hover {{
            background-color: #218838;
            transform: scale(1.05);
        }}
        
        .download-button:active {{
            transform: scale(0.98);
        }}
        
        .btn-voltar-topo {{
            position: fixed;
            bottom: 30px;
            right: 30px;
            width: 50px;
            height: 50px;
            background-color: #0070C0;
            color: white;
            border: none;
            border-radius: 50%;
            cursor: pointer;
            font-size: 24px;
            display: none;
            align-items: center;
            justify-content: center;
            box-shadow: 0 4px 12px rgba(0, 112, 192, 0.4);
            transition: all 0.3s ease;
            z-index: 1000;
        }}
        
        .btn-voltar-topo:hover {{
            background-color: #003d7a;
            transform: translateY(-3px);
            box-shadow: 0 6px 16px rgba(0, 112, 192, 0.6);
        }}
        
        .btn-voltar-topo:active {{
            transform: translateY(-1px);
        }}
        
        .btn-voltar-topo.visible {{
            display: flex;
        }}
        
        .tab-content {{
            display: none;
            animation: fadeIn 0.3s;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; }}
            to {{ opacity: 1; }}
        }}
        
        .card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .grafico {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .tabela-estatistica {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        
        .tabela-estatistica th {{
            background-color: #0070C0;
            color: white;
            padding: 10px;
            text-align: center;
            font-weight: 600;
        }}
        
        .tabela-estatistica td {{
            padding: 8px 10px;
            border: 1px solid #ddd;
            text-align: center;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            border-top: 1px solid #ddd;
            margin-top: 40px;
        }}
        
        h2 {{
            color: #0070C0;
            margin-top: 30px;
            margin-bottom: 20px;
            border-bottom: 2px solid #0070C0;
            padding-bottom: 10px;
        }}
        
        .grid-2 {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 20px;
        }}
        
        .insight {{
            background-color: #e8f4f8;
            border-left: 4px solid #0070C0;
            padding: 15px;
            margin: 20px 0;
            border-radius: 4px;
        }}
        
        .insight strong {{
            color: #0070C0;
        }}
        
        @media (max-width: 768px) {{
            .tabs {{
                flex-direction: column;
            }}
            
            .tabs-buttons {{
                width: 100%;
            }}
            
            .download-button {{
                width: 100%;
                justify-content: center;
            }}
            
            .btn-voltar-topo {{
                bottom: 20px;
                right: 20px;
                width: 45px;
                height: 45px;
                font-size: 20px;
            }}
        }}
    </style>
</head>
<body>
    <button class="btn-voltar-topo" id="btnVoltarTopo" title="Voltar ao topo">
        ↑
    </button>
    
    <div class="header">
        <img src="{img_src_base64}" alt="Logo AFPESP">
        <h1>Análise de Clustering - Atestados Médicos (UL)</h1>
        <p>Segmentação de Colaboradores por Padrões de Afastamento</p>
    </div>
    
    <div class="container">
        <div class="tabs">
            <div class="tabs-buttons">
                <button class="tab-button active" onclick="abrirAba(event, 'exploratorio')">📊 Análise Exploratória</button>
                <button class="tab-button" onclick="abrirAba(event, 'clustering')">🎯 Clustering</button>
                <button class="tab-button" onclick="abrirAba(event, 'clusters')">📈 Características dos Clusters</button>
            </div>
            <button class="download-button" onclick="baixarExcel()">
                📥 Baixar Planilha
            </button>
        </div>
        
        <!-- ABA 1: ANÁLISE EXPLORATÓRIA -->
        <div id="exploratorio" class="tab-content active">
            <div class="card">
                <h2>Estatísticas Descritivas</h2>
                {tabela_estat}
            </div>
            
            <div class="grafico">
                <h2>Matriz de Correlação</h2>
                {grafico_corr}
            </div>
            
            <div class="grafico">
                <h2>Boxplots - Distribuição das Variáveis</h2>
                {boxplot_html}
            </div>
        </div>
        
        <!-- ABA 2: CLUSTERING -->
        <div id="clustering" class="tab-content">
            <div class="card">
                <h2>Método Elbow - Identificação do Número de Clusters</h2>
                <div class="insight">
                    <strong>💡 Insight:</strong> O método Elbow ajuda a identificar o número ótimo de clusters observando o "cotovelo" na curva de inércia.
                </div>
                <div class="grafico">{grafico_elbow}</div>
            </div>
            
            <div class="card">
                <h2>Centroides dos Clusters</h2>
                <div class="insight">
                    <strong>💡 Interpretação:</strong> Z-Score facilita comparar perfis entre variáveis. A escala original facilita ações de negócio.
                </div>
                <h3>Z-Score</h3>
                {tabela_centroides_z}
                <h3>Escala Original</h3>
                {tabela_centroides_orig}
            </div>
            
            <div class="grafico">
                <h2>Radar dos Centroids (z-score)</h2>
                {grafico_radar}
            </div>
            
            <div class="grid-2">
                <div class="grafico">{grafico_3d_1}</div>
                <div class="grafico">{grafico_3d_2}</div>
            </div>
            
            <div class="grafico">{grafico_3d_3}</div>
        </div>
        
        <!-- ABA 3: CARACTERÍSTICAS DOS CLUSTERS -->
        <div id="clusters" class="tab-content">
            <div class="card">
                <h2>Validação Estatística das Variáveis (ANOVA)</h2>
                <div class="insight">
                    <strong>📊 Interpretação:</strong> O teste ANOVA verifica se a média da variável muda significativamente entre os clusters.
                    <br>
                    • <strong>P-Value < 0.05:</strong> A variável é estatisticamente relevante para a separação dos grupos (✅).
                    <br>
                    • <strong>F-Score alto:</strong> Indica maior poder de separação entre os clusters.
                </div>
                {tabela_anova}
            </div>

            <div class="card">
                <h2>Médias das Variáveis por Cluster</h2>
                {tabela_medias}
            </div>
            
            <div class="grafico">
                <h2>Distribuição do Silhouette por Cluster</h2>
                {grafico_silhouette}
            </div>
            
            <div class="card">
                <h2>Distribuições por Cluster (Violino) — Variáveis Individuais</h2>
                <p class="insight">
                    <strong>💡 Leitura:</strong> Cada gráfico mostra a distribuição de uma variável por cluster.
                </p>
            </div>
            
            <div class="grafico">
                <h2>1. Dias Afastado</h2>
                {grafico_violin_dias}
            </div>
            
            <div class="grafico">
                <h2>2. Idade</h2>
                {grafico_violin_idade}
            </div>
            
            <div class="grafico">
                <h2>3. Dias até 1º Atestado</h2>
                {grafico_violin_primeiro}
            </div>
            
            <div class="grafico">
                <h2>4. Salário</h2>
                {grafico_violin_salario}
            </div>
            
            <div class="grafico">
                <h2>5. Dias Trabalhado</h2>
                {grafico_violin_tempo}
            </div>
            
            <div class="grafico">
                <h2>Média de Salário por Cluster</h2>
                {grafico_salario}
            </div>
            
            <div class="grafico">
                <h2>Médias de Dias por Cluster</h2>
                {grafico_dias}
            </div>
            
            <div class="grafico">
                <h2>Média de Idade por Cluster</h2>
                {grafico_idade}
            </div>
            
            {f'<div class="grafico"><h2>Distribuição de Gênero por Cluster</h2>{grafico_sexo}</div>' if grafico_sexo else ''}
            
            <div class="grafico">
                <h2>Dias de Afastamento Acumulados</h2>
                {grafico_corrida}
            </div>
            
            <div class="card">
                <h2>Recomendações Estratégicas</h2>
                <div class="insight">
                    <strong>📌 Cluster 0:</strong> Colaboradores com padrão específico de afastamento. Recomenda-se monitoramento contínuo.
                </div>
                <div class="insight">
                    <strong>📌 Cluster 1:</strong> Grupo com características distintas. Considere programas de bem-estar personalizados.
                </div>
                <div class="insight">
                    <strong>📌 Cluster 2:</strong> Padrão diferenciado. Análise individual recomendada para intervenções específicas.
                </div>
            </div>
        </div>
    </div>
    
    <div class="footer">
        <p>Relatório gerado automaticamente em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')} | Gestão de Pessoas - AFPESP</p>
    </div>
    
    <script>
        function abrirAba(evt, abaNome) {{
            var i, tabcontent, tabbuttons;
            tabcontent = document.getElementsByClassName("tab-content");
            for (i = 0; i < tabcontent.length; i++) {{
                tabcontent[i].classList.remove("active");
            }}
            tabbuttons = document.getElementsByClassName("tab-button");
            for (i = 0; i < tabbuttons.length; i++) {{
                tabbuttons[i].classList.remove("active");
            }}
            document.getElementById(abaNome).classList.add("active");
            evt.currentTarget.classList.add("active");
        }}
        
        function baixarExcel() {{
            const link = document.createElement('a');
            link.href = '{arquivo_excel}';
            link.download = '{arquivo_excel}';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            alert('Download iniciado: {arquivo_excel}');
        }}
        
        const btnVoltarTopo = document.getElementById('btnVoltarTopo');
        
        window.addEventListener('scroll', function() {{
            if (window.pageYOffset > 300) {{
                btnVoltarTopo.classList.add('visible');
            }} else {{
                btnVoltarTopo.classList.remove('visible');
            }}
        }});
        
        btnVoltarTopo.addEventListener('click', function() {{
            window.scrollTo({{
                top: 0,
                behavior: 'smooth'
            }});
        }});
    </script>
</body>
</html>
    """
    
    # Salvar HTML
    caminho_html = 'Relatório_Clustering_Atestados_UL.html'
    with open(caminho_html, 'w', encoding='utf-8') as f:
        f.write(html_final)
    
    logging.info(f"✓ Relatório HTML gerado: {caminho_html}")
    webbrowser.open('file://' + os.path.abspath(caminho_html))
    logging.info("✓ Relatório aberto no navegador.\n")
    
except Exception as e:
    logging.error(f"ERRO na SEÇÃO 10: {e}")
    import traceback
    traceback.print_exc()
    sys.exit("Script encerrado.")

logging.info("="*80)
logging.info("✓ SCRIPT CONCLUÍDO COM SUCESSO!")
logging.info("="*80)

"""
PROJETO DE CLUSTERING - ANÁLISE DE ATESTADOS MÉDICOS DA SEDE
"""
# ==============================================================================
# 1. CARREGAMENTO, PRÉ-PROCESSAMENTO E FORMATAÇÃO DE COLUNAS
# ==============================================================================

logging.info("="*80)
logging.info("1. CARREGAMENTO, PRÉ-PROCESSAMENTO E FORMATAÇÃO DE COLUNAS")
logging.info("="*80)

try:
    path_arquivo = r'X:\Gestao_de_Pessoas\Analytics\10 - Relatórios\10.4 - HC e Atestados Médicos\Controle_HC e Atestados.xlsb'
    dados_hc = pd.read_excel(path_arquivo, sheet_name='HC', engine='pyxlsb')
    
    # ========== FORMATAÇÃO DE COLUNAS ==========
    
    # Dicionário com todas as colunas a formatar
    colunas_formato = {
        # DATAS (origem Excel: 1899-12-30)
        'nascimento': 'data',
        'data_nasc_conjuge': 'data',
        'data_admissao': 'data',
        'data_rescisao': 'data',
        'ultimo_reajuste_individual': 'data',
        'ultimo_reajuste_coletivo': 'data',
        'data_inicio_ferias': 'data',
        'data_fim_ferias': 'data',
        'exp_45dias': 'data',
        'exp_90dias': 'data',
        'data_inicio_atestado': 'data',
        # NÚMEROS (valores monetários e inteiros)
        'salario_admissao': 'numero',
        'salario_atual': 'numero',
        'salario_total': 'numero',
        'saldo_fgts': 'numero',
        'horas_nao_trabalhadas': 'numero',
        'custo_afastamento': 'numero',
        'primeiro_atestado': 'numero',
        'dias_trabalhado': 'numero',
    }
    
    # Função para converter datas do Excel
    def converter_data_excel(valor):
        """Converte serial date do Excel para datetime"""
        if pd.isna(valor):
            return pd.NaT
        if isinstance(valor, (int, float)):
            # Excel serial date começa em 1899-12-30
            return pd.to_datetime(valor, unit='D', origin='1899-12-30', errors='coerce')
        return pd.to_datetime(valor, errors='coerce')
    
    # Aplicar formatação de DATAS
    for coluna in colunas_formato:
        if coluna in dados_hc.columns:
            if colunas_formato[coluna] == 'data':
                # Converter para datetime
                dados_hc[coluna] = dados_hc[coluna].apply(converter_data_excel)
                logging.info(f"✓ Coluna '{coluna}' convertida para DATA (formato: YYYY-MM-DD)")
    
    # Aplicar formatação de NÚMEROS
    for coluna in colunas_formato:
        if coluna in dados_hc.columns:
            if colunas_formato[coluna] == 'numero':
                # Converter para float, depois para int se apropriado
                dados_hc[coluna] = pd.to_numeric(dados_hc[coluna], errors='coerce')
                # Se for inteiro (dias, primeiro_atestado), converter para int
                if coluna in ['dias_trabalhado', 'primeiro_atestado']:
                    dados_hc[coluna] = dados_hc[coluna].fillna(0).astype(int)
                logging.info(f"✓ Coluna '{coluna}' convertida para NÚMERO")
    
    # Filtrar dados
    dados_hc = dados_hc.loc[(dados_hc['ano_atestado'] >= 2026) & (dados_hc['unidade'] == 'SEDE')].copy()
    logging.info(f"✓ Base de dados carregada com {dados_hc.shape[0]} registros.")
    
    # Selecionar variáveis para clustering
    hc_cluster = dados_hc[['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']].copy()
    
    # Remover valores faltantes
    hc_cluster = hc_cluster.dropna()
    logging.info(f"✓ Dados após limpeza: {hc_cluster.shape[0]} registros.")
    
    # Codificar logo
    caminho_imagem = 'Logo AFPESP.png'
    try:
        with open(caminho_imagem, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
        img_src_base64 = f"data:image/png;base64,{encoded_string}"
    except:
        img_src_base64 = ""
    
    logging.info("✓ SEÇÃO 1 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 1: {e}")
    import traceback
    traceback.print_exc()
    sys.exit("Script encerrado.")

# ==============================================================================
# 2. ANÁLISE EXPLORATÓRIA
# ==============================================================================

logging.info("="*80)
logging.info("2. ANÁLISE EXPLORATÓRIA DOS DADOS")
logging.info("="*80)

try:
    # Estatísticas descritivas
    estat_descritiva = hc_cluster.describe().round(2)
    tabela_estat = estat_descritiva.to_html(classes='tabela-estatistica', border=0)
    
    # Matriz de correlação
    corr = hc_cluster.corr()
    fig_corr = go.Figure(data=go.Heatmap(
        x=corr.columns,
        y=corr.index,
        z=np.array(corr),
        text=corr.values,
        texttemplate='%{text:.2f}',
        colorscale='Blues'
    ))
    fig_corr.update_layout(height=500, width=600, title='Matriz de Correlação')
    grafico_corr = fig_corr.to_html(full_html=False, include_plotlyjs='cdn')
    
    # Boxplots
    fig_boxplot = make_subplots(
        rows=2, cols=3,
        subplot_titles=('Dias Afastado', 'Idade', 'Primeiro Atestado', 'Salário', 'Dias Trabalhado', 'Custo Afastamento'),
        specs=[[{'type': 'box'}, {'type': 'box'}, {'type': 'box'}],
               [{'type': 'box'}, {'type': 'box'}, {'type': 'box'}]]
    )
    
    fig_boxplot.add_trace(go.Box(y=hc_cluster['dias_afastado'], name='Dias Afastado', marker_color='#0070C0'), row=1, col=1)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['idade'], name='Idade', marker_color='#0070C0'), row=1, col=2)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['primeiro_atestado'], name='Primeiro Atestado', marker_color='#0070C0'), row=1, col=3)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['salario_total'], name='Salário', marker_color='#0070C0'), row=2, col=1)
    fig_boxplot.add_trace(go.Box(y=hc_cluster['dias_trabalhado'], name='Dias Trabalhado', marker_color='#0070C0'), row=2, col=2)

    fig_boxplot.update_layout(height=600, showlegend=False, title_text="Boxplots - Variáveis Numéricas")
    boxplot_html = fig_boxplot.to_html(full_html=False, include_plotlyjs=False)
    
    logging.info("✓ SEÇÃO 2 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 2: {e}")
    import traceback
    traceback.print_exc()
    tabela_estat = ""
    grafico_corr = ""
    boxplot_html = ""

# ==============================================================================
# 3. PADRONIZAÇÃO E DETECÇÃO DE OUTLIERS
# ==============================================================================

logging.info("="*80)
logging.info("3. PADRONIZAÇÃO E DETECÇÃO DE OUTLIERS")
logging.info("="*80)

try:
    # Aplicar Z-Score
    hc_pad = hc_cluster.apply(zscore, ddof=1)
    
    # Detectar outliers
    Q1 = hc_pad.quantile(0.25)
    Q3 = hc_pad.quantile(0.75)
    IQR = Q3 - Q1
    lim_inferior = Q1 - 1.5 * IQR
    lim_superior = Q3 + 1.5 * IQR
    
    outliers = (hc_pad < lim_inferior) | (hc_pad > lim_superior)
    num_outliers = outliers.sum()
    
    logging.info(f"✓ Outliers detectados por variável:")
    for col in num_outliers.index:
        logging.info(f"  - {col}: {num_outliers[col]}")
    
    logging.info("✓ SEÇÃO 3 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 3: {e}")
    import traceback
    traceback.print_exc()

# ==============================================================================
# 4. MÉTODO ELBOW
# ==============================================================================

logging.info("="*80)
logging.info("4. MÉTODO ELBOW - IDENTIFICAÇÃO DO Nº DE CLUSTERS")
logging.info("="*80)

try:
    elbow = []
    K = range(1, 10)
    for k in K:
        kmeans_elbow = KMeans(n_clusters=k, init='random', random_state=100, n_init=10).fit(hc_pad)
        elbow.append(kmeans_elbow.inertia_)
    
    fig_elbow = px.line(
        x=list(K), 
        y=elbow, 
        title="Método Elbow - Número de Clusters",
        markers=True,
        line_shape='linear'
    )
    fig_elbow.update_layout(
        xaxis_title="Número de Clusters (K)",
        yaxis_title="Erro Quadrático Interno (WCSS)",
        height=400,
        template='plotly_white'
    )
    grafico_elbow = fig_elbow.to_html(full_html=False, include_plotlyjs=False)
    
    logging.info("✓ SEÇÃO 4 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 4: {e}")
    import traceback
    traceback.print_exc()
    grafico_elbow = ""

# ==============================================================================
# 5. K-MEANS CLUSTERING
# ==============================================================================

logging.info("="*80)
logging.info("5. K-MEANS CLUSTERING")
logging.info("="*80)

try:
    # Aplicar K-Means com 3 clusters
    kmeans_hc = KMeans(n_clusters=3, init='random', random_state=100, n_init=10).fit(hc_pad)
    
    # Adicionar clusters aos dados
    hc_pad['cluster_kmeans'] = kmeans_hc.labels_
    dados_hc['cluster_kmeans'] = kmeans_hc.labels_
    
    # Converter para string
    dados_hc['cluster_kmeans'] = dados_hc['cluster_kmeans'].astype(str)
    hc_pad['cluster_kmeans'] = hc_pad['cluster_kmeans'].astype(str)
    
    logging.info(f"✓ K-Means aplicado com 3 clusters")
    logging.info(f"  Distribuição dos clusters: {np.bincount(kmeans_hc.labels_)}")
    
    # Centroides
    variaveis_cent = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
    cent_finais = pd.DataFrame(kmeans_hc.cluster_centers_, columns=variaveis_cent)
    cent_finais.index.name = 'Cluster'
    tabela_centroides = cent_finais.round(2).to_html(classes='tabela-estatistica', border=0)
    
    logging.info("✓ SEÇÃO 5 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 5: {e}")
    import traceback
    traceback.print_exc()
    tabela_centroides = ""

# === Centroids em z-score ===
variaveis_cent = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
centroides_z = pd.DataFrame(kmeans_hc.cluster_centers_, columns=variaveis_cent)
centroides_z.index.name = 'Cluster'

# === Centroids na escala original ===
means_orig = hc_cluster[variaveis_cent].mean()
stds_orig = hc_cluster[variaveis_cent].std(ddof=1)
centroides_orig = centroides_z.copy()
for c in variaveis_cent:
    centroides_orig[c] = centroides_z[c] * stds_orig[c] + means_orig[c]
centroides_orig.index.name = 'Cluster'

# Tabelas HTML
tabela_centroides_z = centroides_z.round(2).to_html(classes='tabela-estatistica', border=0)
tabela_centroides_orig = centroides_orig.round(2).to_html(classes='tabela-estatistica', border=0)

# === Radar (Spider) ===
eixos = variaveis_cent
fig_radar = go.Figure()
for i in range(centroides_z.shape[0]):
    r_vals = centroides_z.loc[i, eixos].tolist()
    r_vals.append(r_vals[0])
    theta = eixos + [eixos[0]]
    fig_radar.add_trace(go.Scatterpolar(
        r=r_vals,
        theta=theta,
        fill='toself',
        name=f'Cluster {i}',
        opacity=0.6
    ))

fig_radar.update_layout(
    title='Radar dos Centroids (z-score)',
    polar=dict(radialaxis=dict(visible=True, tickformat=".1f")),
    showlegend=True,
    height=500,
    template='plotly_white'
)
grafico_radar = fig_radar.to_html(full_html=False, include_plotlyjs=False)

# === GRÁFICOS 3D ===
logging.info("="*80)
logging.info("6. GRÁFICOS 3D DOS CLUSTERS")
logging.info("="*80)

try:
    cores_personalizadas = ['#440154', '#31688e', '#35b779']
    
    logging.info("Gerando gráfico 3D - 1...")
    fig_3d_1 = px.scatter_3d(
        dados_hc,
        x='idade',
        y='salario_total',
        z='dias_trabalhado',
        color='cluster_kmeans',
        color_discrete_sequence=cores_personalizadas,
        title='Clusters: Idade × Salário × Dias Trabalhado'
    )
    fig_3d_1.update_layout(height=600, showlegend=True)
    
    fig_3d_1.add_trace(go.Scatter3d(
        x=centroides_orig['idade'],
        y=centroides_orig['salario_total'],
        z=centroides_orig['dias_trabalhado'],
        mode='markers+text',
        marker=dict(size=10, color=['#440154', '#31688e', '#35b779'], symbol='diamond'),
        text=[f'C{i}' for i in centroides_orig.index],
        textposition='top center',
        name='Centroides'
    ))
    
    grafico_3d_1 = fig_3d_1.to_html(full_html=False, include_plotlyjs=False)
    logging.info("✓ Gráfico 3D - 1 gerado com sucesso")
    
    logging.info("Gerando gráfico 3D - 2...")
    fig_3d_2 = px.scatter_3d(
        dados_hc,
        x='dias_afastado',
        y='salario_total',
        z='primeiro_atestado',
        color='cluster_kmeans',
        color_discrete_sequence=cores_personalizadas,
        title='Clusters: Dias Afastado × Salário × Primeiro Atestado'
    )
    fig_3d_2.update_layout(height=600, showlegend=True)
    
    fig_3d_2.add_trace(go.Scatter3d(
        x=centroides_orig['dias_afastado'],
        y=centroides_orig['salario_total'],
        z=centroides_orig['primeiro_atestado'],
        mode='markers+text',
        marker=dict(size=10, color=['#440154', '#31688e', '#35b779'], symbol='diamond'),
        text=[f'C{i}' for i in centroides_orig.index],
        textposition='top center',
        name='Centroides'
    ))
    
    grafico_3d_2 = fig_3d_2.to_html(full_html=False, include_plotlyjs=False)
    logging.info("✓ Gráfico 3D - 2 gerado com sucesso")
    
    logging.info("Gerando gráfico 3D - 3...")
    fig_3d_3 = px.scatter_3d(
        dados_hc,
        x='dias_afastado',
        y='idade',
        z='primeiro_atestado',
        color='cluster_kmeans',
        color_discrete_sequence=cores_personalizadas,
        title='Clusters: Dias Afastado × Idade × Primeiro Atestado'
    )
    fig_3d_3.update_layout(height=600, showlegend=True)
    
    fig_3d_3.add_trace(go.Scatter3d(
        x=centroides_orig['dias_afastado'],
        y=centroides_orig['idade'],
        z=centroides_orig['primeiro_atestado'],
        mode='markers+text',
        marker=dict(size=10, color=['#440154', '#31688e', '#35b779'], symbol='diamond'),
        text=[f'C{i}' for i in centroides_orig.index],
        textposition='top center',
        name='Centroides'
    ))
    
    grafico_3d_3 = fig_3d_3.to_html(full_html=False, include_plotlyjs=False)
    logging.info("✓ Gráfico 3D - 3 gerado com sucesso")
    logging.info("✓ SEÇÃO 6 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 6: {e}")
    import traceback
    traceback.print_exc()
    grafico_3d_1 = "<p>Erro ao gerar gráfico 3D - 1</p>"
    grafico_3d_2 = "<p>Erro ao gerar gráfico 3D - 2</p>"
    grafico_3d_3 = "<p>Erro ao gerar gráfico 3D - 3</p>"

# ==============================================================================
# 7. ANÁLISE ANOVA
# ==============================================================================

logging.info("="*80)
logging.info("7. ANÁLISE ANOVA E VALIDAÇÃO DAS VARIÁVEIS")
logging.info("="*80)

try:
    variaveis_anova = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
    lista_resultados_anova = []

    for var in variaveis_anova:
        # Executa ANOVA
        resultado = pg.anova(dv=var, between='cluster_kmeans', data=hc_pad, detailed=True)
        
        # Extrai métricas
        f_value = resultado['F'].values[0]
        p_value = resultado['p-unc'].values[0]
        eta_sq = resultado['np2'].values[0] # Tamanho do efeito (Eta-squared parcial)
        
        # Verifica significância (p < 0.05)
        significativo = "✅ Sim" if p_value < 0.05 else "❌ Não"
        
        lista_resultados_anova.append({
            'Variável': var,
            'F-Score': f_value,
            'P-Value': p_value,
            'Eta² (Tamanho Efeito)': eta_sq,
            'Estat. Significativo?': significativo
        })
        
        logging.info(f"✓ ANOVA {var}: p={p_value:.4f} | {significativo}")

    # Cria DataFrame Resumo
    df_anova_resumo = pd.DataFrame(lista_resultados_anova)
    
    # Ordena pelo F-Score (quanto maior, mais a variável separa os grupos)
    df_anova_resumo = df_anova_resumo.sort_values(by='F-Score', ascending=False)

    # Gera HTML da tabela ANOVA
    # Formatando p-value e F-score para leitura amigável
    formatters = {
        'F-Score': '{:.2f}'.format,
        'P-Value': '{:.4f}'.format,
        'Eta² (Tamanho Efeito)': '{:.3f}'.format
    }
    tabela_anova = df_anova_resumo.to_html(
        classes='tabela-estatistica', 
        index=False, 
        border=0,
        formatters=formatters,
        escape=False # Permite renderizar os emojis
    )

    logging.info("✓ Tabela ANOVA gerada com sucesso.")
    logging.info("✓ SEÇÃO 7 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 7: {e}")
    import traceback
    traceback.print_exc()
    tabela_anova = "<p>Erro ao gerar tabela ANOVA</p>"

# ==============================================================================
# 8. CARACTERÍSTICAS DOS CLUSTERS
# ==============================================================================

logging.info("="*80)
logging.info("8. CARACTERÍSTICAS DOS CLUSTERS")
logging.info("="*80)

try:
    ordered_vars = ['dias_afastado', 'idade', 'primeiro_atestado', 'salario_total', 'dias_trabalhado']
    df_violin = dados_hc[ordered_vars + ['cluster_kmeans']].dropna().copy()
    
    # Médias por cluster
    medias_cluster = dados_hc.groupby('cluster_kmeans')[variaveis_cent].mean().round(3)
    tabela_medias = medias_cluster.to_html(classes='tabela-estatistica', border=0)
    
    # Gráfico de Salário
    medias_salario = medias_cluster[['salario_total']].reset_index()
    medias_salario.columns = ['cluster_kmeans', 'Salário']
    
    fig_salario = px.bar(
        medias_salario,
        x='cluster_kmeans',
        y='Salário',
        color='cluster_kmeans',
        color_discrete_sequence=['#440154', '#31688e', '#35b779'],
        title='Média de Salário por Cluster',
        labels={'cluster_kmeans': 'Cluster', 'Salário': 'Salário (R$)'}
    )
    fig_salario.update_layout(height=400, showlegend=False, xaxis_title='Cluster', yaxis_title='Salário (R$)')
    grafico_salario = fig_salario.to_html(full_html=False, include_plotlyjs=False)
    
    # Gráfico de Dias
    medias_dias = medias_cluster[['dias_afastado', 'primeiro_atestado', 'dias_trabalhado']].reset_index()
    medias_dias_pivot = medias_dias.set_index('cluster_kmeans')
    medias_dias_pivot.columns = ['Dias Afastado', 'Dias até 1º Atestado', 'Dias Trabalhado']
    
    fig_dias = go.Figure(data=go.Heatmap(
        z=medias_dias_pivot.values,
        x=medias_dias_pivot.columns,
        y=[f'Cluster {i}' for i in medias_dias_pivot.index],
        text=medias_dias_pivot.values.round(1),
        texttemplate='%{text:.1f}',
        colorscale='Viridis',
        hovertemplate='%{y}<br>%{x}<br>Média: %{z:.2f}<extra></extra>'
    ))
    fig_dias.update_layout(title='Médias de Dias por Cluster (Heatmap)', height=350, xaxis_title='Variável', yaxis_title='Cluster', coloraxis_colorbar=dict(title='Dias'))
    grafico_dias = fig_dias.to_html(full_html=False, include_plotlyjs=False)
    
    # Gráfico de Idade
    medias_idade = medias_cluster[['idade']].reset_index()
    medias_idade.columns = ['cluster_kmeans', 'Idade']
    
    fig_idade = px.bar(
        medias_idade,
        x='cluster_kmeans',
        y='Idade',
        color='cluster_kmeans',
        color_discrete_sequence=['#440154', '#31688e', '#35b779'],
        title='Média de Idade por Cluster',
        labels={'cluster_kmeans': 'Cluster', 'Idade': 'Idade (anos)'}
    )
    fig_idade.update_layout(height=400, showlegend=False, xaxis_title='Cluster', yaxis_title='Idade (anos)')
    grafico_idade = fig_idade.to_html(full_html=False, include_plotlyjs=False)
    
    # Gráfico de Sexo
    if 'sexo' in dados_hc.columns:
        dist_sexo = pd.crosstab(dados_hc['cluster_kmeans'], dados_hc['sexo'])
        fig_sexo = px.bar(
            dist_sexo.reset_index().melt(id_vars='cluster_kmeans', var_name='Sexo', value_name='Contagem'),
            x='cluster_kmeans',
            y='Contagem',
            color='Sexo',
            color_discrete_sequence=['#F48FB1', '#64B5F6'],
            barmode='stack',
            title='Distribuição de Gênero por Cluster',
            labels={'cluster_kmeans': 'Cluster', 'Contagem': 'Contagem'}
        )
        fig_sexo.update_layout(height=400, xaxis_title='Cluster', yaxis_title='Contagem')
        grafico_sexo = fig_sexo.to_html(full_html=False, include_plotlyjs=False)
    else:
        grafico_sexo = ""
    
    # Gráficos de Violino
    cores_violin = ['#440154', '#31688e', '#35b779']
    
    def criar_grafico_violino(df, var, titulo, label_y):
        """Cria gráfico de violino com tratamento de outliers"""
        serie = df[[var, 'cluster_kmeans']].dropna().copy()
        q1, q99 = serie[var].quantile([0.01, 0.99])
        pad = (q99 - q1) * 0.05
        y_low = float(q1 - pad)
        y_high = float(q99 + pad)
        
        if not np.isfinite(y_low) or not np.isfinite(y_high) or y_low >= y_high:
            y_low, y_high = None, None
        
        fig = px.violin(
            serie,
            x='cluster_kmeans',
            y=var,
            color='cluster_kmeans',
            box=True,
            points=False,
            color_discrete_sequence=cores_violin,
            title=titulo,
            labels={'cluster_kmeans': 'Cluster', var: label_y}
        )
        
        fig.update_yaxes(title_text=label_y, range=[y_low, y_high] if (y_low is not None and y_high is not None) else None)
        fig.update_layout(height=450, showlegend=False, template='plotly_white', xaxis_title='Cluster')
        
        return fig.to_html(full_html=False, include_plotlyjs=False)
    
    grafico_violin_dias = criar_grafico_violino(df_violin, 'dias_afastado', 'Distribuição: Dias Afastado', 'Dias')
    grafico_violin_idade = criar_grafico_violino(df_violin, 'idade', 'Distribuição: Idade', 'Idade (anos)')
    grafico_violin_primeiro = criar_grafico_violino(df_violin, 'primeiro_atestado', 'Distribuição: Dias até 1º Atestado', 'Dias')
    grafico_violin_tempo = criar_grafico_violino(df_violin, 'dias_trabalhado', 'Distribuição: Dias Trabalhado', 'Dias')
    
    var = 'salario_total'
    serie = df_violin[[var, 'cluster_kmeans']].dropna().copy()
    q1, q99 = serie[var].quantile([0.01, 0.99])
    pad = (q99 - q1) * 0.05
    y_low = float(q1 - pad)
    y_high = float(q99 + pad)
    if not np.isfinite(y_low) or not np.isfinite(y_high) or y_low >= y_high:
        y_low, y_high = None, None
    
    fig_v_sal = px.violin(
        serie,
        x='cluster_kmeans',
        y=var,
        color='cluster_kmeans',
        box=True,
        points=False,
        color_discrete_sequence=cores_violin,
        title='Distribuição: Salário',
        labels={'cluster_kmeans': 'Cluster', var: 'Salário (R$)'}
    )
    
    fig_v_sal.update_yaxes(title_text='Salário (R$)', range=[y_low, y_high] if (y_low is not None and y_high is not None) else None)
    fig_v_sal.update_layout(height=450, showlegend=False, template='plotly_white', xaxis_title='Cluster')
    grafico_violin_salario = fig_v_sal.to_html(full_html=False, include_plotlyjs=False)
    
    # Silhouette
    labels_int = kmeans_hc.labels_.astype(int)
    s_vals = silhouette_samples(hc_pad[variaveis_cent], labels_int)
    
    # CORREÇÃO AQUI: Forçar cluster_kmeans para string para garantir cores discretas
    df_sil = pd.DataFrame({
        'cluster_kmeans': labels_int.astype(str), 
        'silhouette': s_vals
    })
    
    # Ordenar para garantir consistência na legenda
    df_sil = df_sil.sort_values('cluster_kmeans')
    
    fig_sil = px.histogram(
        df_sil,
        x='silhouette',
        color='cluster_kmeans',
        nbins=30,
        barmode='overlay',
        opacity=0.6,
        color_discrete_sequence=['#440154', '#31688e', '#35b779'],
        title=f'Distribuição do Silhouette por Cluster (score médio = {silhouette_score(hc_pad[variaveis_cent], labels_int):.3f})'
    )
    fig_sil.update_layout(height=400, xaxis_title='Silhouette', yaxis_title='Frequência')
    grafico_silhouette = fig_sil.to_html(full_html=False, include_plotlyjs=False)
    
    logging.info("✓ SEÇÃO 8 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 8: {e}")
    import traceback
    traceback.print_exc()
    tabela_medias = ""
    grafico_salario = ""
    grafico_dias = ""
    grafico_idade = ""
    grafico_sexo = ""
    grafico_violin_dias = ""
    grafico_violin_idade = ""
    grafico_violin_primeiro = ""
    grafico_violin_tempo = ""
    grafico_violin_salario = ""
    grafico_silhouette = ""

# ==============================================================================
# 9. GRÁFICO DE CORRIDA
# ==============================================================================

logging.info("="*80)
logging.info("9. GRÁFICO DE CORRIDA - DIAS AFASTADO ACUMULADO")
logging.info("="*80)

try:
    if 'empresa_resumo' in dados_hc.columns and 'data_inicio_atestado' in dados_hc.columns:
        colunas_corrida = ['dias_afastado', 'empresa_resumo', 'data_inicio_atestado']
        dados_corrida = dados_hc[colunas_corrida].copy()
        
        # data_inicio_atestado já deve estar em datetime após formatação
        dados_corrida = dados_corrida.dropna(subset=['data_inicio_atestado'])
        
        if len(dados_corrida) > 0:
            todas_empresas = dados_corrida['empresa_resumo'].unique()
            todas_datas = pd.date_range(
                start=dados_corrida['data_inicio_atestado'].min(),
                end=dados_corrida['data_inicio_atestado'].max()
            )
            
            combinacoes = pd.MultiIndex.from_product(
                [todas_empresas, todas_datas],
                names=['empresa_resumo', 'data_inicio_atestado']
            ).to_frame(index=False)
            
            dados_corrida_completo = pd.merge(combinacoes, dados_corrida, on=['empresa_resumo', 'data_inicio_atestado'], how='left')
            dados_corrida_completo['dias_afastado'] = dados_corrida_completo['dias_afastado'].fillna(0)
            dados_corrida_completo = dados_corrida_completo.sort_values(by=['empresa_resumo', 'data_inicio_atestado'])
            dados_corrida_completo['Dias Acumulado'] = dados_corrida_completo.groupby('empresa_resumo')['dias_afastado'].cumsum()
            dados_corrida_completo = dados_corrida_completo.rename(columns={'empresa_resumo': 'Empresa'})
            
            fig_corrida = px.bar(
                dados_corrida_completo,
                x='Dias Acumulado',
                y='Empresa',
                orientation='h',
                color='Empresa',
                text='Dias Acumulado',
                animation_frame='data_inicio_atestado',
                animation_group='Empresa',
                title='Dias de Afastamento Acumulados por Empresa'
            )
            fig_corrida.update_layout(height=600, margin=dict(l=150, r=40, t=60, b=40))
            grafico_corrida = fig_corrida.to_html(full_html=False, include_plotlyjs=False)
        else:
            grafico_corrida = "<p>Sem dados para gráfico de corrida</p>"
    else:
        grafico_corrida = "<p>Colunas necessárias não encontradas</p>"
    
    logging.info("✓ SEÇÃO 9 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 9: {e}")
    import traceback
    traceback.print_exc()
    grafico_corrida = ""

# ==============================================================================
# 10. EXPORTAR DADOS FORMATADOS E GERAR RELATÓRIO
# ==============================================================================

logging.info("="*80)
logging.info("10. EXPORTAR DADOS E GERAR RELATÓRIO")
logging.info("="*80)

try:
    # Salvar dados em Excel
    arquivo_excel = 'Clusters_Atestados_SEDE.xlsx'
    
    # Usar ExcelWriter para controlar formato
    with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
        dados_hc.to_excel(writer, sheet_name='Dados', index=False)
        
        # Após escrever, formatar
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
    
    # Recarregar e formatar
    wb = load_workbook(arquivo_excel)
    ws = wb['Dados']
    
    # Dicionário de formatação
    formato_colunas = {
        'nascimento': 'mm/dd/yyyy',
        'data_nasc_conjuge': 'mm/dd/yyyy',
        'data_admissao': 'mm/dd/yyyy',
        'data_rescisao': 'mm/dd/yyyy',
        'ultimo_reajuste_individual': 'mm/dd/yyyy',
        'ultimo_reajuste_coletivo': 'mm/dd/yyyy',
        'data_inicio_ferias': 'mm/dd/yyyy',
        'data_fim_ferias': 'mm/dd/yyyy',
        'exp_45dias': 'mm/dd/yyyy',
        'exp_90dias': 'mm/dd/yyyy',
        'data_inicio_atestado': 'mm/dd/yyyy',
        'salario_admissao': '#,##0.00',
        'salario_atual': '#,##0.00',
        'salario_total': '#,##0.00',
        'saldo_fgts': '#,##0.00',
        'horas_nao_trabalhadas': '#,##0.00',
        'custo_afastamento': '#,##0.00',
        'primeiro_atestado': '0',
        'dias_trabalhado': '0',
    }
    
    # Aplicar formatação
    for col_num, col_title in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        col_letter = col_title[0].column_letter
        header_value = col_title[0].value
        
        if header_value in formato_colunas:
            fmt = formato_colunas[header_value]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.number_format = fmt
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(arquivo_excel)
    logging.info(f"✓ Arquivo Excel formatado salvo: {arquivo_excel}")
    
    # Gerar HTML
    html_final = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Análise de Clustering - Atestados Médicos (SEDE)</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f5f5;
            color: #333;
        }}
        
        .header {{
            background: linear-gradient(135deg, #0070C0 0%, #003d7a 100%);
            color: white;
            padding: 40px 20px;
            text-align: center;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        
        .header img {{
            max-width: 150px;
            margin-bottom: 20px;
        }}
        
        .header h1 {{
            font-size: 32px;
            margin-bottom: 10px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
        }}
        
        .tabs {{
            display: flex;
            gap: 10px;
            margin-bottom: 30px;
            border-bottom: 2px solid #ddd;
            flex-wrap: wrap;
            justify-content: space-between;
            align-items: center;
        }}
        
        .tabs-buttons {{
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }}
        
        .tab-button {{
            padding: 12px 24px;
            background-color: #f0f0f0;
            border: none;
            cursor: pointer;
            font-size: 16px;
            border-radius: 5px 5px 0 0;
            transition: all 0.3s;
            font-weight: 500;
        }}
        
        .tab-button.active {{
            background-color: #0070C0;
            color: white;
        }}
        
        .tab-button:hover {{
            background-color: #0070C0;
            color: white;
        }}
        
        .download-button {{
            padding: 12px 24px;
            background-color: #28a745;
            color: white;
            border: none;
            cursor: pointer;
            font-size: 16px;
            border-radius: 5px;
            transition: all 0.3s;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        
        .download-button:hover {{
            background-color: #218838;
            transform: scale(1.05);
        }}
        
        .download-button:active {{
            transform: scale(0.98);
        }}
        
        .btn-voltar-topo {{
            position: fixed;
            bottom: 30px;
            right: 30px;
            width: 50px;
            height: 50px;
            background-color: #0070C0;
            color: white;
            border: none;
            border-radius: 50%;
            cursor: pointer;
            font-size: 24px;
            display: none;
            align-items: center;
            justify-content: center;
            box-shadow: 0 4px 12px rgba(0, 112, 192, 0.4);
            transition: all 0.3s ease;
            z-index: 1000;
        }}
        
        .btn-voltar-topo:hover {{
            background-color: #003d7a;
            transform: translateY(-3px);
            box-shadow: 0 6px 16px rgba(0, 112, 192, 0.6);
        }}
        
        .btn-voltar-topo:active {{
            transform: translateY(-1px);
        }}
        
        .btn-voltar-topo.visible {{
            display: flex;
        }}
        
        .tab-content {{
            display: none;
            animation: fadeIn 0.3s;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; }}
            to {{ opacity: 1; }}
        }}
        
        .card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .grafico {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .tabela-estatistica {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        
        .tabela-estatistica th {{
            background-color: #0070C0;
            color: white;
            padding: 10px;
            text-align: center;
            font-weight: 600;
        }}
        
        .tabela-estatistica td {{
            padding: 8px 10px;
            border: 1px solid #ddd;
            text-align: center;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            border-top: 1px solid #ddd;
            margin-top: 40px;
        }}
        
        h2 {{
            color: #0070C0;
            margin-top: 30px;
            margin-bottom: 20px;
            border-bottom: 2px solid #0070C0;
            padding-bottom: 10px;
        }}
        
        .grid-2 {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 20px;
        }}
        
        .insight {{
            background-color: #e8f4f8;
            border-left: 4px solid #0070C0;
            padding: 15px;
            margin: 20px 0;
            border-radius: 4px;
        }}
        
        .insight strong {{
            color: #0070C0;
        }}
        
        @media (max-width: 768px) {{
            .tabs {{
                flex-direction: column;
            }}
            
            .tabs-buttons {{
                width: 100%;
            }}
            
            .download-button {{
                width: 100%;
                justify-content: center;
            }}
            
            .btn-voltar-topo {{
                bottom: 20px;
                right: 20px;
                width: 45px;
                height: 45px;
                font-size: 20px;
            }}
        }}
    </style>
</head>
<body>
    <button class="btn-voltar-topo" id="btnVoltarTopo" title="Voltar ao topo">
        ↑
    </button>
    
    <div class="header">
        <img src="{img_src_base64}" alt="Logo AFPESP">
        <h1>Análise de Clustering - Atestados Médicos (SEDE)</h1>
        <p>Segmentação de Colaboradores por Padrões de Afastamento</p>
    </div>
    
    <div class="container">
        <div class="tabs">
            <div class="tabs-buttons">
                <button class="tab-button active" onclick="abrirAba(event, 'exploratorio')">📊 Análise Exploratória</button>
                <button class="tab-button" onclick="abrirAba(event, 'clustering')">🎯 Clustering</button>
                <button class="tab-button" onclick="abrirAba(event, 'clusters')">📈 Características dos Clusters</button>
            </div>
            <button class="download-button" onclick="baixarExcel()">
                📥 Baixar Planilha
            </button>
        </div>
        
        <!-- ABA 1: ANÁLISE EXPLORATÓRIA -->
        <div id="exploratorio" class="tab-content active">
            <div class="card">
                <h2>Estatísticas Descritivas</h2>
                {tabela_estat}
            </div>
            
            <div class="grafico">
                <h2>Matriz de Correlação</h2>
                {grafico_corr}
            </div>
            
            <div class="grafico">
                <h2>Boxplots - Distribuição das Variáveis</h2>
                {boxplot_html}
            </div>
        </div>
        
        <!-- ABA 2: CLUSTERING -->
        <div id="clustering" class="tab-content">
            <div class="card">
                <h2>Método Elbow - Identificação do Número de Clusters</h2>
                <div class="insight">
                    <strong>💡 Insight:</strong> O método Elbow ajuda a identificar o número ótimo de clusters observando o "cotovelo" na curva de inércia.
                </div>
                <div class="grafico">{grafico_elbow}</div>
            </div>
            
            <div class="card">
                <h2>Centroides dos Clusters</h2>
                <div class="insight">
                    <strong>💡 Interpretação:</strong> Z-Score facilita comparar perfis entre variáveis. A escala original facilita ações de negócio.
                </div>
                <h3>Z-Score</h3>
                {tabela_centroides_z}
                <h3>Escala Original</h3>
                {tabela_centroides_orig}
            </div>
            
            <div class="grafico">
                <h2>Radar dos Centroids (z-score)</h2>
                {grafico_radar}
            </div>
            
            <div class="grid-2">
                <div class="grafico">{grafico_3d_1}</div>
                <div class="grafico">{grafico_3d_2}</div>
            </div>
            
            <div class="grafico">{grafico_3d_3}</div>
        </div>
        
        <!-- ABA 3: CARACTERÍSTICAS DOS CLUSTERS -->
        <div id="clusters" class="tab-content">
            <div class="card">
                <h2>Validação Estatística das Variáveis (ANOVA)</h2>
                <div class="insight">
                    <strong>📊 Interpretação:</strong> O teste ANOVA verifica se a média da variável muda significativamente entre os clusters.
                    <br>
                    • <strong>P-Value < 0.05:</strong> A variável é estatisticamente relevante para a separação dos grupos (✅).
                    <br>
                    • <strong>F-Score alto:</strong> Indica maior poder de separação entre os clusters.
                </div>
                {tabela_anova}
            </div>

            <div class="card">
                <h2>Médias das Variáveis por Cluster</h2>
                {tabela_medias}
            </div>
            
            <div class="grafico">
                <h2>Distribuição do Silhouette por Cluster</h2>
                {grafico_silhouette}
            </div>
            
            <div class="card">
                <h2>Distribuições por Cluster (Violino) — Variáveis Individuais</h2>
                <p class="insight">
                    <strong>💡 Leitura:</strong> Cada gráfico mostra a distribuição de uma variável por cluster.
                </p>
            </div>
            
            <div class="grafico">
                <h2>1. Dias Afastado</h2>
                {grafico_violin_dias}
            </div>
            
            <div class="grafico">
                <h2>2. Idade</h2>
                {grafico_violin_idade}
            </div>
            
            <div class="grafico">
                <h2>3. Dias até 1º Atestado</h2>
                {grafico_violin_primeiro}
            </div>
            
            <div class="grafico">
                <h2>4. Salário</h2>
                {grafico_violin_salario}
            </div>
            
            <div class="grafico">
                <h2>5. Dias Trabalhado</h2>
                {grafico_violin_tempo}
            </div>
            
            <div class="grafico">
                <h2>Média de Salário por Cluster</h2>
                {grafico_salario}
            </div>
            
            <div class="grafico">
                <h2>Médias de Dias por Cluster</h2>
                {grafico_dias}
            </div>
            
            <div class="grafico">
                <h2>Média de Idade por Cluster</h2>
                {grafico_idade}
            </div>
            
            {f'<div class="grafico"><h2>Distribuição de Gênero por Cluster</h2>{grafico_sexo}</div>' if grafico_sexo else ''}
            
            <div class="grafico">
                <h2>Dias de Afastamento Acumulados</h2>
                {grafico_corrida}
            </div>
            
            <div class="card">
                <h2>Recomendações Estratégicas</h2>
                <div class="insight">
                    <strong>📌 Cluster 0:</strong> Colaboradores com padrão específico de afastamento. Recomenda-se monitoramento contínuo.
                </div>
                <div class="insight">
                    <strong>📌 Cluster 1:</strong> Grupo com características distintas. Considere programas de bem-estar personalizados.
                </div>
                <div class="insight">
                    <strong>📌 Cluster 2:</strong> Padrão diferenciado. Análise individual recomendada para intervenções específicas.
                </div>
            </div>
        </div>
    </div>
    
    <div class="footer">
        <p>Relatório gerado automaticamente em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')} | Gestão de Pessoas - AFPESP</p>
    </div>
    
    <script>
        function abrirAba(evt, abaNome) {{
            var i, tabcontent, tabbuttons;
            tabcontent = document.getElementsByClassName("tab-content");
            for (i = 0; i < tabcontent.length; i++) {{
                tabcontent[i].classList.remove("active");
            }}
            tabbuttons = document.getElementsByClassName("tab-button");
            for (i = 0; i < tabbuttons.length; i++) {{
                tabbuttons[i].classList.remove("active");
            }}
            document.getElementById(abaNome).classList.add("active");
            evt.currentTarget.classList.add("active");
        }}
        
        function baixarExcel() {{
            const link = document.createElement('a');
            link.href = '{arquivo_excel}';
            link.download = '{arquivo_excel}';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            alert('Download iniciado: {arquivo_excel}');
        }}
        
        const btnVoltarTopo = document.getElementById('btnVoltarTopo');
        
        window.addEventListener('scroll', function() {{
            if (window.pageYOffset > 300) {{
                btnVoltarTopo.classList.add('visible');
            }} else {{
                btnVoltarTopo.classList.remove('visible');
            }}
        }});
        
        btnVoltarTopo.addEventListener('click', function() {{
            window.scrollTo({{
                top: 0,
                behavior: 'smooth'
            }});
        }});
    </script>
</body>
</html>
    """
    
    # Salvar HTML
    caminho_html = 'Relatório_Clustering_Atestados_SEDE.html'
    with open(caminho_html, 'w', encoding='utf-8') as f:
        f.write(html_final)
    
    logging.info(f"✓ Relatório HTML gerado: {caminho_html}")
    webbrowser.open('file://' + os.path.abspath(caminho_html))
    logging.info("✓ Relatório aberto no navegador.\n")
    
except Exception as e:
    logging.error(f"ERRO na SEÇÃO 10: {e}")
    import traceback
    traceback.print_exc()
    sys.exit("Script encerrado.")

logging.info("="*80)
logging.info("✓ SCRIPT CONCLUÍDO COM SUCESSO!")
logging.info("="*80)