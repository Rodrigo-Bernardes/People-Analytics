import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import base64
import os
import webbrowser
import logging
import sys
import gc
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
import traceback

# Carregamento da base de controle de processos

id = 22

path_registros_processos = r'X:\Gestao_de_Pessoas\Analytics\03 - Bases\1. BASES TRATADAS\PROCESSOS.xlsx'

registros_processos = pd.read_excel(path_registros_processos, sheet_name = "REGISTROS", engine='openpyxl')

wb_p = load_workbook(path_registros_processos)

ws_p = wb_p['REGISTROS']

# Controle de atualização de processo: Etapa 1

linha_0 = [id, datetime.today(), 1]

ws_p.append(linha_0)

wb_p.save(path_registros_processos)

# ==============================================================================
# Configurar logging
# ==============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

print("✓ Bibliotecas importadas com sucesso.\n")

# ==============================================================================
# FUNÇÃO DE LIMPEZA DE DADOS
# ==============================================================================

def limpar_dados_completo(df):
    """Limpa valores inválidos do DataFrame"""
    for coluna in df.columns:
        try:
            if df[coluna].dtype == 'object':
                df[coluna] = df[coluna].apply(lambda x: 
                    np.nan if pd.isna(x) or (isinstance(x, str) and x.strip().upper() in ['NAN', 'NONE', '', 'NULL', 'N/A']) 
                    else x
                )
            elif df[coluna].dtype in ['float64', 'int64']:
                df[coluna] = pd.to_numeric(df[coluna], errors='coerce')
        except Exception as e:
            continue
    return df

# ==============================================================================
# FUNÇÃO UNIFICADA DE FORMATAÇÃO EXCEL
# ==============================================================================

def formatar_excel_profissional(caminho_arquivo, cor_cabecalho='005A9C', cor_fonte_cabecalho='FFFFFF'):
    """
    Formata arquivo Excel com cabeçalho azul e regras específicas de colunas.
    """
    try:
        wb = load_workbook(caminho_arquivo)
        
        for ws in wb.sheetnames:
            sheet = wb[ws]
            
            # ===== 1. MAPEAMENTO DE COLUNAS =====
            # Cria um dicionário para aplicar regras por nome
            col_map = {}
            for cell in sheet[1]:
                col_map[cell.column] = str(cell.value).lower().strip() if cell.value else ""

            # ===== 2. FORMATAR CABEÇALHO =====
            header_fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type='solid')
            header_font = Font(bold=True, color=cor_fonte_cabecalho, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            header_border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
            
            # ===== 3. FORMATAR DADOS =====
            data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            data_font = Font(size=10, color='333333')
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            data_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            # Colunas que devem ser TEXTO
            cols_texto = ['registro', 'nome', 'rg', 'cpf']
            
            # Colunas que devem ser INTEIRO (sem casas decimais)
            cols_inteiro = ['dias_trabalhado']

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.fill = data_fill
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border
                    
                    # Identifica o nome da coluna atual
                    col_name = col_map.get(cell.column, "")
                    
                    # --- REGRA 1: FORÇAR TEXTO ---
                    if col_name in cols_texto:
                        cell.number_format = '@'  # Formato Texto no Excel
                    
                    # --- REGRA 2: FORÇAR INTEIRO ---
                    elif col_name in cols_inteiro:
                        cell.number_format = '0'  # Inteiro
                        # Para separador de milhar sem decimais (ex: 1.200), usar '#,##0'
                    
                    # --- REGRA 3: DATAS ---
                    elif isinstance(cell.value, str) and '/' in str(cell.value):
                        try:
                            parts = str(cell.value).split('/')
                            if len(parts) == 3 and len(parts[2]) == 4:
                                cell.number_format = 'dd/mm/yyyy'
                        except:
                            pass
                    
                    # --- REGRA 4: OUTROS NÚMEROS ---
                    elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        if isinstance(cell.value, float) or (isinstance(cell.value, int) and cell.value > 999):
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
            
            # ===== 4. AJUSTAR LARGURA DAS COLUNAS =====
            for col_idx, col_name in enumerate(sheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                header_cell = sheet[f'{column_letter}1']
                if header_cell.value:
                    max_length = len(str(header_cell.value))
                
                for row_idx in range(2, min(sheet.max_row + 1, 101)):
                    cell = sheet[f'{column_letter}{row_idx}']
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max(max_length + 2, 12), 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            sheet.freeze_panes = 'A2'
            logging.info(f"   ✓ Aba '{ws}' formatada com sucesso")
        
        wb.save(caminho_arquivo)
        logging.info(f"✅ Arquivo '{caminho_arquivo}' formatado e salvo!\n")
        return True
        
    except Exception as e:
        logging.error(f"❌ Erro ao formatar Excel '{caminho_arquivo}': {e}\n")
        return False

# ==============================================================================
# 1. CARREGAMENTO E PRÉ-PROCESSAMENTO
# ==============================================================================

logging.info("="*80)
logging.info("1. CARREGAMENTO E PRÉ-PROCESSAMENTO DOS DADOS")
logging.info("="*80)

try:
    path_arquivo = r'X:\Gestao_de_Pessoas\Analytics\10 - Relatórios\10.4 - HC e Atestados Médicos\Controle_HC e Atestados.xlsb'
    dados_hc = pd.read_excel(path_arquivo, sheet_name='HC', engine='pyxlsb')
    logging.info(f"✓ Base de dados carregada com {dados_hc.shape[0]} registros e {dados_hc.shape[1]} colunas.")

    cols_texto = ['registro', 'nome', 'rg', 'cpf']
    
    for col in cols_texto:
        if col in dados_hc.columns:

            dados_hc[col] = dados_hc[col].astype(str).str.replace(r'\.0$', '',  regex=True).str.strip()
            
    if 'dias_trabalhado' in dados_hc.columns:
        dados_hc['dias_trabalhado'] = pd.to_numeric(dados_hc['dias_trabalhado'], errors='coerce').fillna(0).astype(int)

    # Limpar dados
    dados_hc = limpar_dados_completo(dados_hc)

    # Converter colunas de data
    colunas_data = ['nascimento', 'data_admissao', 'data_rescisao', 'data_nasc_conjuge']
    data_base = pd.Timestamp('1899-12-30')

    for coluna in colunas_data:
        try:
            if coluna in dados_hc.columns:
                dados_hc[coluna] = pd.to_numeric(dados_hc[coluna], errors='coerce')
                dados_hc[coluna] = data_base + pd.to_timedelta(dados_hc[coluna], unit='D')
                dados_hc[coluna] = dados_hc[coluna].dt.strftime('%d/%m/%Y')
        except:
            pass

    logging.info("✓ Colunas de data convertidas.")

    # Preencher valores faltantes
    dados_hc['filhos'] = dados_hc['filhos'].fillna('Não informado')
    dados_hc['secao'] = dados_hc['secao'].fillna('Não especificado')    
    dados_hc['centro_custo'] = dados_hc['centro_custo'].fillna('Não especificado')
    dados_hc['etnia_raca'] = dados_hc['etnia_raca'].fillna('Não informado')
    dados_hc['estado_civil'] = dados_hc['estado_civil'].fillna('Não informado')

    # Codificar logo
    caminho_imagem = 'Logo AFPESP.png'
    try:
        with open(caminho_imagem, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
        img_src_base64 = f"data:image/png;base64,{encoded_string}"
    except:
        img_src_base64 = ""

    logging.info("✓ SEÇÃO 1 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 1: {e}")
    sys.exit("Script encerrado.")

# ==============================================================================
# 2. ANÁLISE DE COLABORADORES ATIVOS
# ==============================================================================

logging.info("="*80)
logging.info("2. ANÁLISE DE COLABORADORES ATIVOS")
logging.info("="*80)

try:
    hc_ativos = dados_hc.loc[dados_hc['descricao_rescisao'] == 'ATIVO'].copy()
    logging.info(f"✓ {hc_ativos.shape[0]} colaboradores ativos identificados.")

    # Estatísticas descritivas
    colunas_quant = ['salario_total', 'idade', 'dias_trabalhado', 'primeiro_atestado', 'dias_afastado', 'custo_afastamento']
    df_quant_ativo = hc_ativos[colunas_quant].copy()
    df_quant_ativo.columns = ['Salário', 'Idade', 'Dias Trabalhado', 'Primeiro Atestado', 'Dias Afastado', 'Custo Afastamento']
    estat_ativo = df_quant_ativo.describe().round(2)

    # Card headcount
    headcount_ativo = hc_ativos.shape[0]
    fig_card_ativo = go.Figure(go.Indicator(
        mode="number",
        value=headcount_ativo,
        title={"text": "Headcount Ativo"},
        number={'font': {'size': 60, 'color': '#0070C0'}},
    ))
    fig_card_ativo.update_layout(height=150, margin=dict(l=0, r=0, t=30, b=0), paper_bgcolor="white")
    card_ativo_html = fig_card_ativo.to_html(full_html=False, include_plotlyjs='cdn')

    # Gráfico de Pizza - Sexo
    fig_sexo_ativo = px.pie(
        hc_ativos['sexo'].value_counts().reset_index(),
        values='count',
        names='sexo',
        title='Distribuição por Gênero',
        hole=0.3,
        color_discrete_sequence=['#0070C0', '#FF6B6B']
    )
    fig_sexo_ativo.update_layout(height=400, margin=dict(l=40, r=40, t=60, b=40))
    grafico_sexo_ativo = fig_sexo_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Pizza - Filhos
    fig_filhos_ativo = px.pie(
        hc_ativos['filhos'].value_counts().reset_index(),
        values='count',
        names='filhos',
        title='Distribuição por Filhos',
        hole=0.3,
        color_discrete_sequence=['#0070C0', '#FF6B6B', '#4ECDC4']
    )
    fig_filhos_ativo.update_layout(height=400, margin=dict(l=40, r=40, t=60, b=40))
    grafico_filhos_ativo = fig_filhos_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Barras - Etnia e Raça
    fig_etnia_ativo = px.bar(
        hc_ativos['etnia_raca'].value_counts().reset_index().sort_values('count'),
        x='count',
        y='etnia_raca',
        orientation='h',
        title='Distribuição por Etnia e Raça',
        color_discrete_sequence=['#0070C0']
    )
    fig_etnia_ativo.update_layout(
        height=400,
        margin=dict(l=120, r=40, t=60, b=40),
        showlegend=False,
        xaxis_title='Quantidade',
        yaxis_title='Etnia e Raça')
    grafico_etnia_ativo = fig_etnia_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Barras - Estado Civil
    fig_civil_ativo = px.bar(
        hc_ativos['estado_civil'].value_counts().reset_index().sort_values('count'),
        x='count',
        y='estado_civil',
        orientation='h',
        title='Distribuição por Estado Civil',
        color_discrete_sequence=['#0070C0']
    )
    fig_civil_ativo.update_layout(
        height=400, 
        margin=dict(l=120, r=40, t=60, b=40), 
        showlegend=False,
        xaxis_title='Quantidade',
        yaxis_title='Estado Civil')
    grafico_civil_ativo = fig_civil_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Barras - Formação
    fig_form_ativo = px.bar(
        hc_ativos['formacoes'].value_counts().reset_index().sort_values('count'),
        x='count',
        y='formacoes',
        orientation='h',
        title='Distribuição por Formação',
        color_discrete_sequence=['#0070C0']
    )
    fig_form_ativo.update_layout(
        height=500, 
        margin=dict(l=250, r=40, t=60, b=40), 
        showlegend=False, 
        xaxis_title='Quantidade',
        yaxis_title='Formação')
    grafico_form_ativo = fig_form_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Barras - Cargo
    fig_cargo_ativo = px.bar(
        hc_ativos['cargo_completo'].value_counts().head(15).reset_index().sort_values('count'),
        x='count',
        y='cargo_completo',
        orientation='h',
        title='Top 15 Cargos',
        color_discrete_sequence=['#0070C0']
    )
    fig_cargo_ativo.update_layout(
        height=500, 
        margin=dict(l=200, r=40, t=60, b=40), 
        showlegend=False, 
        xaxis_title='Quantidade', 
        yaxis_title='Cargo')
    grafico_cargo_ativo = fig_cargo_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Barras - Centro de Custo
    fig_cc_ativo = px.bar(
        hc_ativos['centro_custo'].value_counts().reset_index().sort_values('count'),
        x='count',
        y='centro_custo',
        orientation='h',
        title='Distribuição por Centro de Custo',
        color_discrete_sequence=['#0070C0']
    )
    fig_cc_ativo.update_layout(
        height=500, 
        margin=dict(l=200, r=40, t=60, b=40), 
        showlegend=False, 
        xaxis_title='Quantidade', 
        yaxis_title='Centro de Custo')
    grafico_cc_ativo = fig_cc_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Barras - Unidade
    fig_unidade_ativo = px.bar(
        hc_ativos['empresa_resumo'].value_counts().reset_index().sort_values('count'),
        x='count',
        y='empresa_resumo',
        orientation='h',
        title='Distribuição por Unidade',
        color_discrete_sequence=['#0070C0']
    )
    fig_unidade_ativo.update_layout(
        height=600, 
        margin=dict(l=150, r=40, t=60, b=40), 
        showlegend=False, 
        xaxis_title='Quantidade', 
        yaxis_title='Empresa')
    grafico_unidade_ativo = fig_unidade_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Boxplots
    fig_boxplot_ativo = make_subplots(
        rows=3, cols=2,
        subplot_titles=('Salário', 'Idade', 'Dias Trabalhado', 'Primeiro Atestado', 'Dias Afastado', 'Custo Afastamento'),
        specs=[[{'type': 'box'}, {'type': 'box'}], [{'type': 'box'}, {'type': 'box'}], [{'type': 'box'}, {'type': 'box'}]]
    )

    fig_boxplot_ativo.add_trace(go.Box(y=hc_ativos['salario_total'], name='Salário', marker_color='#0070C0'), row=1, col=1)
    fig_boxplot_ativo.add_trace(go.Box(y=hc_ativos['idade'], name='Idade', marker_color='#0070C0'), row=1, col=2)
    fig_boxplot_ativo.add_trace(go.Box(y=hc_ativos['dias_trabalhado'], name='Dias Trabalhado', marker_color='#0070C0'), row=2, col=1)
    fig_boxplot_ativo.add_trace(go.Box(y=hc_ativos['primeiro_atestado'], name='Primeiro Atestado', marker_color='#0070C0'), row=2, col=2)
    fig_boxplot_ativo.add_trace(go.Box(y=hc_ativos['dias_afastado'], name='Dias Afastado', marker_color='#0070C0'), row=3, col=1)
    fig_boxplot_ativo.add_trace(go.Box(y=hc_ativos['custo_afastamento'], name='Custo Afastamento', marker_color='#0070C0'), row=3, col=2)  

    fig_boxplot_ativo.update_layout(height=1000, showlegend=False, title_text="Boxplots - Variáveis Numéricas")
    boxplot_ativo_html = fig_boxplot_ativo.to_html(full_html=False, include_plotlyjs=False)

    # Tabela de estatísticas
    tabela_estat_ativo = estat_ativo.to_html(classes='tabela-estatistica', border=0)

    # Tabela de colaboradores
    colunas_tabela_ativos = ['registro', 'nome', 'idade', 'rg', 'cpf', 'data_admissao', 'situacao', 'dias_trabalhado',
                             'cargo_completo', 'salario_total', 'secao', 'centro_custo', 'empresa_resumo', 'formacoes']
    tabela_ativos = hc_ativos[colunas_tabela_ativos].to_html(index=False, classes='tabela-dados', border=0)

    logging.info("✓ SEÇÃO 2 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 2: {e}")
    traceback.print_exc()
    card_ativo_html = "Erro ao processar dados de ativos"
    grafico_sexo_ativo = ""
    grafico_filhos_ativo = ""
    grafico_etnia_ativo = ""
    grafico_civil_ativo = ""
    grafico_form_ativo = ""
    grafico_cargo_ativo = ""
    grafico_cc_ativo = ""
    grafico_unidade_ativo = ""
    boxplot_ativo_html = ""
    tabela_estat_ativo = ""
    tabela_ativos = ""

# ==============================================================================
# 3. ANÁLISE DE DEMISSÕES
# ==============================================================================

logging.info("="*80)
logging.info("3. ANÁLISE DE DEMISSÕES")
logging.info("="*80)

try:
    hc_inativos = dados_hc.loc[dados_hc['ano_rescisao'] == 2026].copy()
    logging.info(f"✓ {hc_inativos.shape[0]} demissões em 2026 identificadas.")

    # Card headcount
    headcount_inativo = hc_inativos.shape[0]
    fig_card_inativo = go.Figure(go.Indicator(
        mode="number",
        value=headcount_inativo,
        title={"text": "Demissões 2026"},
        number={'font': {'size': 60, 'color': '#FF6B6B'}},
    ))
    fig_card_inativo.update_layout(height=150, margin=dict(l=0, r=0, t=30, b=0), paper_bgcolor="white")
    card_inativo_html = fig_card_inativo.to_html(full_html=False, include_plotlyjs='cdn')

    # Gráfico de Corrida
    hc_inativos['Data de Rescisão'] = pd.to_datetime(hc_inativos['data_rescisao'], format='%d/%m/%Y', errors='coerce')
    hc_inativos_corrida = hc_inativos.dropna(subset=['Data de Rescisão'])

    if len(hc_inativos_corrida) > 0:
        contagem_corrida = hc_inativos_corrida.groupby(['empresa_resumo', 'Data de Rescisão']).size().reset_index(name='Demitidos')
        
        todas_empresas = contagem_corrida['empresa_resumo'].unique()
        todas_datas = pd.date_range(
            start=contagem_corrida['Data de Rescisão'].min(),
            end=contagem_corrida['Data de Rescisão'].max()
        )
        
        combinacoes = pd.MultiIndex.from_product(
            [todas_empresas, todas_datas],
            names=['empresa_resumo', 'Data de Rescisão']
        ).to_frame(index=False)
        
        contagem_completo = pd.merge(combinacoes, contagem_corrida, on=['empresa_resumo', 'Data de Rescisão'], how='left')
        contagem_completo['Demitidos'] = contagem_completo['Demitidos'].fillna(0)
        contagem_completo = contagem_completo.sort_values(by=['empresa_resumo', 'Data de Rescisão'])
        contagem_completo['Demissão Acumulado'] = contagem_completo.groupby('empresa_resumo')['Demitidos'].cumsum()
        contagem_completo = contagem_completo.rename(columns={'empresa_resumo': 'Empresa'})

        # ANTES DE GERAR O FIG: Formatar a data para o padrão BR e remover horas
        contagem_completo['Data Formatada'] = contagem_completo['Data de Rescisão'].dt.strftime('%d/%m/%Y')

        fig_corrida_inativo = px.bar(
            contagem_completo,
            x='Demissão Acumulado',
            y='Empresa',
            orientation='h',
            color='Empresa',
            text='Demissão Acumulado',
            animation_frame='Data Formatada',
            animation_group='Empresa',
            title='Demissões Acumuladas por Unidade'
        )
        fig_corrida_inativo.update_layout(
            height=600,
            margin=dict(l=150, r=40, t=60, b=40),
            yaxis_title='Empresa')
        grafico_corrida_inativo = fig_corrida_inativo.to_html(full_html=False, include_plotlyjs=False)
    else:
        grafico_corrida_inativo = "<p>Sem dados para gráfico de corrida</p>"

    # Gráfico por Mês
    hc_inativos['mes_rescisao'] = pd.to_numeric(hc_inativos['mes_rescisao'], errors='coerce')
    meses_nome = {1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun', 
                  7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'}

    contagem_mes = hc_inativos['mes_rescisao'].value_counts().sort_index()
    nomes_meses = contagem_mes.index.map(meses_nome)

    # --- AJUSTE PARA COMPARATIVO ANO CONTRA ANO (DEMISSÕES) ---
    df_yoy_dem = dados_hc.loc[dados_hc['ano_rescisao'].isin([2025, 2026])].copy()
    
    # Converter ano para string para o Plotly tratar como categoria
    df_yoy_dem['ano_rescisao'] = df_yoy_dem['ano_rescisao'].astype(int).astype(str)
    df_yoy_dem['mes_rescisao'] = pd.to_numeric(df_yoy_dem['mes_rescisao'], errors='coerce')
    
    contagem_yoy_dem = df_yoy_dem.groupby(['mes_rescisao', 'ano_rescisao']).size().reset_index(name='Quantidade')
    contagem_yoy_dem['Mês'] = contagem_yoy_dem['mes_rescisao'].map(meses_nome)
    contagem_yoy_dem = contagem_yoy_dem.sort_values('mes_rescisao')

    fig_mes_inativo = px.bar(
        contagem_yoy_dem,
        x='Mês',
        y='Quantidade',
        color='ano_rescisao', 
        barmode='group',     # Agora o agrupamento funcionará corretamente
        title='Demissões: 2025 vs 2026',
        color_discrete_map={'2025': '#D97373', '2026': '#FC813F'} # Chaves como string
    )
    
    fig_mes_inativo.update_layout(
        height=400, 
        margin=dict(l=40, r=40, t=60, b=40), 
        legend_title_text='Ano',
        xaxis_title='Mês',
        yaxis_title='Quantidade')
    
    grafico_mes_inativo = fig_mes_inativo.to_html(full_html=False, include_plotlyjs=False)

    # Gráfico de Tipo de Rescisão
    # --- AJUSTE PARA COMPARATIVO ANO CONTRA ANO (TIPO DE RESCISÃO) ---
    # DataFrame temporário incluindo 2025 e 2026
    df_yoy_tipo = dados_hc.loc[dados_hc['ano_rescisao'].isin([2025, 2026])].copy()
    
    # Ano em texto para a legenda ficar limpa (sem .0)
    df_yoy_tipo['ano_rescisao'] = df_yoy_tipo['ano_rescisao'].astype(int).astype(str)
    
    # Agrupado por Tipo de Rescisão e Ano
    contagem_tipo = df_yoy_tipo.groupby(['descricao_rescisao', 'ano_rescisao']).size().reset_index(name='Quantidade')
    
    # Ordenação: os tipos com mais demissões apareçam primeiro
    ordem_tipos = contagem_tipo.groupby('descricao_rescisao')['Quantidade'].sum().sort_values(ascending=True).index

    fig_tipo_inativo = px.bar(
        contagem_tipo,
        x='Quantidade',
        y='descricao_rescisao',
        orientation='h',
        color='ano_rescisao',
        barmode='group', # Coloca as barras de 2025 e 2026 lado a lado
        title='Distribuição por Tipo de Rescisão: 2025 vs 2026',
        color_discrete_map={'2025': '#D97373', '2026': '#FC813F'},
        category_orders={'descricao_rescisao': list(ordem_tipos)} # Mantém a ordem visual
    )
    
    fig_tipo_inativo.update_layout(
        height=500,
        margin=dict(l=250, r=40, t=60, b=40),
        legend_title_text='Ano',
        yaxis_title='Descrição da Rescisão',
        xaxis_title='Quantidade')
    
    grafico_tipo_inativo = fig_tipo_inativo.to_html(full_html=False, include_plotlyjs=False)

    # Tabela de demitidos
    colunas_tabela_inativos = ['registro', 'nome', 'idade', 'rg', 'cpf', 'data_admissao', 'data_rescisao', 'dias_trabalhado',
                               'descricao_rescisao', 'cargo_completo', 'salario_total', 'empresa_resumo', 'formacoes']
    tabela_inativos = hc_inativos[colunas_tabela_inativos].to_html(index=False, classes='tabela-dados', border=0)

    logging.info("✓ SEÇÃO 3 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 3: {e}")
    traceback.print_exc()
    card_inativo_html = "Erro ao processar dados de demissões"
    grafico_corrida_inativo = ""
    grafico_mes_inativo = ""
    grafico_tipo_inativo = ""
    tabela_inativos = ""

# ==============================================================================
# 4. ANÁLISE DE ADMISSÕES
# ==============================================================================

logging.info("="*80)
logging.info("4. ANÁLISE DE ADMISSÕES")
logging.info("="*80)

try:
    hc_admitidos = dados_hc.loc[dados_hc['ano_admissao'] == 2026].copy()
    logging.info(f"✓ {hc_admitidos.shape[0]} admissões em 2026 identificadas.")

    # Card headcount
    headcount_admitido = hc_admitidos.shape[0]
    fig_card_admitido = go.Figure(go.Indicator(
        mode="number",
        value=headcount_admitido,
        title={"text": "Admissões 2026"},
        number={'font': {'size': 60, 'color': '#28a745'}},
    ))
    fig_card_admitido.update_layout(height=150, margin=dict(l=0, r=0, t=30, b=0), paper_bgcolor="white")
    card_admitido_html = fig_card_admitido.to_html(full_html=False, include_plotlyjs='cdn')

    # Gráfico de Corrida
    if len(hc_admitidos) > 0:
        hc_admitidos['Data de Admissão'] = pd.to_datetime(hc_admitidos['data_admissao'], format='%d/%m/%Y', errors='coerce')
        hc_admitidos_corrida = hc_admitidos.dropna(subset=['Data de Admissão'])
        
        if len(hc_admitidos_corrida) > 0:
            contagem_corrida_adm = hc_admitidos_corrida.groupby(['empresa_resumo', 'Data de Admissão']).size().reset_index(name='Admitidos')
            
            todas_empresas_adm = contagem_corrida_adm['empresa_resumo'].unique()
            todas_datas_adm = pd.date_range(
                start=contagem_corrida_adm['Data de Admissão'].min(),
                end=contagem_corrida_adm['Data de Admissão'].max()
            )
            
            combinacoes_adm = pd.MultiIndex.from_product(
                [todas_empresas_adm, todas_datas_adm],
                names=['empresa_resumo', 'Data de Admissão']
            ).to_frame(index=False)
            
            contagem_completo_adm = pd.merge(combinacoes_adm, contagem_corrida_adm, on=['empresa_resumo', 'Data de Admissão'], how='left')
            contagem_completo_adm['Admitidos'] = contagem_completo_adm['Admitidos'].fillna(0)
            contagem_completo_adm = contagem_completo_adm.sort_values(by=['empresa_resumo', 'Data de Admissão'])
            contagem_completo_adm['Admissão Acumulado'] = contagem_completo_adm.groupby('empresa_resumo')['Admitidos'].cumsum()
            contagem_completo_adm = contagem_completo_adm.rename(columns={'empresa_resumo': 'Empresa'})

            # ANTES DE GERAR O FIG: Formatar a data para o padrão BR e remover horas
            contagem_completo_adm['Data Formatada'] = contagem_completo_adm['Data de Admissão'].dt.strftime('%d/%m/%Y')
            
            fig_corrida_admitido = px.bar(
                contagem_completo_adm,
                x='Admissão Acumulado',
                y='Empresa',
                orientation='h',
                color='Empresa',
                text='Admissão Acumulado',
                animation_frame='Data Formatada',
                animation_group='Empresa',
                title='Admissões Acumuladas por Unidade'
            )
            fig_corrida_admitido.update_layout(
                height=600,
                margin=dict(l=150, r=40, t=60, b=40),
                yaxis_title='Empresa')
            grafico_corrida_admitido = fig_corrida_admitido.to_html(full_html=False, include_plotlyjs=False)
        else:
            grafico_corrida_admitido = "<p>Sem dados para gráfico de corrida</p>"
    else:
        grafico_corrida_admitido = "<p>Sem dados de admissões em 2026</p>"

    # Gráfico por Mês
    if len(hc_admitidos) > 0:
        hc_admitidos['mes_admissao'] = pd.to_numeric(hc_admitidos['mes_admissao'], errors='coerce')
        meses_nome = {1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun', 
                      7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'}
        
        contagem_mes_adm = hc_admitidos['mes_admissao'].value_counts().sort_index()
        
        if len(contagem_mes_adm) > 0:
            nomes_meses_adm = contagem_mes_adm.index.map(meses_nome)
            
    # --- AJUSTE PARA COMPARATIVO ANO CONTRA ANO (ADMISSÕES) ---
    df_yoy_adm = dados_hc.loc[dados_hc['ano_admissao'].isin([2025, 2026])].copy()
    
    # Converter ano para string
    df_yoy_adm['ano_admissao'] = df_yoy_adm['ano_admissao'].astype(int).astype(str)
    df_yoy_adm['mes_admissao'] = pd.to_numeric(df_yoy_adm['mes_admissao'], errors='coerce')

    contagem_yoy_adm = df_yoy_adm.groupby(['mes_admissao', 'ano_admissao']).size().reset_index(name='Quantidade')
    contagem_yoy_adm['Mês'] = contagem_yoy_adm['mes_admissao'].map(meses_nome)
    contagem_yoy_adm = contagem_yoy_adm.sort_values('mes_admissao')

    fig_mes_admitido = px.bar(
        contagem_yoy_adm,
        x='Mês',
        y='Quantidade',
        color='ano_admissao', 
        barmode='group',
        title='Admissões: 2025 vs 2026',
        color_discrete_map={'2025': '#05BEE8', '2026': '#3F62FC'} # Chaves como string
    )
    
    fig_mes_admitido.update_layout(
        height=400, 
        margin=dict(l=40, r=40, t=60, b=40), 
        legend_title_text='Ano',
        xaxis_title='Mês',
        yaxis_title='Quantidade')
    
    grafico_mes_admitido = fig_mes_admitido.to_html(full_html=False, include_plotlyjs=False)

    # Tabela de admitidos
    colunas_tabela_admitidos = ['registro', 'nome', 'idade', 'rg', 'cpf', 'data_admissao', 'situacao', 'data_rescisao', 
                                'dias_trabalhado', 'descricao_rescisao', 'cargo_completo', 'salario_total', 'secao',
                                'centro_custo', 'empresa_resumo', 'formacoes']

    tabela_admitidos = hc_admitidos[colunas_tabela_admitidos].to_html(index=False, classes='tabela-dados', border=0)

    logging.info("✓ SEÇÃO 4 CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 4: {e}")
    traceback.print_exc()
    card_admitido_html = "Erro ao processar dados de admissões"
    grafico_corrida_admitido = "Erro ao gerar gráfico de corrida"
    grafico_mes_admitido = "Erro ao gerar gráfico por mês"
    tabela_admitidos = "Erro ao gerar tabela"

# ==============================================================================
# PREPARAÇÃO FINAL DOS DADOS
# ==============================================================================

logging.info("⚙️ Aplicando tipagem forte nas colunas antes da exportação...")

# Lista de DataFrames para processar
dfs_para_ajustar = [hc_ativos, hc_inativos, hc_admitidos]

for df in dfs_para_ajustar:
    # 1. Forçar colunas de TEXTO
    cols_texto = ['registro', 'nome', 'rg', 'cpf']
    for col in cols_texto:
        if col in df.columns:
            # Converte para string e remove '.0' caso tenha vindo como float
            df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')

    # 2. Forçar coluna de INTEIRO (dias_trabalhado)
    if 'dias_trabalhado' in df.columns:
        # Preenche vazios com 0, converte para inteiro
        df['dias_trabalhado'] = df['dias_trabalhado'].fillna(0).astype(int)

# ==============================================================================
# 5. GERAÇÃO DO RELATÓRIO HTML E EXCEL
# ==============================================================================

logging.info("="*80)
logging.info("5. GERAÇÃO DO RELATÓRIO HTML E EXCEL")
logging.info("="*80)

try:
    # Salvar arquivos Excel com formatação unificada
    arquivo_ativos = 'HC_Ativos.xlsx'
    arquivo_demissoes = 'HC_Demissoes_2026.xlsx'
    arquivo_admissoes = 'HC_Admissoes_2026.xlsx'

    def converter_datas_para_string(df, colunas_data_str):
        df_copia = df.copy()
        
        for col in colunas_data_str:
            if col in df_copia.columns:
                if pd.api.types.is_datetime64_any_dtype(df_copia[col]):
                    df_copia[col] = df_copia[col].dt.strftime('%d/%m/%Y')
                df_copia[col] = df_copia[col].fillna('')
        
        return df_copia

    # Definir colunas de data
    colunas_datas = ['nascimento', 'data_admissao', 'data_rescisao', 'data_nasc_conjuge']
    
    logging.info("💾 Salvando e formatando Excel - Ativos...")
    hc_ativos_export = converter_datas_para_string(hc_ativos[colunas_tabela_ativos], colunas_datas)
    hc_ativos_export.to_excel(arquivo_ativos, index=False, sheet_name='Ativos')
    formatar_excel_profissional(arquivo_ativos)
    
    logging.info("💾 Salvando e formatando Excel - Demissões...")
    hc_inativos_export = converter_datas_para_string(hc_inativos[colunas_tabela_inativos], colunas_datas)
    hc_inativos_export.to_excel(arquivo_demissoes, index=False, sheet_name='Demissões 2026')
    formatar_excel_profissional(arquivo_demissoes)
    
    logging.info("💾 Salvando e formatando Excel - Admissões...")
    hc_admitidos_export = converter_datas_para_string(hc_admitidos[colunas_tabela_admitidos], colunas_datas)
    hc_admitidos_export.to_excel(arquivo_admissoes, index=False, sheet_name='Admissões 2026')
    formatar_excel_profissional(arquivo_admissoes)

    logging.info("✓ Arquivos Excel salvos e formatados com sucesso\n")

except Exception as e:
    logging.error(f"ERRO ao salvar/formatar Excel na SEÇÃO 5: {e}")
    traceback.print_exc()

# ==============================================================================
# 6. ANÁLISE DE TURNOVER
# ==============================================================================
logging.info("Gerando análise de Turnover 2025 vs 2026...")

# 1. Converter as colunas corretas para datetime
dados_hc['data_rescisao'] = pd.to_datetime(dados_hc['data_rescisao'], dayfirst=True, errors='coerce')
dados_hc['data_admissao'] = pd.to_datetime(dados_hc['data_admissao'], dayfirst=True, errors='coerce')

# --- DEBUG RÁPIDO PARA CONFERÊNCIA --- #
# Imprimir no terminal as datas de demissão de 2026
demissoes_2026_check = dados_hc[dados_hc['data_rescisao'].dt.year == 2026]['data_rescisao']
logging.info(f"Datas de demissão identificadas em 2026:\n{demissoes_2026_check}")

# Gera lista de meses para análise (Jan/25 a Dez/26)
meses_analise = pd.date_range(start='2025-01-01', end='2026-12-31', freq='ME')
dados_turnover = []

for data_corte in meses_analise:
    ano = data_corte.year
    mes = data_corte.month
    
    # 2. Demissões no mês
    demissoes_mes = dados_hc[
        (dados_hc['data_rescisao'].dt.year == ano) & 
        (dados_hc['data_rescisao'].dt.month == mes)
    ].shape[0]
    
    # 3. Ativos no fim do mês
    ativos_fim_mes = dados_hc[
        (dados_hc['data_admissao'] <= data_corte) & 
        ((dados_hc['data_rescisao'].isna()) | (dados_hc['data_rescisao'] > data_corte))
    ].shape[0]
    
    # 4. Cálculo da Taxa %
    taxa = (demissoes_mes / ativos_fim_mes * 100) if ativos_fim_mes > 0 else 0
    
    dados_turnover.append({
        'Ano': str(ano),
        'Mes_Nome': data_corte.strftime('%b'),
        'Mes_Num': mes,
        'Taxa': taxa
    })

df_turnover = pd.DataFrame(dados_turnover)

# Criar Figura Plotly
fig_turnover = go.Figure()

# Linha 2025
df_2025 = df_turnover[df_turnover['Ano'] == '2025']
fig_turnover.add_trace(go.Scatter(
    x=df_2025['Mes_Nome'], y=df_2025['Taxa'],
    mode='lines+markers', name='2025',
    line=dict(color='#D97373', width=2)
))

# Linha 2026
df_2026 = df_turnover[df_turnover['Ano'] == '2026']
fig_turnover.add_trace(go.Scatter(
    x=df_2026['Mes_Nome'], y=df_2026['Taxa'],
    mode='lines+markers', name='2026',
    line=dict(color='#FC813F', width=4)
))

# Linha de Controle (Ex: 2%)
fig_turnover.add_hline(y=2, line_dash="dot", annotation_text="Limite 2%", line_color="red")

fig_turnover.update_layout(
    title='<b>Taxa de Turnover Mensal: 2025 vs 2026</b>',
    yaxis_title='Turnover (%)',
    template='plotly_white',
    height=400
)

# Converter para HTML string
grafico_turnover = fig_turnover.to_html(full_html=False, include_plotlyjs='cdn')

# ==============================================================================
# 7. ANÁLISE DE EARLY TURNOVER (< 90 DIAS)
# ==============================================================================
logging.info("Gerando análise de Early Turnover (Qualidade de Contratação)...")

# 1. Filtrar demissões de 2026
df_early = dados_hc[
    (dados_hc['data_rescisao'].dt.year == 2026)
].copy()

# 2. Calcular tempo de casa em dias
df_early['dias_trabalhados'] = (df_early['data_rescisao'] - df_early['data_admissao']).dt.days

# 3. Criar categoria (Early vs Orgânico)
df_early['tipo_saida'] = df_early['dias_trabalhados'].apply(
    lambda x: 'Early Turnover (<90 dias)' if x <= 90 else 'Turnover Orgânico (>90 dias)'
)

# --- CÁLCULOS MATEMÁTICOS ---
total_demissoes_2026 = df_early.shape[0]
early_count = df_early[df_early['dias_trabalhados'] <= 90].shape[0]

# Evitar divisão por zero
pct_early = (early_count / total_demissoes_2026 * 100) if total_demissoes_2026 > 0 else 0

# --- KPI GERAL (INDICADOR - GAUGE) ---
fig_early_gauge = go.Figure(go.Indicator(
    mode = "gauge+number",
    value = pct_early,
    number = {'suffix': "%", 'font': {'size': 40}}, 
    title = {'text': "<b>Early Turnover 2026 (Saídas < 90 dias)</b>"},
    gauge = {
        'axis': {'range': [None, 100]},
        'bar': {'color': "#ef4444"},
        'steps': [
            {'range': [0, 15], 'color': "#dcfce7"},
            {'range': [15, 30], 'color': "#fef9c3"},
            {'range': [30, 100], 'color': "#fee2e2"}
        ],
        'threshold': {
            'line': {'color': "red", 'width': 4},
            'thickness': 0.75,
            'value': 20
        }
    }
))

# Ajuste de Margens
fig_early_gauge.update_layout(
    height=350,
    margin=dict(l=30, r=30, t=80, b=30)
)

# --- GRÁFICO TEMPORAL (BARRAS EMPILHADAS) ---
# Agrupar por mês e tipo de saída
df_early['mes_num'] = df_early['data_rescisao'].dt.month
df_early['mes_nome'] = df_early['data_rescisao'].dt.strftime('%b')

resumo_early = df_early.groupby(['mes_num', 'mes_nome', 'tipo_saida']).size().reset_index(name='qtd')
resumo_early = resumo_early.sort_values('mes_num')

fig_early_bar = px.bar(
    resumo_early, 
    x='mes_nome', 
    y='qtd', 
    color='tipo_saida',
    title="<b>Perfil das Demissões: Recentes vs. Antigos (Mês a Mês)</b>",
    color_discrete_map={
        'Early Turnover (<90 dias)': '#ef4444',
        'Turnover Orgânico (>90 dias)': '#3b82f6'
    },
    text_auto=True
)

# Ajuste de Layout Vertical
fig_early_bar.update_layout(
    barmode='stack', 
    xaxis_title="Mês", 
    yaxis_title="Qtd. Demissões",
    legend_title="Tempo de Casa",
    template='plotly_white',
    height=400,
    margin=dict(t=50, b=50)
)

# Converter para HTML
grafico_early_gauge = fig_early_gauge.to_html(full_html=False, include_plotlyjs='cdn')
grafico_early_bar = fig_early_bar.to_html(full_html=False, include_plotlyjs='cdn')

# ==============================================================================
# 8. ANÁLISE DE SOBREVIVÊNCIA (TENURE / TEMPO DE CASA)
# ==============================================================================
logging.info("Gerando análise de Sobrevivência (Tenure dos Demitidos)...")

# 1. Filtrar demissões de 2026
# Vamos usar 2026 para focar no problema atual, mas pode alterar para pegar tudo.
df_tenure = dados_hc[
    (dados_hc['data_rescisao'].dt.year == 2026)
].copy()

# 2. Calcular tempo de casa em ANOS (para facilitar leitura)
df_tenure['anos_casa'] = (df_tenure['data_rescisao'] - df_tenure['data_admissao']).dt.days / 365.25

# --- GRÁFICO 1: HISTOGRAMA (DISTRIBUIÇÃO GERAL) ---
# Define bins (faixas): 0-1 ano, 1-2 anos, etc.
fig_tenure_hist = px.histogram(
    df_tenure, 
    x="anos_casa",
    nbins=20, # Ajuste a granularidade
    title="<b>Ponto de Ruptura: Com quanto tempo de casa as pessoas saem?</b>",
    labels={'anos_casa': 'Tempo de Casa (Anos)'},
    color_discrete_sequence=['#6366f1'], # Cor Indigo
    text_auto=True
)

fig_tenure_hist.update_layout(
    yaxis_title="Qtd. Demissões",
    xaxis_title="Anos de Empresa",
    bargap=0.1, # Espaço entre barras
    template='plotly_white',
    height=400
)

# Adiciona uma linha vertical na Média
media_anos = df_tenure['anos_casa'].mean()
fig_tenure_hist.add_vline(
    x=media_anos, 
    line_dash="dash", 
    line_color="red", 
    annotation_text=f"Média: {media_anos:.1f} anos", 
    annotation_position="top right"
)

# --- GRÁFICO 2: BOXPLOT POR CARGO (TOP 10 MAIS DEMITIDOS) ---
# Descobrir os cargos com mais demissões para não poluir o gráfico
top_cargos_demissao = df_tenure['cargo_completo'].value_counts().nlargest(10).index
df_tenure_top = df_tenure[df_tenure['cargo_completo'].isin(top_cargos_demissao)]

fig_tenure_box = px.box(
    df_tenure_top, 
    x="anos_casa", 
    y="cargo_completo",
    title="<b>Variação do Tempo de Casa por Cargo (Top 10 Demissões)</b>",
    color="cargo_completo",
    points="all" # Mostra os pontos individuais
)

fig_tenure_box.update_layout(
    showlegend=False,
    xaxis_title="Anos de Empresa",
    yaxis_title="",
    template='plotly_white',
    height=500,
    margin=dict(l=10)
)

# Converter para HTML
grafico_tenure_hist = fig_tenure_hist.to_html(full_html=False, include_plotlyjs='cdn')
grafico_tenure_box = fig_tenure_box.to_html(full_html=False, include_plotlyjs='cdn')

# ==============================================================================
# 9. ANÁLISE FINANCEIRA (VISÃO FP&A - CUSTO DE SUBSTITUIÇÃO)
# ==============================================================================
logging.info("Gerando análise Financeira de Turnover...")

# 1. Configuração do Fator de Custo (Pode ajustar para 2.0 se quiser ser mais agressivo)
FATOR_CUSTO = 1.5 

# 2. Filtrar demissões de 2026
df_financeiro = dados_hc[
    (dados_hc['data_rescisao'].dt.year == 2026)
].copy()

# 3. Tratamento da Coluna de Salário (Garantir que é número)
col_salario = 'salario_atual' if 'salario_atual' in df_financeiro.columns else 'salario_total'

# Remove caracteres de moeda se for string e converte
if df_financeiro[col_salario].dtype == 'object':
    df_financeiro[col_salario] = df_financeiro[col_salario].astype(str).str.replace('R$', '', regex=False).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)

df_financeiro[col_salario] = pd.to_numeric(df_financeiro[col_salario], errors='coerce').fillna(0)

# 4. Cálculos Financeiros
soma_salarios_demitidos = df_financeiro[col_salario].sum()
custo_total_estimado = soma_salarios_demitidos * FATOR_CUSTO
custo_medio_por_demissao = custo_total_estimado / len(df_financeiro) if len(df_financeiro) > 0 else 0

# --- VISUALIZAÇÃO (BIG NUMBERS) ---
fig_financeiro = go.Figure()

fig_financeiro.add_trace(go.Indicator(
    mode = "number",
    value = custo_total_estimado,
    number = {'prefix': "R$ ", 'valueformat': ",.2f", 'font': {'size': 50, 'color': '#b91c1c'}}, # Vermelho Escuro
    title = {
        'text': f"<b>Impacto Financeiro Estimado (YTD 2026)</b><br><span style='font-size:0.6em;color:gray'>Baseado em {FATOR_CUSTO}x o salário nominal (Rescisão + Recrutamento + Treinamento)</span>",
        'font': {'size': 18}
    },
    domain = {'x': [0, 1], 'y': [0, 1]}
))

fig_financeiro.update_layout(
    height=250,
    margin=dict(l=20, r=20, t=60, b=20),
    template='plotly_white'
)

# ==============================================================================
# 10. PAINEL DE KPIs ESTRATÉGICOS (CONSOLIDAÇÃO)
# ==============================================================================
logging.info("Gerando Painel de KPIs Estratégicos...")

# --- 1. CÁLCULO DO TURNOVER YTD (ACUMULADO 2026) ---
# Fórmula: Total Demissões 2026 / Média de Ativos 2026
total_demissoes_ytd = dados_hc[dados_hc['data_rescisao'].dt.year == 2026].shape[0]

# Média de ativos mensal em 2026
meses_2026 = pd.date_range(start='2026-01-01', end=pd.Timestamp.now(), freq='ME')
soma_ativos = 0
for data_corte in meses_2026:
    ativos = dados_hc[
        (dados_hc['data_admissao'] <= data_corte) & 
        ((dados_hc['data_rescisao'].isna()) | (dados_hc['data_rescisao'] > data_corte))
    ].shape[0]
    soma_ativos += ativos

media_ativos_ytd = soma_ativos / len(meses_2026) if len(meses_2026) > 0 else 0
taxa_turnover_ytd = (total_demissoes_ytd / media_ativos_ytd * 100) if media_ativos_ytd > 0 else 0

# --- 2. GERAÇÃO DOS CARDS (HTML PURO PARA LEVEZA) ---
# HTML para ter flexibilidade total no layout
kpi_dashboard_html = f"""
<div style="display: flex; gap: 20px; margin-bottom: 30px; flex-wrap: wrap;">
    
    <!-- CARD 1: TURNOVER YTD -->
    <div style="flex: 1; min-width: 200px; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-left: 5px solid #3b82f6;">
        <h3 style="margin: 0; font-size: 14px; color: #64748b;">Turnover YTD (2026)</h3>
        <p style="margin: 10px 0 0 0; font-size: 28px; font-weight: bold; color: #1e293b;">{taxa_turnover_ytd:.1f}%</p>
        <span style="font-size: 12px; color: #94a3b8;">Acumulado do ano</span>
    </div>

    <!-- CARD 2: EARLY TURNOVER -->
    <div style="flex: 1; min-width: 200px; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-left: 5px solid #ef4444;">
        <h3 style="margin: 0; font-size: 14px; color: #64748b;">Early Turnover</h3>
        <p style="margin: 10px 0 0 0; font-size: 28px; font-weight: bold; color: #1e293b;">{pct_early:.1f}%</p>
        <span style="font-size: 12px; color: #94a3b8;">Saídas < 90 dias</span>
    </div>

    <!-- CARD 3: TEMPO MÉDIO (TENURE) -->
    <div style="flex: 1; min-width: 200px; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-left: 5px solid #8b5cf6;">
        <h3 style="margin: 0; font-size: 14px; color: #64748b;">Tempo Médio de Casa</h3>
        <p style="margin: 10px 0 0 0; font-size: 28px; font-weight: bold; color: #1e293b;">{media_anos:.1f} anos</p>
        <span style="font-size: 12px; color: #94a3b8;">Média dos demitidos</span>
    </div>

    <!-- CARD 4: CUSTO ESTIMADO -->
    <div style="flex: 1; min-width: 200px; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-left: 5px solid #10b981;">
        <h3 style="margin: 0; font-size: 14px; color: #64748b;">Custo Est. (YTD)</h3>
        <p style="margin: 10px 0 0 0; font-size: 28px; font-weight: bold; color: #1e293b;">R$ {custo_total_estimado:,.0f}</p>
        <span style="font-size: 12px; color: #94a3b8;">Impacto Financeiro</span>
    </div>

</div>
"""

# Converter para HTML
grafico_financeiro = fig_financeiro.to_html(full_html=False, include_plotlyjs='cdn')

    # HTML com referência aos arquivos Excel
try:
    html_final = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Análise de HC - Indicadores de Gestão de Pessoas</title>
    <script src="https://cdn.plot.ly/plotly-2.18.2.min.js"></script>
    <!-- CRÍTICO: SheetJS importado ANTES do script customizado -->
    <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f5f5;
            color: #333;
        }}
        
        .header {{
            background: linear-gradient(135deg, #0070C0 0%, #003d7a 100%);
            color: white;
            padding: 40px 20px;
            text-align: center;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        
        .header img {{
            max-width: 150px;
            margin-bottom: 20px;
        }}
        
        .header h1 {{
            font-size: 32px;
            margin-bottom: 10px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
        }}
        
        .tabs {{
            display: flex;
            gap: 10px;
            margin-bottom: 30px;
            border-bottom: 2px solid #ddd;
            flex-wrap: wrap;
        }}
        
        .tab-button {{
            padding: 12px 24px;
            background-color: #f0f0f0;
            border: none;
            cursor: pointer;
            font-size: 16px;
            border-radius: 5px 5px 0 0;
            transition: all 0.3s;
            font-weight: 500;
        }}
        
        .tab-button.active {{
            background-color: #0070C0;
            color: white;
        }}
        
        .tab-button:hover {{
            background-color: #0070C0;
            color: white;
        }}
        
        .tab-content {{
            display: none;
            animation: fadeIn 0.3s;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; }}
            to {{ opacity: 1; }}
        }}
        
        .card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .card-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
        }}
        
        .card-header h2 {{
            margin: 0;
            color: #0070C0;
            border: none;
            padding: 0;
        }}
        
        .download-button {{
            padding: 10px 20px;
            background-color: #28a745;
            color: white;
            border: none;
            cursor: pointer;
            font-size: 14px;
            border-radius: 5px;
            transition: all 0.3s;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        
        .download-button:hover {{
            background-color: #218838;
            transform: scale(1.05);
        }}
        
        .download-button:active {{
            transform: scale(0.98);
        }}
        
        /* ===== ESTILO DO BOTÃO VOLTAR AO TOPO ===== */
        .btn-voltar-topo {{
            position: fixed;
            bottom: 30px;
            right: 30px;
            width: 50px;
            height: 50px;
            background-color: #0070C0;
            color: white;
            border: none;
            border-radius: 50%;
            cursor: pointer;
            font-size: 24px;
            display: none;
            align-items: center;
            justify-content: center;
            box-shadow: 0 4px 12px rgba(0, 112, 192, 0.4);
            transition: all 0.3s ease;
            z-index: 1000;
        }}
        
        .btn-voltar-topo:hover {{
            background-color: #003d7a;
            transform: translateY(-3px);
            box-shadow: 0 6px 16px rgba(0, 112, 192, 0.6);
        }}
        
        .btn-voltar-topo:active {{
            transform: translateY(-1px);
        }}
        
        .btn-voltar-topo.visible {{
            display: flex;
        }}
        /* ===== FIM DO ESTILO DO BOTÃO ===== */
        
        .metrics {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .grafico {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .tabela-dados {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 13px;
            max-height: 600px;
            overflow-y: auto;
            display: block;
        }}
        
        .tabela-dados thead {{
            display: table;
            width: 100%;
            background-color: #0070C0;
            color: white;
            position: sticky;
            top: 0;
        }}
        
        .tabela-dados tbody {{
            display: block;
            width: 100%;
            overflow-y: auto;
            max-height: 500px;
        }}
        
        .tabela-dados tr {{
            display: table;
            width: 100%;
            table-layout: fixed;
        }}
        
        .tabela-dados th {{
            background-color: #0070C0;
            color: white;
            padding: 10px;
            text-align: left;
            font-weight: 600;
        }}
        
        .tabela-dados td {{
            padding: 8px 10px;
            border-bottom: 1px solid #ddd;
        }}
        
        .tabela-dados tbody tr:hover {{
            background-color: #f9f9f9;
        }}
        
        .tabela-estatistica {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        
        .tabela-estatistica th {{
            background-color: #0070C0;
            color: white;
            padding: 10px;
            text-align: center;
            font-weight: 600;
        }}
        
        .tabela-estatistica td {{
            padding: 8px 10px;
            border: 1px solid #ddd;
            text-align: center;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            border-top: 1px solid #ddd;
            margin-top: 40px;
        }}
        
        h2 {{
            color: #0070C0;
            margin-top: 30px;
            margin-bottom: 20px;
            border-bottom: 2px solid #0070C0;
            padding-bottom: 10px;
        }}
        
        .grid-2 {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 20px;
        }}
        
        @media (max-width: 768px) {{
            .card-header {{
                flex-direction: column;
                gap: 10px;
            }}
            
            .download-button {{
                width: 100%;
                justify-content: center;
            }}
            
            .btn-voltar-topo {{
                bottom: 20px;
                right: 20px;
                width: 45px;
                height: 45px;
                font-size: 20px;
            }}
        }}
    </style>
</head>
<body>
    <!-- ===== BOTÃO VOLTAR AO TOPO ===== -->
    <button class="btn-voltar-topo" id="btnVoltarTopo" title="Voltar ao topo">
        ↑
    </button>
    <!-- ===== FIM DO BOTÃO ===== -->
    
    <div class="header">
        <img src="{img_src_base64}" alt="Logo AFPESP">
        <h1>Análise de HC - Indicadores de Gestão de Pessoas</h1>
        <p>Relatório de Ativos, Demissões e Admissões</p>
    </div>
    
        <div class="container">
        <!-- DASHBOARD -->
        <h2>Resumo Estratégico 2026</h2>
        {kpi_dashboard_html}
    </div>
</body>

    <div class="container">
        <div class="tabs">
            <button class="tab-button active" onclick="abrirAba(event, 'ativos')">👥 Colaboradores Ativos</button>
            <button class="tab-button" onclick="abrirAba(event, 'demissoes')">📉 Demissões 2026</button>
            <button class="tab-button" onclick="abrirAba(event, 'admissoes')">📈 Admissões 2026</button>
        </div>
        
        <!-- ABA 1: ATIVOS -->
        <div id="ativos" class="tab-content active">
            <div class="metrics">
                {card_ativo_html}
            </div>
            
            <div class="card">
                <h2>Estatísticas Descritivas - Variáveis Quantitativas</h2>
                {tabela_estat_ativo}
            </div>
            
            <div class="grid-2">
                <div class="grafico">{grafico_sexo_ativo}</div>
                <div class="grafico">{grafico_filhos_ativo}</div>
            </div>
            
            <div class="grid-2">
                <div class="grafico">{grafico_etnia_ativo}</div>
                <div class="grafico">{grafico_civil_ativo}</div>
            </div>
            
            <div class="grafico">{grafico_form_ativo}</div>
            <div class="grafico">{grafico_cargo_ativo}</div>
            <div class="grafico">{grafico_cc_ativo}</div>
            <div class="grafico">{grafico_unidade_ativo}</div>
            
            <div class="grafico">{boxplot_ativo_html}</div>
            
            <!-- ===== SEÇÃO DE FILTROS - ATIVOS ===== -->
            <div class="card">
                <h2>🔍 Filtros de Pesquisa</h2>
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px;">
                    
                    <!-- Filtro por Nome -->
                    <div>
                        <label for="filtroNome" style="font-weight: 600; display: block; margin-bottom: 5px;">Nome</label>
                        <input type="text" id="filtroNome" placeholder="Digite o nome..." 
                               style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                    </div>
                    
                    <!-- Filtro por Data de Admissão -->
                    <div>
                        <label for="filtroDataAdmissao" style="font-weight: 600; display: block; margin-bottom: 5px;">Data de Admissão (De)</label>
                        <input type="date" id="filtroDataAdmissao" 
                               style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                    </div>
                    
                    <!-- Filtro por Empresa -->
                    <div>
                        <label for="filtroEmpresa" style="font-weight: 600; display: block; margin-bottom: 5px;">Empresa</label>
                        <select id="filtroEmpresa" 
                                style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                            <option value="">-- Todas as Empresas --</option>
                        </select>
                    </div>
                    
                    <!-- Botão Limpar -->
                    <div style="display: flex; align-items: flex-end;">
                        <button onclick="limparFiltros()" 
                                style="width: 100%; padding: 8px 15px; background-color: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: 600;">
                            🔄 Limpar Filtros
                        </button>
                    </div>
                </div>
            </div>
            
            <!-- ===== TABELA COM FILTROS ===== -->
            <div class="card">
                <div class="card-header">
                    <h2>Tabela de Colaboradores Ativos</h2>
                    <div>
                        <button class="download-button" onclick="baixarExcel('{arquivo_ativos}')">
                            📥 Baixar Planilha Completa
                        </button>
                        <button class="download-button" style="background-color: #0070C0;
                        " onclick="exportarFiltradosExcel('ativos', 'HC_Ativos_Filtrados.xlsx')">
                            ⬇️ Baixar Dados Filtrados
                        </button>
                    </div>
                </div>
                <div style="overflow-x: auto;">
                    {tabela_ativos}
                </div>
                <p id="totalRegistros" style="margin-top: 10px; color: #666; font-size: 14px;"></p>
            </div>
        </div>
        
        <!-- ABA 2: DEMISSÕES -->
        <div id="demissoes" class="tab-content">
            <div class="metrics">
                {card_inativo_html}
            </div>
            
            <div class="grafico">{grafico_corrida_inativo}</div>
            <div class="grafico">{grafico_mes_inativo}</div>
            <div class="card">
                <h2>Comparativo de Turnover (Rotatividade)</h2>
                <p>Acompanhamento da taxa mensal (Demissões / Ativos) independente do crescimento do HC.</p>
                {grafico_turnover}
            </div>
            <div class="grafico">{grafico_tipo_inativo}</div>
            
            <!-- Dentro da variável html_final -->
            <div class="card">
                <h2>Qualidade da Contratação (Early Turnover)</h2>
                <p>Monitoramento de desligamentos ocorridos durante o período de experiência (90 dias).</p>
                
                <!-- 1. Indicador Macro (Centralizado) -->
                <div style="width: 100%; max-width: 600px; margin: 0 auto;">
                    {grafico_early_gauge}
                </div>
                
                <hr style="border: 0; border-top: 1px solid #eee; margin: 20px 0;">

                <!-- 2. Gráfico Temporal (Largura Total) -->
                <div style="width: 100%;">
                    {grafico_early_bar}
                </div>
                
                <p style="font-size: 12px; color: gray; margin-top: 10px;">
                    * <b>Early Turnover (<90 dias):</b> Indica possíveis falhas no Recrutamento (perfil) ou Onboarding (integração).<br>
                    * <b>Turnover Orgânico (>90 dias):</b> Indica questões de Gestão, Clima, Remuneração ou Mercado.
                </p>
            </div>

                <div class="card">
                    <h2>Análise de Sobrevivência (Tenure)</h2>
                    <p>Identificação do "Ponto de Ruptura": Em qual momento da jornada o colaborador decide sair?</p>
                    
                    <!-- Gráfico de Distribuição -->
                    <div style="margin-bottom: 30px;">
                        {grafico_tenure_hist}
                        <p style="font-size: 12px; color: gray; text-align: center;">
                            * Barras altas indicam o período crítico onde ocorrem mais desligamentos.
                        </p>
                    </div>

                    <hr style="border: 0; border-top: 1px solid #eee; margin: 20px 0;">

                    <!-- Gráfico de Boxplot -->
                    <div>
                        {grafico_tenure_box}
                        <p style="font-size: 12px; color: gray;">
                            * <b>Interpretação do Boxplot:</b> A linha dentro da caixa é a Mediana. A caixa representa 50% das pessoas. 
                            Pontos fora das linhas são "outliers" (casos isolados).
                        </p>
                    </div>
                </div>

                    <div class="card" style="border-left: 5px solid #b91c1c;"> <!-- Borda vermelha para chamar atenção -->
                        <h2>Custo da Rotatividade</h2>
                        <p>Estimativa do impacto financeiro direto e indireto gerado pelas demissões no ano corrente.</p>
                        
                        {grafico_financeiro}
                        
                        <div style="background-color: #fef2f2; padding: 15px; border-radius: 8px; margin-top: 10px;">
                            <p style="margin: 0; color: #991b1b; font-size: 14px;">
                                <b>💡 O que compõe este custo (Fator 1.5x)?</b><br>
                                Não é apenas a rescisão. Inclui: Multa do FGTS, custos de recrutamento (anúncios, tempo de triagem), exames admissionais, 
                                treinamento do novo colaborador e a <i>curva de aprendizado</i> (tempo até o novo atingir a produtividade do anterior).
                            </p>
                        </div>
                    </div>

            <!-- ===== SEÇÃO DE FILTROS - DEMISSÕES ===== -->
            <div class="card">
                <h2>🔍 Filtros de Pesquisa</h2>
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px;">
                    
                    <!-- Filtro por Nome -->
                    <div>
                        <label for="filtroNomeDemissoes" style="font-weight: 600; display: block; margin-bottom: 5px;">Nome</label>
                        <input type="text" id="filtroNomeDemissoes" placeholder="Digite o nome..." 
                               style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                    </div>
                    
                    <!-- Filtro por Data de Rescisão -->
                    <div>
                        <label for="filtroDataRescisao" style="font-weight: 600; display: block; margin-bottom: 5px;">Data de Rescisão (De)</label>
                        <input type="date" id="filtroDataRescisao" 
                               style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                    </div>
                    
                    <!-- Filtro por Empresa -->
                    <div>
                        <label for="filtroEmpresaDemissoes" style="font-weight: 600; display: block; margin-bottom: 5px;">Empresa</label>
                        <select id="filtroEmpresaDemissoes" 
                                style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                            <option value="">-- Todas as Empresas --</option>
                        </select>
                    </div>
                    
                    <!-- Filtro por Tipo de Rescisão -->
                    <div>
                        <label for="filtroTipoRescisao" style="font-weight: 600; display: block; margin-bottom: 5px;">Tipo de Rescisão</label>
                        <select id="filtroTipoRescisao" 
                                style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                            <option value="">-- Todos os Tipos --</option>
                        </select>
                    </div>
                    
                    <!-- Botão Limpar -->
                    <div style="display: flex; align-items: flex-end;">
                        <button onclick="limparFiltrosDemissoes()" 
                                style="width: 100%; padding: 8px 15px; background-color: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: 600;">
                            🔄 Limpar Filtros
                        </button>
                    </div>
                </div>
            </div>
            
            <div class="card">
                <div class="card-header">
                    <h2>Tabela de Demissões em 2026</h2>
                    <div>
                        <button class="download-button" onclick="baixarExcel('{arquivo_demissoes}')">
                            📥 Baixar Planilha Completa
                        </button>
                        <button class="download-button" style="background-color: #0070C0;" onclick="exportarFiltradosExcel('demissoes', 'HC_Demissoes_Filtrados.xlsx')">
                            ⬇️ Baixar Dados Filtrados
                        </button>
                    </div>
                </div>
                <div style="overflow-x: auto;">
                    {tabela_inativos}
                </div>
                <p id="totalRegistrosDemissoes" style="margin-top: 10px; color: #666; font-size: 14px;"></p>
            </div>
        </div>
        
        <!-- ABA 3: ADMISSÕES -->
        <div id="admissoes" class="tab-content">
            <div class="metrics">
                {card_admitido_html}
            </div>
            
            <div class="grafico">{grafico_corrida_admitido}</div>
            <div class="grafico">{grafico_mes_admitido}</div>
            
            <!-- ===== SEÇÃO DE FILTROS - ADMISSÕES ===== -->
            <div class="card">
                <h2>🔍 Filtros de Pesquisa</h2>
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px;">
                    
                    <!-- Filtro por Nome -->
                    <div>
                        <label for="filtroNomeAdmissoes" style="font-weight: 600; display: block; margin-bottom: 5px;">Nome</label>
                        <input type="text" id="filtroNomeAdmissoes" placeholder="Digite o nome..." 
                               style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                    </div>
                    
                    <!-- Filtro por Data de Admissão -->
                    <div>
                        <label for="filtroDataAdmissaoAdm" style="font-weight: 600; display: block; margin-bottom: 5px;">Data de Admissão (De)</label>
                        <input type="date" id="filtroDataAdmissaoAdm" 
                               style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                    </div>
                    
                    <!-- Filtro por Empresa -->
                    <div>
                        <label for="filtroEmpresaAdmissoes" style="font-weight: 600; display: block; margin-bottom: 5px;">Empresa</label>
                        <select id="filtroEmpresaAdmissoes" 
                                style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 5px;">
                            <option value="">-- Todas as Empresas --</option>
                        </select>
                    </div>
                    
                    <!-- Botão Limpar -->
                    <div style="display: flex; align-items: flex-end;">
                        <button onclick="limparFiltrosAdmissoes()" 
                                style="width: 100%; padding: 8px 15px; background-color: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: 600;">
                            🔄 Limpar Filtros
                        </button>
                    </div>
                </div>
            </div>
            
            <div class="card">
                <div class="card-header">
                    <h2>Tabela de Admissões em 2026</h2>
                    <div>
                        <button class="download-button" onclick="baixarExcel('{arquivo_admissoes}')">
                            📥 Baixar Planilha Completa
                        </button>
                        <button class="download-button" style="background-color: #0070C0;" onclick="exportarFiltradosExcel('admissoes', 'HC_Admissoes_Filtrados.xlsx')">
                            ⬇️ Baixar Dados Filtrados
                        </button>
                    </div>
                </div>
                <div style="overflow-x: auto;">
                    {tabela_admitidos}
                </div>
                <p id="totalRegistrosAdmissoes" style="margin-top: 10px; color: #666; font-size: 14px;"></p>
            </div>
        </div>
    </div>
    
    <div class="footer">
        <p>Relatório gerado automaticamente em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')} | Gestão de Pessoas - AFPESP</p>
    </div>
    
    <script>
        function abrirAba(evt, abaNome) {{
            var i, tabcontent, tabbuttons;
            tabcontent = document.getElementsByClassName("tab-content");
            for (i = 0; i < tabcontent.length; i++) {{
                tabcontent[i].classList.remove("active");
            }}
            tabbuttons = document.getElementsByClassName("tab-button");
            for (i = 0; i < tabbuttons.length; i++) {{
                tabbuttons[i].classList.remove("active");
            }}
            document.getElementById(abaNome).classList.add("active");
            evt.currentTarget.classList.add("active");
        }}
        
        function baixarExcel(nomeArquivo) {{
            const link = document.createElement('a');
            link.href = nomeArquivo;
            link.download = nomeArquivo;
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            alert('Download da planilha completa iniciado: ' + nomeArquivo);
        }}

        // Fallback para CSV caso o XLSX não carregue ou falhe
        function exportarFiltradosCSV(tabId, filename) {{
            const tabContent = document.getElementById(tabId);
            const tabela = tabContent.querySelector('.tabela-dados'); 
            if (!tabela) {{
                alert('Tabela não encontrada para download CSV.');
                return;
            }}

            const headers = Array.from(tabela.querySelectorAll('thead th')).map(th => th.textContent.trim());
            let csvContent = headers.map(header => `"${{header.replace(/"/g, '""')}}"`.replace(/\\n/g, ' ')).join(',') + '\\n';

            const linhasVisiveis = Array.from(tabela.querySelectorAll('tbody tr')).filter(tr => tr.style.display !== 'none');

            if (linhasVisiveis.length === 0) {{
                alert('Nenhum registro visível para exportar em CSV.');
                return;
            }}

            linhasVisiveis.forEach(linha => {{
                const rowData = Array.from(linha.querySelectorAll('td')).map(td => {{
                    let text = td.textContent.trim();
                    // Escape double quotes and wrap in double quotes if it contains comma or double quotes
                    if (text.includes(',') || text.includes('"') || text.includes('\\n')) {{
                        text = `"${{text.replace(/"/g, '""')}}"`;
                    }}
                    return text;
                }});
                csvContent += rowData.join(',') + '\\n';
            }});

            const blob = new Blob([csvContent], {{ type: 'text/csv;charset=utf-8;' }});
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = filename;
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            alert(`Download dos dados filtrados iniciado: ${{filename}} (${{linhasVisiveis.length}} registros)`);
        }}

        // Função para exportar dados filtrados para Excel com formatação
function exportarFiltradosExcel(tabId, filename) {{
    const colunas_data = ['nascimento', 'data_admissao', 'data_rescisao', 'data_nasc_conjuge'];
    let retryCount = 0;
    const maxRetries = 10;
    const retryDelay = 200;

    function doExport() {{
        if (typeof XLSX === 'undefined') {{
            if (retryCount < maxRetries) {{
                console.warn(`XLSX is not defined yet. Retrying in ${{retryDelay}}ms... (Attempt ${{retryCount + 1}}/${{maxRetries}})`);
                retryCount++;
                setTimeout(doExport, retryDelay);
            }} else {{
                console.error('XLSX library failed to load after multiple retries. Falling back to CSV export.');
                alert('A biblioteca de exportação Excel não carregou após várias tentativas. Exportando em formato CSV.');
                exportarFiltradosCSV(tabId, filename.replace('.xlsx', '.csv'));
            }}
            return;
        }}

        try {{
            const tabContent = document.getElementById(tabId);
            const tabela = tabContent.querySelector('.tabela-dados'); 
            if (!tabela) {{
                alert('Tabela não encontrada para download Excel.');
                return;
            }}

            const headers = Array.from(tabela.querySelectorAll('thead th')).map(th => th.textContent.trim());
            const linhasVisiveis = Array.from(tabela.querySelectorAll('tbody tr')).filter(tr => tr.style.display !== 'none');

            if (linhasVisiveis.length === 0) {{
                alert('Nenhum registro visível para exportar em Excel.');
                return;
            }}

            console.log('Total de linhas visíveis:', linhasVisiveis.length);
            console.log('Headers encontrados:', headers);

            const wb = XLSX.utils.book_new();
            const ws = XLSX.utils.aoa_to_sheet([]);

            // PASSO 1: Adicionar cabeçalhos
            XLSX.utils.sheet_add_aoa(ws, [headers], {{ origin: 'A1' }});

            // PASSO 2: Preparar dados
            const dataRows = [];
            linhasVisiveis.forEach(function(linha) {{
                const cells = Array.from(linha.querySelectorAll('td'));
                const row = [];
                cells.forEach(function(td) {{
                    row.push(td.textContent.trim());
                }});
                dataRows.push(row);
            }});

            // PASSO 3: Adicionar dados ao worksheet
            XLSX.utils.sheet_add_aoa(ws, dataRows, {{ origin: 'A2' }});

            console.log('Dados adicionados ao worksheet:', dataRows.length, 'linhas');

            // PASSO 4: Definir estilos
            const headerStyle = {{
                fill: {{ fgColor: {{ rgb: "005A9C" }} }},
                font: {{ bold: true, color: {{ rgb: "FFFFFF" }}, sz: 11 }},
                alignment: {{ horizontal: "center", vertical: "center", wrapText: true }},
                border: {{
                    top: {{ style: "thin", color: {{ rgb: "FFFFFF" }} }},
                    bottom: {{ style: "thin", color: {{ rgb: "FFFFFF" }} }},
                    left: {{ style: "thin", color: {{ rgb: "FFFFFF" }} }},
                    right: {{ style: "thin", color: {{ rgb: "FFFFFF" }} }}
                }}
            }};

            const dataStyle = {{
                fill: {{ fgColor: {{ rgb: "FFFFFF" }} }},
                font: {{ sz: 10, color: {{ rgb: "333333" }} }},
                alignment: {{ horizontal: "left", vertical: "center", wrapText: true }},
                border: {{
                    top: {{ style: "thin", color: {{ rgb: "DDDDDD" }} }},
                    bottom: {{ style: "thin", color: {{ rgb: "DDDDDD" }} }},
                    left: {{ style: "thin", color: {{ rgb: "DDDDDD" }} }},
                    right: {{ style: "thin", color: {{ rgb: "DDDDDD" }} }}
                }}
            }};

            // PASSO 5: Aplicar estilos ao cabeçalho
            for (let C = 0; C < headers.length; ++C) {{
                const cell_address = XLSX.utils.encode_cell({{ r: 0, c: C }});
                if (!ws[cell_address]) ws[cell_address] = {{}};
                ws[cell_address].s = headerStyle;
            }}

            // PASSO 6: Aplicar estilos aos dados
            for (let R = 0; R < dataRows.length; ++R) {{
                for (let C = 0; C < headers.length; ++C) {{
                    const cell_address = XLSX.utils.encode_cell({{ r: R + 1, c: C }});
                    if (!ws[cell_address]) ws[cell_address] = {{}};
                    const cell = ws[cell_address];

                    cell.s = dataStyle;

                    const cellValue = cell.v;
                    if (typeof cellValue === 'string') {{
                        const dateMatch = cellValue.match(/^(\d{{2}})\/(\d{{2}})\/(\d{{4}})$/);
                        if (dateMatch) {{
                            cell.t = 's';
                            cell.z = '@';
                        }} else {{
                            const numValue = parseFloat(cellValue.replace(/\./g, '').replace(/,/g, '.'));
                            if (!isNaN(numValue) && isFinite(numValue)) {{
                                cell.t = 'n';
                                cell.v = numValue;
                                if (cellValue.includes(',') || cellValue.includes('.')) {{
                                    cell.z = '#,##0.00';
                                }} else {{
                                    cell.z = '#,##0';
                                }}
                            }}
                        }}
                    }} else if (typeof cellValue === 'number') {{
                        cell.t = 'n';
                        if (cellValue % 1 !== 0) {{
                            cell.z = '#,##0.00';
                        }} else {{
                            cell.z = '#,##0';
                        }}
                    }}
                }}
            }}

            // PASSO 7: Calcular largura das colunas
            const colWidths = [];
            for (let C = 0; C < headers.length; ++C) {{
                let maxLength = headers[C].length;
                for (let R = 0; R < dataRows.length; ++R) {{
                    const cell_address = XLSX.utils.encode_cell({{ r: R + 1, c: C }});
                    const cell = ws[cell_address];
                    if (cell && cell.v !== undefined && cell.v !== null) {{
                        maxLength = Math.max(maxLength, String(cell.v).length);
                    }}
                }}
                colWidths.push({{ wch: Math.min(Math.max(maxLength + 2, 12), 50) }});
            }}
            ws['!cols'] = colWidths;

            // PASSO 8: Congelar painel
            ws['!freeze'] = {{ xSplit: 0, ySplit: 1, topLeftCell: 'A2', activePane: 'bottomLeft', state: 'frozen' }};

            // PASSO 9: Adicionar worksheet ao workbook e salvar
            XLSX.utils.book_append_sheet(wb, ws, "Dados Filtrados");
            XLSX.writeFile(wb, filename);
            alert(`Download dos dados filtrados iniciado: ${{filename}} (${{linhasVisiveis.length}} registros)`);

        }} catch (error) {{
            console.error('Erro ao exportar dados para Excel:', error);
            alert('Erro ao exportar dados para Excel. Tentando exportar em CSV. Erro: ' + error.message);
            exportarFiltradosCSV(tabId, filename.replace('.xlsx', '.csv'));
        }}
    }}
    doExport();
}}
        
        // ===== FUNCIONALIDADE BOTÃO VOLTAR AO TOPO =====
        const btnVoltarTopo = document.getElementById('btnVoltarTopo');
        
        window.addEventListener('scroll', function() {{
            if (window.pageYOffset > 300) {{
                btnVoltarTopo.classList.add('visible');
            }} else {{
                btnVoltarTopo.classList.remove('visible');
            }}
        }});
        
        btnVoltarTopo.addEventListener('click', function() {{
            window.scrollTo({{
                top: 0,
                behavior: 'smooth'
            }});
        }});
        // ===== FIM DA FUNCIONALIDADE =====
        
        // ===== FUNÇÃO DE FILTROS - ATIVOS =====
        function inicializarFiltros() {{
            const tabela = document.querySelector('#ativos .tabela-dados');
            if (!tabela) return;
            
            const linhas = Array.from(tabela.querySelectorAll('tbody tr'));
            
            const empresas = new Set();
            linhas.forEach(linha => {{
                const empresa = linha.querySelector('td:nth-child(13)');
                if (empresa) empresas.add(empresa.textContent.trim());
            }});
            
            const selectEmpresa = document.getElementById('filtroEmpresa');
            Array.from(empresas).sort().forEach(empresa => {{
                if (empresa) {{
                    const option = document.createElement('option');
                    option.value = empresa;
                    option.textContent = empresa;
                    selectEmpresa.appendChild(option);
                }}
            }});
            
            document.getElementById('filtroNome').addEventListener('keyup', aplicarFiltros);
            document.getElementById('filtroDataAdmissao').addEventListener('change', aplicarFiltros);
            document.getElementById('filtroEmpresa').addEventListener('change', aplicarFiltros);
            aplicarFiltros(); // Aplicar filtros iniciais para mostrar a contagem
        }}
        
        function aplicarFiltros() {{
            const tabela = document.querySelector('#ativos .tabela-dados');
            if (!tabela) return;
            
            const linhas = tabela.querySelectorAll('tbody tr');
            
            const nome = document.getElementById('filtroNome').value.toUpperCase();
            const data = document.getElementById('filtroDataAdmissao').value;
            const empresa = document.getElementById('filtroEmpresa').value;
            
            let visiveisCount = 0;
            
            linhas.forEach(linha => {{
                const colunaNome = linha.querySelector('td:nth-child(2)');
                const colunaData = linha.querySelector('td:nth-child(6)');
                const colunaEmpresa = linha.querySelector('td:nth-child(13)');
                
                let mostrar = true;
                
                if (nome && colunaNome && !colunaNome.textContent.toUpperCase().includes(nome)) {{
                    mostrar = false;
                }}
                
                if (data && colunaData) {{
                    const dataLinha = colunaData.textContent.trim();
                    const [dia, mes, ano] = dataLinha.split('/');
                    const dataLinhaObj = new Date(ano, mes - 1, dia);
                    const dataFiltro = new Date(data);
                    
                    if (dataLinhaObj < dataFiltro) {{
                        mostrar = false;
                    }}
                }}
                
                if (empresa && colunaEmpresa && colunaEmpresa.textContent.trim() !== empresa) {{
                    mostrar = false;
                }}
                
                linha.style.display = mostrar ? 'table-row' : 'none';
                if (mostrar) visiveisCount++;
            }});
            
            document.getElementById('totalRegistros').textContent = 
                `📊 Exibindo ${{visiveisCount}} de ${{linhas.length}} registros`;
        }}
        
        function limparFiltros() {{
            document.getElementById('filtroNome').value = '';
            document.getElementById('filtroDataAdmissao').value = '';
            document.getElementById('filtroEmpresa').value = '';
            aplicarFiltros();
        }}
        
        // ===== FUNÇÃO DE FILTROS - DEMISSÕES =====
        function inicializarFiltrosDemissoes() {{
            const tabelaDemissoes = document.querySelector('#demissoes .tabela-dados');
            if (!tabelaDemissoes) return;
            
            const linhas = Array.from(tabelaDemissoes.querySelectorAll('tbody tr'));
            
            const empresas = new Set();
            const tipos = new Set();
            linhas.forEach(linha => {{
                const empresa = linha.querySelector('td:nth-child(12)');
                const tipo = linha.querySelector('td:nth-child(9)');
                if (empresa) empresas.add(empresa.textContent.trim());
                if (tipo) tipos.add(tipo.textContent.trim());
            }});
            
            const selectEmpresa = document.getElementById('filtroEmpresaDemissoes');
            Array.from(empresas).sort().forEach(empresa => {{
                if (empresa) {{
                    const option = document.createElement('option');
                    option.value = empresa;
                    option.textContent = empresa;
                    selectEmpresa.appendChild(option);
                }}
            }});
            
            const selectTipo = document.getElementById('filtroTipoRescisao');
            Array.from(tipos).sort().forEach(tipo => {{
                if (tipo) {{
                    const option = document.createElement('option');
                    option.value = tipo;
                    option.textContent = tipo;
                    selectTipo.appendChild(option);
                }}
            }});
            
            document.getElementById('filtroNomeDemissoes').addEventListener('keyup', aplicarFiltrosDemissoes);
            document.getElementById('filtroDataRescisao').addEventListener('change', aplicarFiltrosDemissoes);
            document.getElementById('filtroEmpresaDemissoes').addEventListener('change', aplicarFiltrosDemissoes);
            document.getElementById('filtroTipoRescisao').addEventListener('change', aplicarFiltrosDemissoes);
            aplicarFiltrosDemissoes(); // Aplicar filtros iniciais para mostrar a contagem
        }}
        
        function aplicarFiltrosDemissoes() {{
            const tabelaDemissoes = document.querySelector('#demissoes .tabela-dados');
            if (!tabelaDemissoes) return;
            
            const linhas = tabelaDemissoes.querySelectorAll('tbody tr');
            
            const nome = document.getElementById('filtroNomeDemissoes').value.toUpperCase();
            const data = document.getElementById('filtroDataRescisao').value;
            const empresa = document.getElementById('filtroEmpresaDemissoes').value;
            const tipo = document.getElementById('filtroTipoRescisao').value;
            
            let visiveisCount = 0;
            
            linhas.forEach(linha => {{
                const colunaNome = linha.querySelector('td:nth-child(2)');
                const colunaData = linha.querySelector('td:nth-child(7)');
                const colunaEmpresa = linha.querySelector('td:nth-child(12)');
                const colunaTipo = linha.querySelector('td:nth-child(9)');
                
                let mostrar = true;
                
                if (nome && colunaNome && !colunaNome.textContent.toUpperCase().includes(nome)) {{
                    mostrar = false;
                }}
                
                if (data && colunaData) {{
                    const dataLinha = colunaData.textContent.trim();
                    const [dia, mes, ano] = dataLinha.split('/');
                    const dataLinhaObj = new Date(ano, mes - 1, dia);
                    const dataFiltro = new Date(data);
                    
                    if (dataLinhaObj < dataFiltro) {{
                        mostrar = false;
                    }}
                }}
                
                if (empresa && colunaEmpresa && colunaEmpresa.textContent.trim() !== empresa) {{
                    mostrar = false;
                }}
                
                if (tipo && colunaTipo && colunaTipo.textContent.trim() !== tipo) {{
                    mostrar = false;
                }}
                
                linha.style.display = mostrar ? 'table-row' : 'none';
                if (mostrar) visiveisCount++;
            }});
            
            document.getElementById('totalRegistrosDemissoes').textContent = 
                `📊 Exibindo ${{visiveisCount}} de ${{linhas.length}} registros`;
        }}
        
        function limparFiltrosDemissoes() {{
            document.getElementById('filtroNomeDemissoes').value = '';
            document.getElementById('filtroDataRescisao').value = '';
            document.getElementById('filtroEmpresaDemissoes').value = '';
            document.getElementById('filtroTipoRescisao').value = '';
            aplicarFiltrosDemissoes();
        }}
        
        // ===== FUNÇÃO DE FILTROS - ADMISSÕES =====
        function inicializarFiltrosAdmissoes() {{
            const tabelaAdmissoes = document.querySelector('#admissoes .tabela-dados');
            if (!tabelaAdmissoes) return;
            
            const linhas = Array.from(tabelaAdmissoes.querySelectorAll('tbody tr'));
            
            const empresas = new Set();
            linhas.forEach(linha => {{
                const empresa = linha.querySelector('td:nth-child(15)');
                if (empresa) empresas.add(empresa.textContent.trim());
            }});
            
            const selectEmpresa = document.getElementById('filtroEmpresaAdmissoes');
            Array.from(empresas).sort().forEach(empresa => {{
                if (empresa) {{
                    const option = document.createElement('option');
                    option.value = empresa;
                    option.textContent = empresa;
                    selectEmpresa.appendChild(option);
                }}
            }});
            
            document.getElementById('filtroNomeAdmissoes').addEventListener('keyup', aplicarFiltrosAdmissoes);
            document.getElementById('filtroDataAdmissaoAdm').addEventListener('change', aplicarFiltrosAdmissoes);
            document.getElementById('filtroEmpresaAdmissoes').addEventListener('change', aplicarFiltrosAdmissoes);
            aplicarFiltrosAdmissoes(); // Aplicar filtros iniciais para mostrar a contagem
        }}
        
        function aplicarFiltrosAdmissoes() {{
            const tabelaAdmissoes = document.querySelector('#admissoes .tabela-dados');
            if (!tabelaAdmissoes) return;
            
            const linhas = tabelaAdmissoes.querySelectorAll('tbody tr');
            
            const nome = document.getElementById('filtroNomeAdmissoes').value.toUpperCase();
            const data = document.getElementById('filtroDataAdmissaoAdm').value;
            const empresa = document.getElementById('filtroEmpresaAdmissoes').value;
            
            let visiveisCount = 0;
            
            linhas.forEach(linha => {{
                const colunaNome = linha.querySelector('td:nth-child(2)');
                const colunaData = linha.querySelector('td:nth-child(6)');
                const colunaEmpresa = linha.querySelector('td:nth-child(15)');
                
                let mostrar = true;
                
                if (nome && colunaNome && !colunaNome.textContent.toUpperCase().includes(nome)) {{
                    mostrar = false;
                }}
                
                if (data && colunaData) {{
                    const dataLinha = colunaData.textContent.trim();
                    const [dia, mes, ano] = dataLinha.split('/');
                    const dataLinhaObj = new Date(ano, mes - 1, dia);
                    const dataFiltro = new Date(data);
                    
                    if (dataLinhaObj < dataFiltro) {{
                        mostrar = false;
                    }}
                }}
                
                if (empresa && colunaEmpresa && colunaEmpresa.textContent.trim() !== empresa) {{
                    mostrar = false;
                }}
                
                linha.style.display = mostrar ? 'table-row' : 'none';
                if (mostrar) visiveisCount++;
            }});
            
            document.getElementById('totalRegistrosAdmissoes').textContent = 
                `📊 Exibindo ${{visiveisCount}} de ${{linhas.length}} registros`;
        }}
        
        function limparFiltrosAdmissoes() {{
            document.getElementById('filtroNomeAdmissoes').value = '';
            document.getElementById('filtroDataAdmissaoAdm').value = '';
            document.getElementById('filtroEmpresaAdmissoes').value = '';
            aplicarFiltrosAdmissoes();
        }}
        
        // Inicializar todos os filtros ao carregar a página
        window.addEventListener('load', function() {{
            inicializarFiltros();
            inicializarFiltrosDemissoes();
            inicializarFiltrosAdmissoes();
        }});
    </script>
</body>
</html>
"""

    # Salvar HTML
    caminho_html = 'Relatório_HC.html'
    with open(caminho_html, 'w', encoding='utf-8') as f:
        f.write(html_final)

    logging.info(f"✓ Relatório HTML gerado: {caminho_html}")

    # Abrir no navegador
    webbrowser.open('file://' + os.path.abspath(caminho_html))
    logging.info("✓ Relatório aberto no navegador.\n")

except Exception as e:
    logging.error(f"ERRO ao gerar HTML na SEÇÃO 5: {e}")
    traceback.print_exc()
    sys.exit("Script encerrado.")

logging.info("="*80)
logging.info("✓ SCRIPT CONCLUÍDO COM SUCESSO!")
logging.info("="*80)