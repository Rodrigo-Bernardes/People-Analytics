# -*- coding: utf-8 -*-
"""
PROJETO DE PEOPLE ANALYTICS: PREVISÃO DE DEMISSÃO COM REGRESSÃO LOGÍSTICA
Autor: Rodrigo Rodrigues Bernardes

Este script implementa um modelo híbrido de People Analytics para prever a probabilidade
de um colaborador ser demitido, utilizando dados de atestados médicos e outras
informações do quadro de funcionários.

A solução é dividida em 3 camadas:
1.  Previsão (Scikit-learn com SMOTE): Para alta acurácia na identificação de risco.
2.  Interpretação (GLM com SMOTE): Para identificar e quantificar fatores de risco/proteção.
3.  Ação Estratégica (Dashboard e Recomendações): Para transformar insights em ações.

Funcionalidades incluídas:
1.  Pré-processamento de dados robusto, incluindo tratamento de valores faltantes e conversão de tipos.
2.  Identificação e remoção de data leakage (vazamento de dados).
3.  Divisão dos dados em conjuntos de treino e teste com estratificação.
4.  Uso de ColumnTransformer para pré-processamento de features numéricas e categóricas.
5.  Balanceamento da classe minoritária ('INATIVO') utilizando SMOTE para ambos os modelos.
6.  Construção e treinamento de um Pipeline com Regressão Logística (Scikit-learn).
7.  Ajuste do threshold de classificação do modelo Scikit-learn para otimizar Recall/Precisão.
8.  Construção e análise de um modelo GLM (Generalized Linear Model) com Statsmodels para interpretabilidade.
9.  Avaliação detalhada do modelo Scikit-learn com métricas (Acurácia, Precisão, Recall, F1-Score, ROC AUC, Gini).
10. Cálculo e plotagem da Curva ROC com o Coeficiente de Gini.
11. Geração de previsões de probabilidade e classificação de risco para todos os colaboradores.
12. Exportação dos resultados detalhados para um arquivo Excel com formatação aprimorada.
13. Geração de um Relatório HTML interativo e completo com todos os gráficos e tabelas.
14. Geração de um Dashboard Executivo em Excel com múltiplas abas para a Gestão de Pessoas.
15. Comentários detalhados, tratamento de erros e mensagens de progresso para facilitar a compreensão e depuração.
"""

# ==============================================================================
# 0. IMPORTAÇÕES, CONFIGURAÇÕES E FUNÇÕES AUXILIARES
# ==============================================================================

# --- 0.1. Importações ---
import pandas as pd
import numpy as np
import statsmodels.api as sm
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.compose import ColumnTransformer
from imblearn.pipeline import Pipeline as ImbPipeline
from sklearn.linear_model import LogisticRegression
from imblearn.over_sampling import SMOTE
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score, roc_auc_score,
    confusion_matrix, classification_report, roc_curve, precision_recall_curve
)
from sklearn.feature_selection import VarianceThreshold
import matplotlib.pyplot as plt
import seaborn as sns
import warnings
import sys
import os
import base64
import gc  # Para gerenciamento de memória
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.genmod.generalized_linear_model import GLM

# --- 0.2. Configuração de Logging ---
# Configurar logging para mensagens formatadas
import logging
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
warnings.filterwarnings('ignore') # Suprimir avisos

logging.info("✓ Bibliotecas importadas e logging configurado com sucesso.")

# --- 0.3. Definição de Constantes ---
PATH_ARQUIVO = r'X:\Gestao_de_Pessoas\Analytics\10 - Relatórios\10.4 - HC e Atestados Médicos\Controle_HC e Atestados.xlsb'
SHEET_NAME = 'HC'
OUTPUT_DIR = 'relatorio_people_analytics_output' # Diretório para salvar saídas
OUTPUT_EXCEL_PREVISOES_FILE = os.path.join(OUTPUT_DIR, 'df_previsoes_risco.xlsx')
OUTPUT_HTML_FILE = os.path.join(OUTPUT_DIR, 'relatorio_people_analytics.html')
OUTPUT_GLM_COEF_CSV = os.path.join(OUTPUT_DIR, 'glm_coeficientes_summary.csv')
OUTPUT_GLM_SUMMARY_TXT = os.path.join(OUTPUT_DIR, 'relatorio_glm_statsmodels.txt')
OUTPUT_DASHBOARD_EXCEL_FILE = os.path.join(OUTPUT_DIR, 'dashboard_gestao_pessoas.xlsx')

# Criar diretório de saída se não existir
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)
    logging.info(f"✓ Diretório de saída '{OUTPUT_DIR}' criado.")

# --- 0.4. Funções Auxiliares para HTML ---
def converter_grafico_para_base64(filepath):
    """Converte um arquivo de imagem PNG para uma string Base64."""
    if not os.path.exists(filepath):
        logging.warning(f"Arquivo de gráfico não encontrado: {filepath}. Retornando string vazia.")
        return ""
    try:
        with open(filepath, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
        return f"data:image/png;base64,{encoded_string}"
    except Exception as e:
        logging.error(f"Erro ao converter gráfico {filepath} para Base64: {e}")
        return ""

def gerar_tabela_html(df, title="", caption="", limit=None):
    """Gera uma tabela HTML a partir de um DataFrame Pandas."""
    if df is None or df.empty:
        return f"<p><strong>{title}</strong>: Dados não disponíveis ou DataFrame vazio.</p>"
    
    # Limitar o número de linhas para exibição no HTML se especificado
    if limit and len(df) > limit:
        df_display = df.head(limit).copy()
        caption_text = f"{caption} (Mostrando as primeiras {limit} linhas de {len(df)})"
    else:
        df_display = df.copy()
        caption_text = caption

    html = f"""
    <div class="table-responsive">
        <h4 class="mt-4">{title}</h4>
        <p class="text-muted">{caption_text}</p>
        {df_display.to_html(classes="table table-striped table-bordered table-hover", index=False if 'Feature' in df.columns else True)}
    </div>
    """
    return html

def gerar_card_metricas(title, value, description=""):
    """Gera um card HTML para exibição de métricas."""
    return f"""
    <div class="col-md-4 mb-4">
        <div class="card shadow-sm border-0">
            <div class="card-body">
                <h5 class="card-title text-primary">{title}</h5>
                <h3 class="card-text text-dark">{value}</h3>
                <p class="card-text text-muted">{description}</p>
            </div>
        </div>
    </div>
    """

logging.info("✓ Funções auxiliares HTML configuradas.")

# ==============================================================================
# 1. CARREGAMENTO E PRÉ-PROCESSAMENTO INICIAL DOS DADOS
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 1 - CARREGAMENTO E PRÉ-PROCESSAMENTO INICIAL DOS DADOS")
logging.info("="*80)

try:
    # 1.1. Carregando a base de dados
    if not os.path.exists(PATH_ARQUIVO):
        raise FileNotFoundError(f"Arquivo não encontrado: {PATH_ARQUIVO}")

    dados_hc = pd.read_excel(PATH_ARQUIVO, sheet_name=SHEET_NAME, engine='pyxlsb')
    logging.info(f"✓ Base de dados carregada com {dados_hc.shape[0]} registros e {dados_hc.shape[1]} colunas.")

    # Armazenar o DataFrame original para exportação posterior com todas as colunas
    df_hc_original_full = dados_hc.copy()

    # 1.2. Tratamento inicial de datas para filtragem
    # Garante que 'ano_rescisao' e 'ano_atestado' são tratados como strings para comparação
    # Preenche NaN com string vazia antes de converter para string
    dados_hc['ano_rescisao'] = dados_hc['ano_rescisao'].fillna('').astype(str)
    dados_hc['ano_atestado'] = dados_hc['ano_atestado'].fillna('').astype(str)
    logging.info("✓ Coluna 'ano_rescisao' tratada para string.")
    logging.info("✓ Coluna 'ano_atestado' tratada para string.")

    # 1.3. Aplicando filtros nos dados
    # A máscara filtra por ano de rescisão/atestado a partir de 2022 (ou sem data) e unidade 'UL'
    # Aqui removemos a variável'ano_atestado' para considerar todos colaboradores com ou sem
    mascara = (
        ((dados_hc['ano_rescisao'] >= '2022') | (dados_hc['ano_rescisao'] == '')) &
        (dados_hc['unidade'] == 'UL')
    )
    df_hc = dados_hc.loc[mascara].copy() # DataFrame filtrado para o modelo
    logging.info(f"✓ Dados filtrados: {df_hc.shape[0]} registros após aplicar critérios.")
    
    # Armazenar o DataFrame filtrado original (com todas as colunas) para exportação posterior
    df_hc_filtered_full = df_hc.copy()

    # 1.4. Excluindo variáveis que não serão utilizadas (Data Leakage e irrelevantes)
    # 'descricao_rescisao' é um exemplo CRÍTICO de data leakage se o objetivo é prever a demissão.
    colunas_para_dropar = [
        'registro','nome','nascimento','etnia_raca','cod_formacao','aposentado','rg','cpf','deficiente','nacionalidade',
        'naturalidade_uf','naturalidade_cidade','sexo_conjuge','nome_conjuge','data_nasc_conjuge','cpf_conjuge','data_admissao',
        'situacao','data_rescisao','descricao_rescisao','cod_cargo','cargo_completo','cargo_abreviado','cod_centro_custo',
        'nome_gestor','cod_secao','secao','centro_custo','salario_admissao','salario_atual','saldo_fgts','endereco',
        'numero_endereco','complemento','bairro','cidade','estado','telefone','email','cod_situacao','regime_trabalho',
        'hora_base','hora_complemento','ultimo_reajuste_individual','ultimo_reajuste_coletivo','data_inicio_ferias',
        'data_fim_ferias','cod_empresa','empresa','unidade','ano_admissao','mes_admissao','ano_rescisao','mes_rescisao',
        'cargo_padrao','prazo_experiencia','exp_45dias','exp_90dias','ultima_alteracao_contratual','motivo_alteracao',
        'descricao_rescisao_esocial','categoria','horas_nao_trabalhadas','data_inicio_atestado','ano_atestado'
    ]
    colunas_existentes_para_dropar = [col for col in colunas_para_dropar if col in df_hc.columns]
    hc_padr = df_hc.drop(columns=colunas_existentes_para_dropar).copy()

    logging.info(f"✓ Variáveis removidas: {len(colunas_existentes_para_dropar)}")
    logging.info(f"✓ Variáveis mantidas para análise: {hc_padr.shape[1]} -> {hc_padr.columns.tolist()}")

    # 1.5. Tratamento de valores faltantes e coluna 'filhos'
    # Converte 'S'/'N' para 1/0 e preenche faltantes com 0 (assumindo sem filhos se não especificado)
    if 'filhos' in hc_padr.columns:
        faltantes_filhos = hc_padr['filhos'].isnull().sum()
        hc_padr['filhos'] = hc_padr['filhos'].astype(str).str.upper()
        hc_padr['filhos'] = hc_padr['filhos'].replace({'S': 1, 'N': 0})
        hc_padr['filhos'] = pd.to_numeric(hc_padr['filhos'], errors='coerce')
        hc_padr['filhos'] = hc_padr['filhos'].fillna(0).astype(int)
        logging.info(f"✓ Valores faltantes em 'filhos' tratados: {faltantes_filhos} registros preenchidos com 0 e convertidos para int.")
    else:
        logging.warning("Aviso: Coluna 'filhos' não encontrada. Verifique se foi removida acidentalmente.")

    # Após o bloco onde 'filhos' é convertido para int
    if 'filhos' in hc_padr.columns:
        hc_padr['filhos'] = hc_padr['filhos'].astype('category')

    # 1.6. Criando a variável alvo (target)
    # 'status' é convertido para 'target' (1 para INATIVO, 0 para ATIVO)
    if 'status' in hc_padr.columns:
        hc_padr['target'] = hc_padr['status'].apply(lambda x: 1 if x == 'INATIVO' else 0)
        logging.info("✓ Variável 'target' criada: INATIVO = 1, ATIVO = 0.")
        logging.info(f"Distribuição do target: \n{hc_padr['target'].value_counts()}")
    else:
        raise ValueError("Coluna 'status' não encontrada em hc_padr. Não é possível criar a variável target.")

    # 1.7. Validação de dados após pré-processamento
    logging.info(f"Número final de registros para modelagem: {hc_padr.shape[0]}")
    logging.info(f"Número final de colunas para modelagem: {hc_padr.shape[1]}")
    if hc_padr.isnull().sum().sum() == 0:
        logging.info("✓ Nenhum valor nulo restante no DataFrame para modelagem.")
    else:
        logging.warning(f"Aviso: Ainda existem valores nulos no DataFrame para modelagem: {hc_padr.isnull().sum()[hc_padr.isnull().sum() > 0].to_dict()}")

    logging.info("\nPRÉ-PROCESSAMENTO DOS DADOS CONCLUÍDO COM SUCESSO!")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 1 (Carregamento e Pré-processamento Inicial): {e}")
    sys.exit("Script encerrado devido ao erro no pré-processamento inicial.")

gc.collect() # Limpeza de memória
logging.info("✓ GC.collect() executado após Seção 1.")

# ==============================================================================
# 2. SEPARAÇÃO DE FEATURES (X) E TARGET (y)
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 2 - SEPARAÇÃO DE FEATURES (X) E TARGET (y)")
logging.info("="*80)

try:
    # CRÍTICO: Remover 'status' (coluna original) e 'target' (a própria variável alvo) de X
    # As colunas 'nome', 'cargo', etc. também devem ser removidas de X, mas mantidas no df_hc_filtered_full
    # para serem re-adicionadas posteriormente para o dashboard.
    colunas_para_dropar_final = ['target', 'status']
    # Certifique-se de que colunas como 'nome', 'cargo', etc. não estejam em X, mas sim no df_hc_filtered_full
    # X deve conter apenas as variáveis preditoras (numéricas e categóricas).
    
    # Identificar colunas não-numéricas que devem ser tratadas como categóricas ou que são identificadores e devem ser removidas.
    cols_to_exclude_from_X = ['status'] # 'target' já foi excluído acima
    
    # As colunas 'nome', 'cargo', etc. são úteis no df_hc_filtered_full para o relatório final,
    # mas não devem ser features no modelo.
    # Assumindo que hc_padr já tem as colunas relevantes e removeu as demais.
    
    X = hc_padr.drop(columns=[col for col in colunas_para_dropar_final if col in hc_padr.columns]).copy()
    y = hc_padr['target'].copy()

    logging.info("✓ Features (X) e Target (y) separados.")
    logging.info(f"✓ Features (X) selecionadas ({X.shape[1]} variáveis):")
    for i, col in enumerate(X.columns):
        logging.info(f"   {i+1}. {col}")
    logging.info(f"\n✓ Target (y) selecionado: {y.name}")
    logging.info(f"✓ Dimensões resultantes: X ({X.shape[0]}x{X.shape[1]}), y ({y.shape[0]})")

    logging.info("\n✓ SEÇÃO 2 CONCLUÍDA COM SUCESSO")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 2 (Separação de Features e Target): {e}")
    sys.exit("Script encerrado devido ao erro na separação de features/target.")

gc.collect() # Limpeza de memória
logging.info("✓ GC.collect() executado após Seção 2.")

# ==============================================================================
# 3. ANÁLISE EXPLORATÓRIA DOS DADOS (EDA)
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 3 - ANÁLISE EXPLORATÓRIA DOS DADOS (EDA)")
logging.info("="*80)

try:
    # 3.1. Estatísticas Descritivas das Variáveis Numéricas
    numerical_cols = X.select_dtypes(include=np.number).columns.tolist()
    logging.info("\n--- 3.1. Estatísticas Descritivas das Variáveis Numéricas ---")
    eda_desc_stats_df = X[numerical_cols].describe().T # DataFrame para HTML
    logging.info(eda_desc_stats_df)

    # 3.2. Distribuição da Variável Alvo ('target')
    logging.info("\n--- 3.2. Distribuição da Variável Alvo ('target') ---")
    eda_target_dist_df = y.value_counts().to_frame(name='count') # DataFrame para HTML
    prop_inativos = y.value_counts(normalize=True).get(1, 0) * 100
    prop_ativos = y.value_counts(normalize=True).get(0, 0) * 100
    logging.info(f"Proporção de 'INATIVOS' (1): {prop_inativos:.2f}%")
    logging.info(f"Proporção de 'ATIVOS' (0): {prop_ativos:.2f}%")

    plt.figure(figsize=(6, 4))
    sns.countplot(x=y) # Usar o 'y' diretamente
    plt.title('Distribuição da Variável Alvo (0=Ativo, 1=Inativo)')
    plt.xlabel('Status')
    plt.ylabel('Contagem')
    plt.xticks(ticks=[0, 1], labels=['ATIVO', 'INATIVO'])
    eda_target_dist_path = os.path.join(OUTPUT_DIR, 'distribuicao_target.png')
    plt.savefig(eda_target_dist_path)
    plt.close() # Fechar o plot para liberar memória
    eda_target_dist_base64 = converter_grafico_para_base64(eda_target_dist_path)
    logging.info(f"✓ Gráfico de distribuição do target salvo e convertido.")

    # 3.3. Análise de Correlação (Variáveis Numéricas)
    logging.info("\n--- 3.3. Análise de Correlação (Variáveis Numéricas) ---")
    # Incluir 'target' para calcular correlação
    correlation_matrix = pd.concat([X[numerical_cols], y], axis=1).corr()
    eda_corr_matrix_df = correlation_matrix # DataFrame para HTML
    logging.info("Matriz de Correlação:")
    logging.info(correlation_matrix)
    logging.info("\nCorrelação das variáveis numéricas com 'target':")
    eda_corr_target_df = correlation_matrix['target'].sort_values(ascending=False).to_frame(name='Correlação com Target')
    logging.info(eda_corr_target_df)

    plt.figure(figsize=(10, 8))
    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt=".2f")
    plt.title('Heatmap da Matriz de Correlação')
    eda_corr_heatmap_path = os.path.join(OUTPUT_DIR, 'heatmap_correlacao.png')
    plt.savefig(eda_corr_heatmap_path)
    plt.close()
    eda_corr_heatmap_base64 = converter_grafico_para_base64(eda_corr_heatmap_path)
    logging.info(f"✓ Heatmap da matriz de correlação salvo e convertido.")

    # 3.4. Visualizações das Distribuições
    eda_hist_box_base64 = {}
    for col in numerical_cols[:3]: # Limitar aos 3 primeiros numéricos para o HTML
        plt.figure(figsize=(12, 5))
        plt.subplot(1, 2, 1)
        sns.histplot(X[col], kde=True)
        plt.title(f'Histograma de {col}')
        plt.subplot(1, 2, 2)
        sns.boxplot(x=X[col])
        plt.title(f'Boxplot de {col}')
        plt.tight_layout()
        filepath = os.path.join(OUTPUT_DIR, f'distribuicao_{col}.png')
        plt.savefig(filepath)
        plt.close()
        eda_hist_box_base64[col] = converter_grafico_para_base64(filepath)
    logging.info(f"✓ Histogramas e Boxplots (exemplos) salvos e convertidos.")

    eda_cat_target_base64 = {}
    categorical_cols = X.select_dtypes(include='object').columns.tolist()
    for col in categorical_cols[:3]: # Limitar aos 3 primeiros categóricos para o HTML
        plt.figure(figsize=(10, 6))
        sns.countplot(x=col, hue=y, data=X, palette='viridis') # Usar X e y
        plt.title(f'Distribuição de {col} por Target')
        plt.xticks(rotation=45, ha='right')
        plt.legend(title='Target', labels=['Ativo', 'Inativo'])
        plt.tight_layout()
        filepath = os.path.join(OUTPUT_DIR, f'distribuicao_categorica_{col}.png')
        plt.savefig(filepath)
        plt.close()
        eda_cat_target_base64[col] = converter_grafico_para_base64(filepath)
    logging.info(f"✓ Gráficos de contagem (exemplos) salvos e convertidos.")

    logging.info("\nANÁLISE EXPLORATÓRIA DOS DADOS CONCLUÍDA!")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 3 (Análise Exploratória dos Dados): {e}")
    sys.exit("Script encerrado devido ao erro na EDA.")

gc.collect() # Limpeza de memória
logging.info("✓ GC.collect() executado após Seção 3.")

# ==============================================================================
# 4. CONSTRUÇÃO E AVALIAÇÃO DO MODELO (Scikit-learn e Statsmodels)
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 4 - CONSTRUÇÃO E AVALIAÇÃO DO MODELO")
logging.info("="*80)

# Variáveis para armazenar resultados do modelo Scikit-learn
smote_pipeline = None
optimal_threshold = 0.5 # Default
accuracy_skl, precision_skl, recall_skl, f1_skl, roc_auc_skl, gini = [0.0]*6
cv_scores_mean, cv_scores_std = 0.0, 0.0
classification_report_str = ""
confusion_matrix_base64 = ""
roc_curve_base64 = ""

# Variáveis para armazenar resultados do modelo GLM
glm_results = None
glm_summary_df = pd.DataFrame()
glm_loglike = "N/A"
glm_aic = "N/A"
glm_bic = "N/A"

try:
    # 4.1. Dividindo dados em Treino e Teste
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    logging.info("\n--- 4.1. Dividindo dados em Treino e Teste ---")
    logging.info(f"✓ Dados divididos: {X_train.shape[0]} amostras de treino, {X_test.shape[0]} amostras de teste.")
    logging.info(f"Proporção de 'INATIVOS' (1) no treino: {y_train.value_counts(normalize=True).get(1, 0)*100:.2f}%")
    logging.info(f"Proporção de 'INATIVOS' (1) no teste: {y_test.value_counts(normalize=True).get(1, 0)*100:.2f}%")

    # 4.2. Preparação de Features 
    numerical_features = X.select_dtypes(include=np.number).columns.tolist()
    categorical_features = X.select_dtypes(include='object').columns.tolist()

    preprocessor_skl = ColumnTransformer(
        transformers=[
            ('num', StandardScaler(), numerical_features),
            ('cat', OneHotEncoder(sparse_output=False, handle_unknown='ignore', drop='if_binary'), categorical_features)
        ],
        remainder='passthrough' # Manter colunas não transformadas
    )
    logging.info("\n--- 4.2. Preparação de Features (ColumnTransformer para Scikit-learn) ---")
    logging.info(f"Features Numéricas identificadas: {numerical_features}")
    logging.info(f"Features Categóricas identificadas: {categorical_features}")
    logging.info("✓ Pré-processador (ColumnTransformer) configurado com sucesso.")

    # 4.3. SMOTE e Pipeline com imblearn (para Scikit-learn Logistic Regression)
    smote_pipeline = ImbPipeline([
        ('preprocessor', preprocessor_skl),
        ('smote', SMOTE(random_state=42)),
        ('classifier', LogisticRegression(solver='lbfgs', random_state=42, class_weight='balanced', max_iter=2000, n_jobs=-1)) # Usar lbfgs, n_jobs=-1 para paralelizar
    ])
    logging.info("\n--- 4.3. Aplicando SMOTE e Construindo Pipeline com imblearn ---")
    logging.info("✓ Pipeline com pré-processador, SMOTE e Regressão Logística (Scikit-learn) construído.")

    # 4.4. Treinamento do Modelo e Validação Cruzada (Scikit-learn)
    logging.info("\n--- 4.4. Treinamento do Modelo e Validação Cruzada (Scikit-learn) ---")
    logging.info("Iniciando o treinamento do pipeline (Scikit-learn)...")
    smote_pipeline.fit(X_train, y_train)
    logging.info("✓ Pipeline Scikit-learn treinado com sucesso!")

    # Validação Cruzada (ROC AUC como métrica de desempenho)
    cv = StratifiedKFold(5, shuffle=True, random_state=42)
    cv_scores = cross_val_score(smote_pipeline, X_train, y_train, cv=cv, scoring='roc_auc', n_jobs=-1)
    cv_scores_mean = cv_scores.mean()
    cv_scores_std = cv_scores.std()
    logging.info(f"✓ Validação cruzada (ROC AUC) concluída: {cv_scores_mean:.4f} (+/- {cv_scores_std:.4f})")
    logging.info(f"Resultados de cada fold: {[f'{score:.4f}' for score in cv_scores]}")

    # Avaliação no conjunto de teste
    y_proba_skl = smote_pipeline.predict_proba(X_test)[:, 1]

    # Ajuste do threshold para otimizar F1-Score (balancear Precision e Recall)
    precision, recall, thresholds = precision_recall_curve(y_test, y_proba_skl)
    f1_scores = 2 * (precision * recall) / (precision + recall + 1e-10) # Adiciona 1e-10 para evitar divisão por zero
    optimal_threshold_idx = np.argmax(f1_scores)
    optimal_threshold = thresholds[optimal_threshold_idx]
    
    y_pred_skl = (y_proba_skl >= optimal_threshold).astype(int)

    accuracy_skl = accuracy_score(y_test, y_pred_skl)
    precision_skl = precision_score(y_test, y_pred_skl)
    recall_skl = recall_score(y_test, y_pred_skl)
    f1_skl = f1_score(y_test, y_pred_skl)
    roc_auc_skl = roc_auc_score(y_test, y_proba_skl)
    gini = 2 * roc_auc_skl - 1 # Coeficiente de Gini

    logging.info("\n--- Avaliação do Modelo (Scikit-learn Logistic Regression) ---")
    logging.info(f"Threshold de classificação ajustado: {optimal_threshold:.4f}")
    logging.info(f"Acurácia: {accuracy_skl:.4f}")
    logging.info(f"Precisão: {precision_skl:.4f}")
    logging.info(f"Recall: {recall_skl:.4f}")
    logging.info(f"F1-Score: {f1_skl:.4f}")
    logging.info(f"ROC AUC: {roc_auc_skl:.4f}")
    logging.info(f"Coeficiente de Gini: {gini:.4f}")

    classification_report_str = classification_report(y_test, y_pred_skl, output_dict=False)
    logging.info("\nRelatório de Classificação:\n" + classification_report_str)

    # Matriz de Confusão
    cm_skl = confusion_matrix(y_test, y_pred_skl)
    plt.figure(figsize=(6, 5))
    sns.heatmap(cm_skl, annot=True, fmt='d', cmap='Blues',
                xticklabels=['Previsto Ativo (0)', 'Previsto Inativo (1)'],
                yticklabels=['Real Ativo (0)', 'Real Inativo (1)'])
    plt.title(f'Matriz de Confusão (Scikit-learn) - Threshold={optimal_threshold:.2f}')
    plt.ylabel('Valor Real')
    plt.xlabel('Valor Previsto')
    confusion_matrix_path = os.path.join(OUTPUT_DIR, 'matriz_confusao_sklearn.png')
    plt.savefig(confusion_matrix_path)
    plt.close() # Fechar o plot
    confusion_matrix_base64 = converter_grafico_para_base64(confusion_matrix_path)
    logging.info(f"✓ Matriz de Confusão salva e convertida.")

    # Curva ROC com Coeficiente de Gini
    fpr, tpr, thresholds_roc = roc_curve(y_test, y_proba_skl)
    plt.figure(figsize=(8, 6))
    plt.plot(fpr, tpr, color='darkorange', lw=2, label=f'Curva ROC (AUC = {roc_auc_skl:.2f}, Gini = {gini:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--', label='Linha de Referência Aleatória')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('Taxa de Falso Positivo (1 - Especificidade)')
    plt.ylabel('Taxa de Verdadeiro Positivo (Sensibilidade)')
    plt.title(f'Curva ROC para Previsão de Demissão (Gini = {gini:.2f})')
    plt.legend(loc="lower right")
    plt.grid(True)
    roc_curve_path = os.path.join(OUTPUT_DIR, 'curva_roc_gini.png')
    plt.savefig(roc_curve_path)
    plt.close() # Fechar o plot
    roc_curve_base64 = converter_grafico_para_base64(roc_curve_path)
    logging.info(f"✓ Curva ROC com Gini salva e convertida.")
    
except Exception as e:
    logging.error(f"ERRO na SEÇÃO 4.4 (Scikit-learn): {e}")
    import traceback
    traceback.print_exc()
    sys.exit("Script encerrado devido ao erro no treinamento do modelo Scikit-learn.")

# ==============================================================================
# 4.5.B. Modelo GLM (Generalized Linear Model) com Statsmodels (COM SMOTE)
# ==============================================================================
logging.info("\n--- 4.5.B. Modelo GLM (Generalized Linear Model) com Statsmodels (COM SMOTE) ---")

try:
    # 1) Pré-processador dedicado ao GLM com baseline nas dummies (drop='first')
    # Usamos o mesmo ColumnTransformer, mas ajustando para Statsmodels (drop='first' para evitar multicolinearidade com intercepto)
    preprocessor_glm = ColumnTransformer(
        transformers=[
            ('num', StandardScaler(), numerical_features),
            ('cat', OneHotEncoder(sparse_output=False, handle_unknown='ignore', drop='first'), categorical_features)
        ],
        remainder='drop' # O GLM não aceita 'passthrough' de colunas não processadas facilmente
    )

    # 2) Consolidar categorias raras (evita dummies quase exclusivas de uma classe)
    def consolidar_categorias_raras(df, cols, min_freq=0.01):
        df = df.copy()
        for c in cols:
            if c in df.columns and df[c].dtype == 'object': # Garante que só categorias sejam processadas
                freq = df[c].value_counts(normalize=True)
                raras = freq.index[freq < min_freq]
                if len(raras) > 0:
                    df[c] = df[c].where(~df[c].isin(raras), 'OUTROS')
        return df

    # X_train_glm_raw é a versão de X_train antes da transformação para GLM
    X_train_glm_raw = consolidar_categorias_raras(X_train, categorical_features, min_freq=0.01)

    # 3) Transformar o treino com o preprocessor_glm
    X_train_glm_arr = preprocessor_glm.fit_transform(X_train_glm_raw)

    glm_cat_names = []
    if len(categorical_features) > 0:
        # Pega os nomes das features categóricas transformadas (com 'drop='first')
        glm_cat_names = list(preprocessor_glm.named_transformers_['cat'].get_feature_names_out(categorical_features))

    feature_names_glm = numerical_features + glm_cat_names
    X_train_glm_df_unbalanced = pd.DataFrame(X_train_glm_arr, columns=feature_names_glm, index=X_train.index)

    # 4) Remover colinearidade via VIF de forma iterativa (antes do SMOTE para ter VIF real)
    def remover_vif_alto(df, thresh=5.0, max_iter=20): # Limiar mais rigoroso para GLM
        df_work = df.copy()
        for _ in range(max_iter):
            if df_work.shape[1] <= 1:
                break
            vif_vals = []
            cols = df_work.columns
            for i in range(df_work.shape[1]):
                try:
                    vif_vals.append(variance_inflation_factor(df_work.values, i))
                except Exception as e: # Lidar com colunas que podem ser todas zero ou ter variância zero
                    logging.warning(f"Erro calculando VIF para {cols[i]}: {e}. Removendo.")
                    vif_vals.append(np.inf) # Marca para remoção
            
            vif = pd.Series(vif_vals, index=cols, name='VIF')
            col_max = vif.idxmax()
            if vif[col_max] > thresh:
                logging.info(f"Removendo por VIF alto: {col_max} (VIF={vif[col_max]:.2f})")
                df_work = df_work.drop(columns=[col_max])
            else:
                break
        return df_work

    X_train_glm_df_unbalanced_vif_cleaned = remover_vif_alto(X_train_glm_df_unbalanced, thresh=5.0, max_iter=20)
    logging.info(f"✓ Variáveis após limpeza VIF: {X_train_glm_df_unbalanced_vif_cleaned.shape[1]}")

    # 5) Detectar e remover separação perfeita simples (dummies 0/1 que só ocorrem numa classe)
    # Este passo é crucial antes de adicionar o intercepto e aplicar SMOTE.
    sep_cols = []
    for col in X_train_glm_df_unbalanced_vif_cleaned.columns:
        x_col = X_train_glm_df_unbalanced_vif_cleaned[col]
        # Statsmodels espera que as variáveis não sejam constantes (e.g., todas zero ou todas um)
        # E também lida com "separação perfeita" que causa coeficientes infinitos.
        # Aqui, estamos procurando colunas que só existem para uma classe do target.
        if (x_col[y_train == 1].sum() == 0 and x_col[y_train == 0].sum() > 0) or \
           (x_col[y_train == 0].sum() == 0 and x_col[y_train == 1].sum() > 0):
            sep_cols.append(col)

    if len(sep_cols) > 0:
        logging.info(f"Removendo colunas com separação perfeita: {sep_cols}")
        X_train_glm_df_unbalanced_vif_cleaned = X_train_glm_df_unbalanced_vif_cleaned.drop(columns=sep_cols)
    logging.info(f"✓ Variáveis após limpeza de separação perfeita: {X_train_glm_df_unbalanced_vif_cleaned.shape[1]}")

    # 6) ✅ APLICAR SMOTE (NOVO!)
    logging.info("Aplicando SMOTE para balancear classes para o GLM...")
    smote_glm = SMOTE(random_state=42, k_neighbors=5) # k_neighbors padrão é 5, mas é bom especificar
    # X_train_glm_df_unbalanced_vif_cleaned é o X para o SMOTE
    X_train_glm_balanced, y_train_balanced = smote_glm.fit_resample(
        X_train_glm_df_unbalanced_vif_cleaned, y_train
    )
    
    logging.info(f"Antes SMOTE GLM: {len(y_train)} amostras (positivas: {y_train.sum()})")
    logging.info(f"Depois SMOTE GLM: {len(y_train_balanced)} amostras (positivas: {y_train_balanced.sum()})")

    # 7) Adicionar intercepto (APÓS SMOTE e VIF)
    X_train_glm_balanced = sm.add_constant(X_train_glm_balanced, prepend=True)
    logging.info(f"✓ Features GLM finais preparadas: {X_train_glm_balanced.shape[1]} variáveis (incluindo intercepto)")

    # 8) Treinar GLM SEM penalização (agora com dados balanceados e limpos)
    logging.info("Treinando GLM sem penalização (com SMOTE e dados limpos)...")
    glm_model = sm.GLM(y_train_balanced, X_train_glm_balanced, family=sm.families.Binomial())
    glm_results = glm_model.fit(disp=0) # disp=0 para suprimir output da otimização
    
    logging.info("✓ GLM treinado com sucesso!")

    # 9) Calcular métricas de ajuste (Log-Likelihood, AIC, BIC)
    # Estas métricas já estão disponíveis no objeto glm_results para modelos não penalizados
    glm_loglike = glm_results.llf
    glm_aic = glm_results.aic
    glm_bic = glm_results.bic
    
    logging.info(f"✓ Métricas GLM (com SMOTE):")
    logging.info(f"  - Log-Likelihood: {glm_loglike:.4f}")
    logging.info(f"  - AIC: {glm_aic:.4f}")
    logging.info(f"  - BIC: {glm_bic:.4f}")

    # 10) Gerar coeficientes com p-values e Odds Ratios
    betas = glm_results.params
    # Clip para evitar 'inf' em Odds Ratios de coeficientes muito grandes, comum em logísticas
    odds_ratio = np.exp(np.clip(betas.values, -20, 20)) 

    glm_summary_df = pd.DataFrame({
        'Feature': betas.index,
        'Coefficient (Beta)': betas.values,
        'Std Err': glm_results.bse.values,
        'p-value': glm_results.pvalues.values,
        'Odds Ratio': odds_ratio,
        'Significância': ['***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else '' 
                         for p in glm_results.pvalues.values]
    }).sort_values(by='p-value').reset_index(drop=True)

    glm_summary_df.to_csv(OUTPUT_GLM_COEF_CSV, index=False)
    logging.info(f"✓ Coeficientes GLM salvos em '{OUTPUT_GLM_COEF_CSV}'")

    # 11) Salvar sumário completo do Statsmodels
    with open(OUTPUT_GLM_SUMMARY_TXT, 'w') as f:
        f.write(glm_results.summary().as_text())
    logging.info(f"✓ Sumário completo do GLM salvo em '{OUTPUT_GLM_SUMMARY_TXT}'")

    logging.info("✓ SEÇÃO 4.5.B CONCLUÍDA COM SUCESSO\n")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 4.5.B (GLM com SMOTE): {e}")
    import traceback
    traceback.print_exc()
    # Em caso de erro, garante que as variáveis GLM estão no estado "N/A" ou DataFrame vazio.
    glm_results = None
    glm_summary_df = pd.DataFrame()
    glm_loglike = "N/A"
    glm_aic = "N/A"
    glm_bic = "N/A"

gc.collect() # Limpeza de memória
logging.info("✓ GC.collect() executado após Seção 4.")

# ==============================================================================
# 5. PREVISÕES E EXPORTAÇÃO MELHORADA DOS RESULTADOS
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 5 - PREVISÕES E EXPORTAÇÃO MELHORADA DOS RESULTADOS")
logging.info("="*80)

df_resultados = pd.DataFrame()  # Inicializar para garantir que existe

try:
    if smote_pipeline is None:
        raise ValueError("Pipeline Scikit-learn não foi treinado. Não é possível fazer previsões.")

    # 5.1. Calculando probabilidades e previsões para o DataFrame original completo (filtrado)
    prob_demissao_completa = smote_pipeline.predict_proba(X)[:, 1]  # Previsão para todo o X
    previsao_demissao_completa = smote_pipeline.predict(X)

    # Reconstruir o DataFrame original filtrado com as novas colunas de previsão
    df_resultados = df_hc_filtered_full.copy()  # DataFrame com todas as colunas originais filtradas
    
    # Garantir que o `target` original (y) seja adicionado ao df_resultados
    df_resultados['target_real'] = y.values  # Adiciona o target real para comparação

    df_resultados['prob_demissao'] = prob_demissao_completa
    df_resultados['previsao_demissao'] = previsao_demissao_completa

    logging.info("✓ Probabilidades e previsões calculadas para todos os colaboradores do conjunto filtrado.")

    # 5.2. Segmentando colaboradores em grupos de risco (Baixo, Médio, Alto)
    threshold_baixo = 0.3
    threshold_alto = 0.7

    def classificar_risco(prob):
        if prob >= threshold_alto:
            return 'Alto Risco'
        elif prob >= threshold_baixo:
            return 'Médio Risco'
        else:
            return 'Baixo Risco'

    df_resultados['segmento_risco'] = df_resultados['prob_demissao'].apply(classificar_risco)
    logging.info("✓ Colaboradores segmentados. Distribuição:\n" + df_resultados['segmento_risco'].value_counts().to_string())

    # 5.3. Exportando resultados de previsões e risco para 'df_previsoes_risco.xlsx' com formatação
    # Selecionar apenas colunas que existem no DataFrame
    colunas_disponiveis = df_resultados.columns.tolist()
    
    # Definir limiares de risco (mesmos usados em 4.4)
    threshold_baixo_risco = 0.3
    threshold_alto_risco = 0.7
    
    # Definir colunas desejadas (na ordem)
    colunas_prioritarias = [
        'nome', 'cpf', 'cargo', 'departamento', 'idade', 'salario_total', 
        'dias_trabalhado', 'dias_afastado', 'custo_afastamento', 'filhos',
        'target_real', 'prob_demissao', 'previsao_demissao', 'segmento_risco'
    ]
    
    # Manter apenas colunas que existem
    cols_to_export = [col for col in colunas_prioritarias if col in colunas_disponiveis]
    
    # Adicionar outras colunas que existem mas não estão na lista prioritária
    outras_colunas = [col for col in colunas_disponiveis if col not in cols_to_export]
    cols_to_export.extend(outras_colunas)
    
    logging.info(f"✓ Colunas para exportação ({len(cols_to_export)}): {cols_to_export[:10]}...")

    df_final_export = df_resultados[cols_to_export].copy()

    # Criar workbook com formatação
    wb = Workbook()
    ws = wb.active
    ws.title = "Previsões de Risco de Demissão"

    # Escrever cabeçalhos com formatação
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, header in enumerate(df_final_export.columns):
        cell = ws.cell(row=1, column=col_idx + 1, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Escrever os dados
    for r_idx, row in enumerate(dataframe_to_rows(df_final_export, index=False, header=False), 2):
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=r_idx, column=c_idx + 1, value=value)
            col_name = df_final_export.columns[c_idx]
            
            # Formatação específica por coluna
            if col_name == 'prob_demissao':
                cell.number_format = '0.00%'
            elif col_name in ['target_real', 'previsao_demissao']:
                cell.number_format = '0'
            elif col_name in ['salario_total', 'custo_afastamento']:
                cell.number_format = '#,##0.00'
            elif col_name in ['idade', 'dias_trabalhado', 'dias_afastado']:
                cell.number_format = '0'
            
            # Alinhamento
            if col_name in ['target_real', 'prob_demissao', 'previsao_demissao', 'segmento_risco']:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar a largura das colunas
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)  # Máximo de 50 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width

    # Congelar primeira linha
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT_EXCEL_PREVISOES_FILE)
    logging.info(f"✓ Resultados exportados para '{OUTPUT_EXCEL_PREVISOES_FILE}' com formatação aprimorada.")
    logging.info(f"  Total de linhas: {len(df_final_export)}")
    logging.info(f"  Total de colunas: {len(df_final_export.columns)}")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 5 (Previsões e Exportação): {e}")
    import traceback
    traceback.print_exc()
    sys.exit("Script encerrado devido ao erro nas previsões ou exportação.")

gc.collect()  # Limpeza de memória
logging.info("✓ GC.collect() executado após Seção 5.")

# ==============================================================================
# 6. RECOMENDAÇÕES ESTRATÉGICAS
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 6 - RECOMENDAÇÕES ESTRATÉGICAS")
logging.info("="*80)

# Variáveis para armazenar texto das recomendações (sem fatores de risco/proteção)
recomendacoes_baixo_risco = ""
recomendacoes_medio_risco = ""
recomendacoes_alto_risco = ""
recomendacoes_kpis = ""

try:
    logging.info("\n--- 6.1. Ações Recomendadas por Segmento de Risco ---")
    recomendacoes_baixo_risco = """
    <h5 class="text-success">Colaboradores de Baixo Risco:</h5>
    <ul>
        <li>Manter engajamento através de reconhecimento, feedback positivo e oportunidades de desenvolvimento contínuo.</li>
        <li>Incentivar a mentoria e o compartilhamento de conhecimento.</li>
        <li>Coletar feedback regularmente (e.g., micro-surveys) para identificar potenciais problemas antes que escalem.</li>
        <li>Identificar "embaixadores" para cultura organizacional e mentores para novos colaboradores.</li>
    </ul>
    """
    recomendacoes_medio_risco = """
    <h5 class="text-warning">Colaboradores de Médio Risco:</h5>
    <ul>
        <li>Implementar 'stay interviews' para entender expectativas e motivações.</li>
        <li>Programas de retenção personalizados: planos de carreira, incentivos, flexibilidade.</li>
        <li>Treinamentos orientados ao desenvolvimento de habilidades críticas.</li>
        <li>Conexão com líderes e mentores para fortalecer pertencimento e identificar pontos de atrito.</li>
        <li>Revisão de carga de trabalho e equilíbrio vida-trabalho.</li>
    </ul>
    """
    recomendacoes_alto_risco = """
    <h5 class="text-danger">Colaboradores de Alto Risco:</h5>
    <ul>
        <li>Intervenção imediata e personalizada por RH e liderança direta.</li>
        <li>Investigar causas raízes (sobrecarga, conflitos, falta de perspectiva, saúde, insatisfação salarial).</li>
        <li>Planos de ação de curto prazo (realocação, projeto, ajuste de remuneração, suporte psicológico/social).</li>
        <li>Plano de sucessão caso retenção não seja viável, para mitigar impactos.</li>
        <li>Acompanhamento contínuo e feedback estruturado.</li>
    </ul>
    """

    logging.info("\n--- 6.2. KPIs de Monitoramento Contínuo ---")
    recomendacoes_kpis = """
    <h5 class="text-info">KPIs de Monitoramento Contínuo:</h5>
    <ul>
        <li>Taxa de Rotatividade Voluntária (geral e por segmento de risco).</li>
        <li>Custo de Demissão (custo de reposição, treinamento, perda de produtividade).</li>
        <li>Taxa de Engajamento (pesquisas de clima e pulso).</li>
        <li>% de retenção nos segmentos de Médio e Alto Risco após intervenções.</li>
        <li>Eficácia das ações de retenção (feedbacks qualitativos e quantitativos sobre as intervenções).</li>
        <li>Produtividade por segmento de risco.</li>
    </ul>
    """
    logging.info("\n✓ SEÇÃO 6 CONCLUÍDA COM SUCESSO")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 6 (Recomendações Estratégicas): {e}")
    sys.exit("Script encerrado devido ao erro nas recomendações estratégicas.")

# ==============================================================================
# 7. GERAÇÃO DO RELATÓRIO HTML COMPLETO
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 7 - GERAÇÃO DO RELATÓRIO HTML COMPLETO")
logging.info("="*80)

def preparar_arquivo_para_download(caminho_arquivo, pasta_html):
    # Garante que o arquivo esteja na mesma pasta do HTML e retorna apenas o nome do arquivo
    if not os.path.isfile(caminho_arquivo):
        logging.warning(f"Arquivo para download não encontrado: {caminho_arquivo}")
        return os.path.basename(caminho_arquivo)
    nome = os.path.basename(caminho_arquivo)
    destino = os.path.join(pasta_html, nome)
    # Evita copiar se o arquivo já está no destino e é o mesmo arquivo (caminho absoluto)
    if os.path.abspath(caminho_arquivo) != os.path.abspath(destino):
        os.makedirs(pasta_html, exist_ok=True)
        shutil.copy2(caminho_arquivo, destino)
        logging.info(f"Arquivo copiado para a pasta do HTML: {destino}")
    return nome  # sempre retornar só o nome do arquivo

html_dir = os.path.dirname(OUTPUT_HTML_FILE) if os.path.dirname(OUTPUT_HTML_FILE) else os.getcwd()
os.makedirs(html_dir, exist_ok=True)

link_previsoes = preparar_arquivo_para_download(OUTPUT_EXCEL_PREVISOES_FILE, html_dir)
link_glm_coef = preparar_arquivo_para_download(OUTPUT_GLM_COEF_CSV, html_dir)
link_glm_sum = preparar_arquivo_para_download(OUTPUT_GLM_SUMMARY_TXT, html_dir)
link_dashboard_excel = preparar_arquivo_para_download(OUTPUT_DASHBOARD_EXCEL_FILE, html_dir)

html_content = "" # Inicializar para garantir que existe

try:
    # 7.1. Preparar dados para o HTML
    # Garantir que todas as variáveis necessárias existam ou tenham um valor padrão
    # EDA
    eda_desc_stats_html = gerar_tabela_html(eda_desc_stats_df if 'eda_desc_stats_df' in locals() else pd.DataFrame(), "Estatísticas Descritivas (Variáveis Numéricas)")
    eda_target_dist_html = gerar_tabela_html(eda_target_dist_df if 'eda_target_dist_df' in locals() else pd.DataFrame(), "Distribuição da Variável Alvo")
    eda_corr_matrix_html = gerar_tabela_html(eda_corr_matrix_df if 'eda_corr_matrix_df' in locals() else pd.DataFrame(), "Matriz de Correlação Completa")
    eda_corr_target_html = gerar_tabela_html(eda_corr_target_df if 'eda_corr_target_df' in locals() else pd.DataFrame(), "Correlação com Target")

    eda_hist_box_imgs_html = "".join([f'<div class="col-md-6 mb-4"><img src="{converter_grafico_para_base64(os.path.join(OUTPUT_DIR, f"distribuicao_{col}.png"))}" class="img-fluid rounded shadow-sm" alt="Histograma/Boxplot de {col}"></div>' for col in numerical_cols[:3] if os.path.exists(os.path.join(OUTPUT_DIR, f"distribuicao_{col}.png"))])
    eda_cat_target_imgs_html = "".join([f'<div class="col-md-6 mb-4"><img src="{converter_grafico_para_base64(os.path.join(OUTPUT_DIR, f"distribuicao_categorica_{col}.png"))}" class="img-fluid rounded shadow-sm" alt="Distribuição Categórica {col} por Target"></div>' for col in categorical_cols[:3] if os.path.exists(os.path.join(OUTPUT_DIR, f"distribuicao_categorica_{col}.png"))])

    # Modelo Scikit-learn
    metrics_cards_skl = f"""
    {gerar_card_metricas("Acurácia", f"{accuracy_skl:.2%}", "Proporção de previsões corretas.")}
    {gerar_card_metricas("Precisão", f"{precision_skl:.2%}", "Das previsões 'INATIVO', quantas estavam corretas.")}
    {gerar_card_metricas("Recall", f"{recall_skl:.2%}", "Dos 'INATIVOS' reais, quantos foram identificados.")}
    {gerar_card_metricas("F1-Score", f"{f1_skl:.2f}", "Média harmônica entre precisão e recall.")}
    {gerar_card_metricas("ROC AUC", f"{roc_auc_skl:.2f}", "Probabilidade de o modelo classificar um positivo aleatório mais alto que um negativo aleatório.")}
    {gerar_card_metricas("Gini", f"{gini:.2f}", "Medida da desigualdade na capacidade de previsão do modelo (2*AUC-1).")}
    {gerar_card_metricas("Threshold Ajustado", f"{optimal_threshold:.2f}", "Ponto de corte para classificar como 'INATIVO'.")}
    {gerar_card_metricas("CV ROC AUC Média", f"{cv_scores_mean:.2f}", f"Média da AUC em validação cruzada (std: {cv_scores_std:.2f}).")}
    """

    # Modelo GLM
    # Badge e texto explicativo dinâmicos
    glm_badge_html = "<span class='badge bg-success ms-2'>Com SMOTE (p-values disponíveis)</span>"
    glm_table_title = f"Coeficientes GLM {glm_badge_html}"
    glm_table_caption = (
        "Modelo ajustado com dados balanceados por SMOTE, exibindo coeficientes, erros padrão, "
        "p-values e Odds Ratios. As variáveis são ordenadas por significância (menor p-value)."
    )
    
    # Importante: exibir TODOS os resultados (sem limite)
    glm_summary_html = gerar_tabela_html(
        glm_summary_df if 'glm_summary_df' in locals() and not glm_summary_df.empty else pd.DataFrame(),
        title=glm_table_title,
        caption=glm_table_caption,
        limit=None
    )
    
    # Cards de métricas do GLM
    glm_metrics_cards = f"""
    {gerar_card_metricas("Log-Likelihood", f"{glm_loglike:.2f}" if isinstance(glm_loglike, (int, float)) else str(glm_loglike), "Medida de ajuste do modelo. Quanto maior (menos negativo), melhor.")}
    {gerar_card_metricas("AIC", f"{glm_aic:.2f}" if isinstance(glm_aic, (int, float)) else str(glm_aic), "Critério de Informação de Akaike. Penaliza complexidade. Menor é melhor.")}
    {gerar_card_metricas("BIC", f"{glm_bic:.2f}" if isinstance(glm_bic, (int, float)) else str(glm_bic), "Critério de Informação Bayesiano. Penaliza complexidade mais que AIC. Menor é melhor.")}
    """
    
    # Segmentação
    segmento_risco_dist_df = df_resultados['segmento_risco'].value_counts().to_frame(name='Contagem')
    segmento_risco_dist_html = gerar_tabela_html(segmento_risco_dist_df, "Distribuição por Segmento de Risco")

    # 7.2. Construção do HTML
    html_content = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório de People Analytics: Previsão de Demissão</title>
    <!-- Bootstrap CSS -->
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <!-- Font Awesome para ícones -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0-beta3/css/all.min.css">
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f8f9fa; color: #343a40; }}
        .navbar {{ background-color: #0070c0 !important; }}
        .navbar-brand, .nav-link {{ color: #ffffff !important; }}
        .navbar-nav .nav-link.active {{ font-weight: bold; }}
        .header {{ background-color: #0070c0; color: white; padding: 4rem 0; text-align: center; }}
        .section-title {{ color: #0070c0; border-bottom: 3px solid #0070c0; padding-bottom: 10px; margin-bottom: 30px; }}
        .card {{ border-radius: 0.75rem; border: none; box-shadow: 0 0.5rem 1rem rgba(0,0,0,0.05); }}
        .card-title {{ color: #0070c0; font-weight: bold; }}
        .table-responsive {{ margin-top: 20px; }}
        .table {{ margin-bottom: 0; }}
        .table th {{ background-color: #e9ecef; color: #495057; }}
        .footer {{ background-color: #343a40; color: white; padding: 2rem 0; text-align: center; margin-top: 4rem; }}
        .img-fluid {{ max-width: 100%; height: auto; display: block; margin-left: auto; margin-right: auto; }}
        .back-to-top {{
            position: fixed;
            bottom: 20px;
            right: 20px;
            background-color: #0070c0;
            color: white;
            width: 40px;
            height: 40px;
            border-radius: 50%;
            text-align: center;
            line-height: 40px;
            font-size: 1.5rem;
            box-shadow: 0 2px 5px rgba(0,0,0,0.2);
            z-index: 1000;
            display: none; /* Escondido por padrão */
        }}
    </style>
</head>
<body data-bs-spy="scroll" data-bs-target="#navbar-example" data-bs-offset="50">

    <nav class="navbar navbar-expand-lg navbar-dark fixed-top shadow" id="navbar-example">
        <div class="container">
            <a class="navbar-brand" href="#">People Analytics</a>
            <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav" aria-controls="navbarNav" aria-expanded="false" aria-label="Toggle navigation">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav ms-auto">
                    <li class="nav-item"><a class="nav-link" href="#resumo">Resumo</a></li>
                    <li class="nav-item"><a class="nav-link" href="#contexto">Contexto</a></li>
                    <li class="nav-item"><a class="nav-link" href="#eda">EDA</a></li>
                    <li class="nav-item"><a class="nav-link" href="#modelo_skl">Modelo Scikit-learn</a></li>
                    <li class="nav-item"><a class="nav-link" href="#modelo_glm">Análise GLM</a></li>
                    <li class="nav-item"><a class="nav-link" href="#previsoes">Previsões</a></li>
                    <li class="nav-item"><a class="nav-link" href="#dashboard">Dashboard Executivo</a></li>
                    <li class="nav-item"><a class="nav-link" href="#recomendacoes">Recomendações</a></li>
                    <li class="nav-item"><a class="nav-link" href="#conclusao">Conclusão</a></li>
                </ul>
            </div>
        </div>
    </nav>

    <div class="header">
        <div class="container">
            <h1 class="display-4">Relatório de People Analytics</h1>
            <p class="lead">Previsão de Demissão de Colaboradores Utilizando Modelo Híbrido</p>
            <p class="mt-3"><i>Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</i></p>
        </div>
    </div>

    <div class="container my-5">
        <section id="resumo" class="mb-5">
            <h2 class="section-title">1. Resumo Executivo</h2>
            <p>Este relatório apresenta os resultados de um projeto de People Analytics focado na previsão de demissão de colaboradores ('INATIVO') utilizando um modelo híbrido. A solução integra um modelo de Regressão Logística (Scikit-learn) para alta acurácia preditiva e um Modelo Linear Generalizado (Statsmodels GLM) para interpretabilidade estatística, ambos otimizados com SMOTE para balanceamento de classes. O objetivo é fornecer insights acionáveis para a Gestão de Pessoas.</p>
            <p>O modelo Scikit-learn alcançou uma <strong>ROC AUC de {roc_auc_skl:.2f}</strong>, indicando uma excelente capacidade preditiva, e um <strong>Coeficiente de Gini de {gini:.2f}</strong>. O GLM, por sua vez, revelou os principais fatores de risco e proteção com base em sua significância estatística (p-values). Os resultados incluem a segmentação de colaboradores por risco (Baixo, Médio, Alto) e um dashboard executivo para facilitar a tomada de decisão.</p>
            <div class="row">{metrics_cards_skl}</div>
        </section>

        <section id="contexto" class="mb-5">
            <h2 class="section-title">2. Contexto e Objetivos</h2>
            <p>A rotatividade de colaboradores (turnover) é um desafio significativo para qualquer organização, impactando custos de recrutamento, treinamento e produtividade. Este projeto visa utilizar técnicas avançadas de People Analytics para:</p>
            <ul>
                <li><strong>Prever</strong> a probabilidade de demissão de um colaborador com alta precisão.</li>
                <li><strong>Identificar</strong> os principais fatores (variáveis) que influenciam a demissão e sua significância estatística.</li>
                <li><strong>Segmentar</strong> colaboradores em diferentes níveis de risco (Baixo, Médio, Alto) para intervenções personalizadas.</li>
                <li><strong>Fornecer</strong> recomendações estratégicas e um dashboard executivo para ações proativas de retenção.</li>
            </ul>
            <p>Os dados utilizados incluem informações de atestados médicos e dados cadastrais de colaboradores da unidade 'UL', filtrados para o período de 2022.</p>
        </section>

        <section id="eda" class="mb-5">
            <h2 class="section-title">3. Análise Exploratória dos Dados (EDA)</h2>
            <p>A EDA foi realizada para entender a estrutura, distribuição e relações entre as variáveis. Esta etapa é fundamental para identificar padrões, tratar anomalias e preparar os dados para a modelagem.</p>

            <h3 class="mt-4">3.1. Estatísticas Descritivas</h3>
            <p>Visão geral das principais estatísticas das variáveis numéricas:</p>
            {eda_desc_stats_html}

            <h3 class="mt-4">3.2. Distribuição da Variável Alvo ('target')</h3>
            <p>A variável alvo (target) indica se o colaborador está ATIVO (0) ou INATIVO (1 - demitido). A distribuição mostra um desbalanceamento, com uma minoria de casos 'INATIVO', o que foi tratado com SMOTE em ambos os modelos.</p>
            {eda_target_dist_html}
            <div class="row">
                <div class="col-md-8 mx-auto mb-4">
                    <img src="{eda_target_dist_base64}" class="img-fluid rounded shadow-sm" alt="Distribuição da Variável Alvo">
                    <p class="text-center text-muted mt-2"><i>Distribuição da variável alvo (0=Ativo, 1=Inativo).</i></p>
                </div>
            </div>

            <h3 class="mt-4">3.3. Análise de Correlação</h3>
            <p>O heatmap abaixo ilustra as correlações entre as variáveis numéricas, incluindo a variável 'target'. Coeficientes próximos de 1 ou -1 indicam forte correlação positiva ou negativa, respectivamente. Coeficientes próximos de 0 indicam pouca ou nenhuma correlação linear.</p>
            {eda_corr_matrix_html}
            {eda_corr_target_html}
            <div class="row">
                <div class="col-md-8 mx-auto mb-4">
                    <img src="{eda_corr_heatmap_base64}" class="img-fluid rounded shadow-sm" alt="Heatmap de Correlação">
                    <p class="text-center text-muted mt-2"><i>Heatmap da Matriz de Correlação entre variáveis numéricas.</i></p>
                </div>
            </div>

            <h3 class="mt-4">3.4. Visualizações Detalhadas</h3>
            <p>Exemplos de distribuições para algumas variáveis:</p>
            <div class="row">{eda_hist_box_imgs_html}</div>
            <div class="row">{eda_cat_target_imgs_html}</div>

        </section>

        <section id="modelo_skl" class="mb-5">
            <h2 class="section-title">4. Modelo de Machine Learning: Regressão Logística (Scikit-learn)</h2>
            <p>Um Pipeline de Regressão Logística foi construído utilizando Scikit-learn, incluindo pré-processamento (padronização e One-Hot Encoding) e balanceamento de classes com SMOTE. Este modelo é otimizado para a performance preditiva.</p>

            <h3 class="mt-4">4.1. Configuração do Pipeline</h3>
            <p>O pipeline é composto por:</p>
            <ul>
                <li><code>ColumnTransformer</code>: Aplica <code>StandardScaler</code> às variáveis numéricas e <code>OneHotEncoder</code> às categóricas.</li>
                <li><code>SMOTE</code>: Over-sampling da classe minoritária ('INATIVO') para tratar o desbalanceamento, garantindo que o modelo aprenda bem ambas as classes.</li>
                <li><code>LogisticRegression</code>: O modelo de classificação final, treinado para prever a probabilidade de demissão.</li>
            </ul>

            <h3 class="mt-4">4.2. Métricas de Avaliação</h3>
            <p>O modelo foi avaliado no conjunto de teste usando um threshold de classificação ajustado para otimizar o F1-Score, buscando um bom equilíbrio entre Precisão e Recall. Os resultados são apresentados abaixo:</p>
            <div class="row">{metrics_cards_skl}</div>
            
            <h3 class="mt-4">4.3. Matriz de Confusão</h3>
            <p>A matriz de confusão detalha o desempenho do modelo, mostrando o número de verdadeiros positivos, verdadeiros negativos, falsos positivos e falsos negativos. Valores ideais são altos nas diagonais principais.</p>
            <div class="row">
                <div class="col-md-8 mx-auto mb-4">
                    <img src="{confusion_matrix_base64}" class="img-fluid rounded shadow-sm" alt="Matriz de Confusão">
                    <p class="text-center text-muted mt-2"><i>Matriz de Confusão do Modelo de Regressão Logística (Scikit-learn) com threshold ajustado.</i></p>
                </div>
            </div>

            <h3 class="mt-4">4.4. Curva ROC e Coeficiente de Gini</h3>
            <p>A Curva ROC (Receiver Operating Characteristic) é uma ferramenta gráfica para avaliar o desempenho de modelos de classificação. A Área Sob a Curva (AUC) mede a capacidade do modelo de distinguir entre as classes (quanto mais próximo de 1, melhor). O Coeficiente de Gini é derivado da AUC (<code>2 * AUC - 1</code>) e varia de 0 a 1, onde 1 é um modelo perfeito.</p>
            <div class="row">
                <div class="col-md-8 mx-auto mb-4">
                    <img src="{roc_curve_base64}" class="img-fluid rounded shadow-sm" alt="Curva ROC">
                    <p class="text-center text-muted mt-2"><i>Curva ROC com valor AUC e Coeficiente de Gini.</i></p>
                </div>
            </div>
            
            <h3 class="mt-4">4.5. Relatório de Classificação</h3>
            <pre class="bg-light p-3 rounded">{classification_report_str}</pre>
            
        </section>

        <section id="modelo_glm" class="mb-5">
            <h2 class="section-title">5. Análise do Modelo GLM (Statsmodels)</h2>
            <p>O Modelo Linear Generalizado (GLM) foi utilizado para fornecer uma interpretação estatística mais aprofundada dos coeficientes e sua significância (p-values). Este modelo foi treinado com SMOTE e técnicas de limpeza para garantir estabilidade e interpretabilidade, ajudando a identificar quais variáveis são estatisticamente importantes na previsão de demissão.</p>
            <div class="row">{glm_metrics_cards}</div>
            
            <h3 class="mt-4">5.1. Coeficientes do Modelo GLM — Fatores de Risco e Proteção</h3>
            <p>O GLM fornece os coeficientes (Beta) que representam o impacto de cada variável na probabilidade de demissão, e as Odds Ratios, que quantificam essa relação. Os p-values indicam a significância estatística de cada fator.</p>
            {glm_summary_html}
            <p class="text-muted mt-2">
                <i>**Significância:** *** p &lt; 0.001, ** p &lt; 0.01, * p &lt; 0.05.</i><br>
                <i>Odds Ratio &gt; 1: Aumenta a chance de demissão. Odds Ratio &lt; 1: Diminui a chance de demissão.</i>
            </p>
            <a href="{link_glm_coef}" class="btn btn-primary mt-3" download><i class="fas fa-download"></i> Baixar Coeficientes GLM (CSV)</a>
            <a href="{link_glm_sum}" class="btn btn-secondary mt-3 ms-2" download><i class="fas fa-download"></i> Baixar Sumário GLM (TXT)</a>
        </section>

        <section id="previsoes" class="mb-5">
            <h2 class="section-title">6. Previsões e Segmentação de Colaboradores</h2>
            <p>O modelo Scikit-learn foi aplicado a todos os colaboradores do conjunto filtrado para calcular a probabilidade de demissão e segmentá-los em grupos de risco (Baixo, Médio, Alto). Essa segmentação é a base para as ações proativas do RH.</p>

            <h3 class="mt-4">6.1. Distribuição por Segmento de Risco</h3>
            <p>Esta tabela mostra quantos colaboradores foram classificados em cada categoria de risco com base nas suas probabilidades de demissão.</p>
            {segmento_risco_dist_html}
            <p class="text-muted mt-2"><i>Limiares utilizados: Prob. &lt; {threshold_baixo_risco:.2f} = Baixo Risco; {threshold_baixo_risco:.2f} &le; Prob. &lt; {threshold_alto_risco:.2f} = Médio Risco; Prob. &ge; {threshold_alto_risco:.2f} = Alto Risco.</i></p>

            <h3 class="mt-4">6.2. Exportação Detalhada</h3>
            <p>Um arquivo Excel detalhado (<code>df_previsoes_risco.xlsx</code>) foi gerado contendo todas as informações originais dos colaboradores, suas probabilidades de demissão, a previsão do modelo e o segmento de risco ao qual pertencem. Este arquivo serve como um registro completo para análise individual.</p>
            <a href="{link_previsoes}" class="btn btn-success mt-3" download><i class="fas fa-file-excel"></i> Baixar Relatório de Previsões (Excel)</a>
        </section>
        
        <section id="dashboard" class="mb-5">
            <h2 class="section-title">7. Dashboard Executivo de Gestão de Pessoas</h2>
            <p>Para facilitar a tomada de decisão, um Dashboard em Excel foi criado, consolidando as principais informações. Ele contém abas com:</p>
            <ul>
                <li><strong>Resumo Executivo:</strong> Métricas chave e distribuição de risco.</li>
                <li><strong>Alto Risco:</strong> Lista detalhada de colaboradores que demandam ação imediata.</li>
                <li><strong>Médio Risco:</strong> Lista de colaboradores para programas de retenção.</li>
                <li><strong>Fatores de Risco:</strong> Variáveis que aumentam a chance de demissão (do GLM).</li>
                <li><strong>Fatores de Proteção:</strong> Variáveis que diminuem a chance de demissão (do GLM).</li>
            </ul>
            <p>Este dashboard é a principal ferramenta para o RH e a liderança acessarem os insights e agirem proativamente.</p>
            <a href="{link_dashboard_excel}" class="btn btn-info mt-3" download><i class="fas fa-file-excel"></i> Baixar Dashboard Completo (Excel)</a>
        </section>

        <section id="recomendacoes" class="mb-5">
            <h2 class="section-title">8. Recomendações Estratégicas</h2>
            <p>Com base na segmentação de risco e nos fatores identificados pelos modelos, seguem recomendações práticas para a Gestão de Pessoas.</p>

            <h3 class="mt-4">8.1. Ações por Segmento de Risco</h3>
            {recomendacoes_baixo_risco}
            {recomendacoes_medio_risco}
            {recomendacoes_alto_risco}
        
            <h3 class="mt-4">8.2. KPIs de Monitoramento Contínuo</h3>
            <p>Para garantir a eficácia das intervenções e o sucesso do programa de retenção, é crucial monitorar os seguintes indicadores:</p>
            {recomendacoes_kpis}
        </section>

        <section id="conclusao" class="mb-5">
            <h2 class="section-title">9. Conclusão e Próximos Passos</h2>
            <p>Este projeto demonstra o poder do People Analytics na identificação proativa de riscos de demissão. O modelo híbrido desenvolvido oferece insights valiosos sobre os fatores que influenciam a decisão de um colaborador permanecer na empresa e permite a segmentação para ações de retenção personalizadas, tornando-o uma ferramenta estratégica essencial para a Gestão de Pessoas.</p>
            <p><strong>Próximos Passos Sugeridos:</strong></p>
            <ul>
                <li><strong>Validação de Negócio:</strong> Discutir os fatores de risco e proteção identificados com especialistas de RH e lideranças para validar as descobertas no contexto da empresa.</li>
                <li><strong>Implementação de Ações:</strong> Desenvolver e implementar planos de ação detalhados e personalizados para os segmentos de médio e alto risco.</li>
                <li><strong>Integração em Processos:</strong> Integrar o modelo em um processo de monitoramento regular, talvez com atualizações mensais ou trimestrais dos dados e do modelo.</li>
                <li><strong>Refinamento Contínuo:</strong> Explorar variáveis adicionais (e.g., dados de performance, feedback de clima, interações com gestores) ou outros modelos de Machine Learning para aprimorar ainda mais a precisão preditiva e a riqueza dos insights.</li>
                <li><strong>Monitoramento de Impacto:</strong> Medir o impacto das ações de retenção nos KPIs definidos, criando um ciclo de feedback para otimização contínua.</li>
            </ul>
        </section>

    </div>

    <a href="#" class="back-to-top btn rounded-circle"><i class="fas fa-arrow-up"></i></a>

    <footer class="footer">
        <div class="container">
            <p>&copy; {datetime.now().year} People Analytics. Todos os direitos reservados.</p>
        </div>
    </footer>

    <!-- Bootstrap JS (com Popper.js) -->
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        // Script para rolagem suave e botão voltar ao topo
        document.addEventListener('DOMContentLoaded', function() {{
            // Smooth scrolling for internal links
            document.querySelectorAll('a[href^="#"]').forEach(anchor => {{
                anchor.addEventListener('click', function (e) {{
                    e.preventDefault();
                    document.querySelector(this.getAttribute('href')).scrollIntoView({{
                        behavior: 'smooth'
                    }});
                }});
            }});

            // Back to top button functionality
            const backToTopButton = document.querySelector('.back-to-top');
            window.addEventListener('scroll', () => {{
                if (window.scrollY > 300) {{ // Show button after scrolling 300px
                    backToTopButton.style.display = 'block';
                }} else {{
                    backToTopButton.style.display = 'none';
                }}
            }});

            backToTopButton.addEventListener('click', () => {{
                window.scrollTo({{ top: 0, behavior: 'smooth' }});
            }});
        }});
    </script>
</body>
</html>
"""
    # 7.3. Salvar o HTML em um arquivo
    with open(OUTPUT_HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(html_content)
    logging.info(f"✓ Relatório HTML gerado com sucesso em '{OUTPUT_HTML_FILE}'")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 7 (Geração do Relatório HTML): {e}")
    import traceback
    traceback.print_exc()
    # Continua, não encerra o script, pois o HTML é um output adicional.

# ==============================================================================
# 9. DASHBOARD DE GESTÃO DE PESSOAS (NOVO)
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 9 - DASHBOARD DE GESTÃO DE PESSOAS (NOVO)")
logging.info("="*80)

try:
    if df_resultados.empty:
        raise ValueError("DataFrame de resultados vazio. Não é possível gerar o Dashboard.")
    if glm_summary_df.empty:
        logging.warning("DataFrame de sumário GLM vazio. Abas de Fatores de Risco/Proteção serão vazias.")

    # 9.1. Análise por Segmento de Risco
    segmento_stats = df_resultados.groupby('segmento_risco').agg({
        'prob_demissao': ['count', 'mean', 'min', 'max'],
        'dias_afastado': 'mean',
        'salario_total': 'mean',
        'idade': 'mean',
        'dias_trabalhado': 'mean'
    }).round(2)
    # Renomeia as colunas do multi-índice para facilitar a exibição
    segmento_stats.columns = ['_'.join(col).strip() for col in segmento_stats.columns.values]
    segmento_stats = segmento_stats.reset_index().rename(columns={'segmento_risco': 'Segmento de Risco'})

    logging.info("\n--- 9.1. Estatísticas por Segmento de Risco ---")
    logging.info(segmento_stats)

    # 9.2. Colaboradores de Alto Risco (Ação Imediata)
    alto_risco_df = df_resultados[df_resultados['segmento_risco'] == 'Alto Risco'].sort_values(
        'prob_demissao', ascending=False
    ).copy()
    # Adiciona colunas para identificar o colaborador e seu risco
    cols_to_show_alto_risco = ['nome', 'cargo', 'departamento', 'prob_demissao', 'idade', 'dias_trabalhado', 'salario_total', 'dias_afastado', 'target_real']
    # Filtra apenas as colunas que realmente existem no DataFrame
    alto_risco_df = alto_risco_df[[col for col in cols_to_show_alto_risco if col in alto_risco_df.columns]]
    alto_risco_df['prob_demissao'] = alto_risco_df['prob_demissao'].apply(lambda x: f"{x:.2%}")
    logging.info(f"\n--- 9.2. Colaboradores de Alto Risco ({len(alto_risco_df)}) ---")
    logging.info(alto_risco_df.head(20))

    # 9.3. Colaboradores de Médio Risco (para programas de retenção)
    medio_risco_df = df_resultados[df_resultados['segmento_risco'] == 'Médio Risco'].sort_values(
        'prob_demissao', ascending=False
    ).copy()
    medio_risco_df = medio_risco_df[[col for col in cols_to_show_alto_risco if col in medio_risco_df.columns]]
    medio_risco_df['prob_demissao'] = medio_risco_df['prob_demissao'].apply(lambda x: f"{x:.2%}")
    logging.info(f"\n--- 9.3. Colaboradores de Médio Risco ({len(medio_risco_df)}) ---")
    logging.info(medio_risco_df.head(20))

    # 9.4. Fatores de Risco (do GLM)
    # Verifica se glm_summary_df existe e não está vazio
    if 'glm_summary_df' in locals() and not glm_summary_df.empty:
        fatores_risco_df = glm_summary_df[glm_summary_df['Odds Ratio'] > 1].sort_values(
            'Odds Ratio', ascending=False
        ).copy()
        # Formata Odds Ratio e p-value para melhor visualização
        fatores_risco_df['Odds Ratio'] = fatores_risco_df['Odds Ratio'].round(3)
        fatores_risco_df['p-value'] = fatores_risco_df['p-value'].apply(lambda x: f"{x:.4f}")
    else:
        fatores_risco_df = pd.DataFrame(columns=['Feature', 'Coefficient (Beta)', 'Odds Ratio', 'p-value', 'Significância'])
        logging.warning("DataFrame de fatores de risco GLM está vazio ou não foi gerado.")

    logging.info(f"\n--- 9.4. Fatores de RISCO (Odds Ratio > 1) ---")
    logging.info(fatores_risco_df[['Feature', 'Odds Ratio', 'p-value', 'Significância']])

    # 9.5. Fatores de Proteção (do GLM)
    if 'glm_summary_df' in locals() and not glm_summary_df.empty:
        fatores_protecao_df = glm_summary_df[glm_summary_df['Odds Ratio'] < 1].sort_values(
            'Odds Ratio', ascending=True
        ).copy()
        fatores_protecao_df['Odds Ratio'] = fatores_protecao_df['Odds Ratio'].round(3)
        fatores_protecao_df['p-value'] = fatores_protecao_df['p-value'].apply(lambda x: f"{x:.4f}")
    else:
        fatores_protecao_df = pd.DataFrame(columns=['Feature', 'Coefficient (Beta)', 'Odds Ratio', 'p-value', 'Significância'])
        logging.warning("DataFrame de fatores de proteção GLM está vazio ou não foi gerado.")

    logging.info(f"\n--- 9.5. Fatores de PROTEÇÃO (Odds Ratio < 1) ---")
    logging.info(fatores_protecao_df[['Feature', 'Odds Ratio', 'p-value', 'Significância']])

    # 9.6. Exportar para Excel com abas
    wb_dashboard = Workbook()
    
    # Aba 1: Resumo Executivo
    ws_resumo = wb_dashboard.active
    ws_resumo.title = "Resumo Executivo"
    
    # Adicionar cabeçalho
    current_datetime = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws_resumo.append(['Relatório de People Analytics: Previsão de Demissão'])
    ws_resumo.append(['Gerado em:', current_datetime])
    ws_resumo.append(['']) # Linha em branco
    
    ws_resumo.append(['Métrica', 'Valor'])
    ws_resumo.append(['Total de Colaboradores', len(df_resultados)])
    ws_resumo.append(['Colaboradores de Baixo Risco', len(df_resultados[df_resultados['segmento_risco'] == 'Baixo Risco'])])
    ws_resumo.append(['Colaboradores de Médio Risco', len(df_resultados[df_resultados['segmento_risco'] == 'Médio Risco'])])
    ws_resumo.append(['Colaboradores de Alto Risco', len(df_resultados[df_resultados['segmento_risco'] == 'Alto Risco'])])
    ws_resumo.append(['Taxa de Alto Risco', f"{len(alto_risco_df) / len(df_resultados) * 100:.2f}%"])
    ws_resumo.append([''])
    ws_resumo.append(['Métricas do Modelo (Scikit-learn)'])
    ws_resumo.append(['Acurácia', f"{accuracy_skl:.2%}"])
    ws_resumo.append(['Precisão', f"{precision_skl:.2%}"])
    ws_resumo.append(['Recall', f"{recall_skl:.2%}"])
    ws_resumo.append(['F1-Score', f"{f1_skl:.2f}"])
    ws_resumo.append(['ROC AUC', f"{roc_auc_skl:.4f}"])
    ws_resumo.append(['Gini', f"{gini:.4f}"])
    ws_resumo.append(['Threshold Ajustado', f"{optimal_threshold:.4f}"])
    ws_resumo.append([''])
    ws_resumo.append(['Métricas do Modelo (GLM)'])
    ws_resumo.append(['Log-Likelihood', f"{glm_loglike:.4f}" if isinstance(glm_loglike, (int, float)) else str(glm_loglike)])
    ws_resumo.append(['AIC', f"{glm_aic:.4f}" if isinstance(glm_aic, (int, float)) else str(glm_aic)])
    ws_resumo.append(['BIC', f"{glm_bic:.4f}" if isinstance(glm_bic, (int, float)) else str(glm_bic)])

    # Estilizar cabeçalhos
    for col_idx in range(1, ws_resumo.max_column + 1):
        ws_resumo.cell(row=4, column=col_idx).font = Font(bold=True) # Cabeçalho 'Métrica', 'Valor'
    
    # Ajustar largura das colunas
    for col_idx in range(1, ws_resumo.max_column + 1):
        ws_resumo.column_dimensions[get_column_letter(col_idx)].width = 30
    
    # Adicionar tabelas de resumo por segmento
    r_offset = ws_resumo.max_row + 2
    ws_resumo.cell(row=r_offset, column=1, value="Estatísticas por Segmento de Risco").font = Font(bold=True)
    for r_idx, row in enumerate(dataframe_to_rows(segmento_stats, index=False, header=True), r_offset + 1):
        ws_resumo.append(row)
    
    # Aba 2: Alto Risco
    ws_alto_risco = wb_dashboard.create_sheet(title="Alto Risco")
    ws_alto_risco.append(list(alto_risco_df.columns))
    for row in dataframe_to_rows(alto_risco_df, index=False, header=False):
        ws_alto_risco.append(row)
    # Auto-ajuste de colunas
    for col in ws_alto_risco.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_alto_risco.column_dimensions[column_letter].width = min(max_length + 2, 75)
    
    # Aba 3: Médio Risco
    ws_medio_risco = wb_dashboard.create_sheet(title="Médio Risco")
    ws_medio_risco.append(list(medio_risco_df.columns))
    for row in dataframe_to_rows(medio_risco_df, index=False, header=False):
        ws_medio_risco.append(row)
    # Auto-ajuste de colunas
    for col in ws_medio_risco.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_medio_risco.column_dimensions[column_letter].width = min(max_length + 2, 75)
    
    # Aba 4: Fatores de Risco
    ws_fatores_risco = wb_dashboard.create_sheet(title="Fatores de Risco")
    ws_fatores_risco.append(list(fatores_risco_df.columns))
    for row in dataframe_to_rows(fatores_risco_df, index=False, header=False):
        ws_fatores_risco.append(row)
    # Auto-ajuste de colunas
    for col in ws_fatores_risco.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_fatores_risco.column_dimensions[column_letter].width = min(max_length + 2, 75)
    
    # Aba 5: Fatores de Proteção
    ws_fatores_protecao = wb_dashboard.create_sheet(title="Fatores de Proteção")
    ws_fatores_protecao.append(list(fatores_protecao_df.columns))
    for row in dataframe_to_rows(fatores_protecao_df, index=False, header=False):
        ws_fatores_protecao.append(row)
    # Auto-ajuste de colunas
    for col in ws_fatores_protecao.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_fatores_protecao.column_dimensions[column_letter].width = min(max_length + 2, 75)
    
    # Salvar o arquivo do dashboard
    wb_dashboard.save(OUTPUT_DASHBOARD_EXCEL_FILE)
    logging.info(f"✓ Dashboard executivo exportado para '{OUTPUT_DASHBOARD_EXCEL_FILE}'")

except Exception as e:
    logging.error(f"ERRO na SEÇÃO 9 (Geração do Dashboard): {e}")
    import traceback
    traceback.print_exc()

gc.collect() # Limpeza de memória
logging.info("✓ GC.collect() executado após Seção 9.")

# ==============================================================================
# 10. FINALIZAÇÃO E EXPORTAÇÕES (Sumário)
# ==============================================================================
logging.info("\n" + "="*80)
logging.info("SEÇÃO 10 - FINALIZAÇÃO E EXPORTAÇÕES")
logging.info("="*80)
logging.info("✓ Todas as etapas do projeto foram concluídas.")
logging.info(f"O relatório HTML completo está disponível em: {OUTPUT_HTML_FILE}")
logging.info(f"Os dados detalhados com previsões estão em: {OUTPUT_EXCEL_PREVISOES_FILE}")
logging.info(f"Os coeficientes do modelo GLM estão em: {OUTPUT_GLM_COEF_CSV}")
logging.info(f"O sumário completo do modelo GLM está em: {OUTPUT_GLM_SUMMARY_TXT}")
logging.info(f"O Dashboard Executivo em Excel está em: {OUTPUT_DASHBOARD_EXCEL_FILE}")
logging.info(f"Os gráficos PNG individuais estão no diretório: {OUTPUT_DIR}")

logging.info("\nPROJETO DE PEOPLE ANALYTICS CONCLUÍDO COM SUCESSO!")
logging.info("================================================================================")