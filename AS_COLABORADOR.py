from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import logging
import os
import string
import base64
from io import BytesIO
from collections import Counter
import webbrowser

logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')

# ============================================================================
# IMPORTAR NLTK COM TRATAMENTO DE ERRO
# ============================================================================

logging.info("📦 Importando NLTK...")
try:
    import nltk
    import subprocess
    import sys
    subprocess.run([sys.executable, "-m", "nltk.downloader", "stopwords", "vader_lexicon"],
                   capture_output=True, timeout=60)
    
    from nltk.sentiment.vader import SentimentIntensityAnalyzer
    from nltk.corpus import stopwords
    logging.info("   ✓ NLTK importado com sucesso")
    NLTK_AVAILABLE = True
    sia = SentimentIntensityAnalyzer()
except Exception as e:
    logging.warning(f"   ⚠️ NLTK não disponível: {e}")
    NLTK_AVAILABLE = False
    sia = None

# ============================================================================
# DEMAIS IMPORTAÇÕES
# ============================================================================

import pandas as pd
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import spacy
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

logging.info("✓ Todas as importações completas")

# ============================================================================
# CONFIGURAÇÕES
# ============================================================================

# Carregamento da base de controle de processos

id = 23

path_registros_processos = r'X:\Gestao_de_Pessoas\Analytics\03 - Bases\1. BASES TRATADAS\PROCESSOS.xlsx'

registros_processos = pd.read_excel(path_registros_processos, sheet_name = "REGISTROS", engine='openpyxl')

wb_p = load_workbook(path_registros_processos)

ws_p = wb_p['REGISTROS']

# Controle de atualização de processo: Etapa 0

linha_0 = [id, datetime.today(), 23]

ws_p.append(linha_0)

wb_p.save(path_registros_processos)

PATH_ENTREVISTA = r"X:\Gestao_de_Pessoas\Analytics\08 - Notebooks Python\08.5 - Estudos e Projetos\Análise de Sentimentos\Entrevista de Desligamento\Entrevista_Desligamento.xlsb"
PATH_HC = r"X:\Gestao_de_Pessoas\Analytics\10 - Relatórios\10.4 - HC e Atestados Médicos\Controle_HC e Atestados.xlsb"
PATH_LOGO = r"X:\Gestao_de_Pessoas\Analytics\08 - Notebooks Python\08.5 - Estudos e Projetos\Análise de Sentimentos\Logo AFPESP.png"
PASTA_DESTINO = r'X:\Gestao_de_Pessoas\Analytics\08 - Notebooks Python\08.5 - Estudos e Projetos\Análise de Sentimentos\Entrevista de Desligamento'
CAMINHO_RELATORIO_HTML = os.path.join(PASTA_DESTINO, 'Relatorio_Sentimentos_Colaborador.html')
CAMINHO_RELATORIO_DADOS = os.path.join(PASTA_DESTINO, 'Dados_Sentimentos_Colaborador.xlsx')

# ============================================================================
# STOPWORDS
# ============================================================================

if NLTK_AVAILABLE:
    STOPWORDS_PT = set(stopwords.words('portuguese'))
else:
    STOPWORDS_PT = {
        'a', 'o', 'e', 'de', 'da', 'do', 'em', 'um', 'uma', 'para', 'com', 'por', 'mas',
        'mais', 'ou', 'quando', 'muito', 'já', 'está', 'eu', 'também', 'só', 'pelo', 'pela',
        'até', 'isso', 'ela', 'entre', 'era', 'depois', 'sem', 'mesmo', 'aos', 'ter', 'seus',
        'pra', 'pro', 'tbm', 'tb', 'vc', 'vcs', 'q', 'p', 'n', 'tá', 'tô', 'ta', 'to', 'né'
    }

# ============================================================================
# ☆☆☆ LÉXICO DE SENTIMENTOS CUSTOMIZADO ☆☆☆
# ============================================================================

PALAVRAS_NEGATIVAS = {
    # Ofensas e desrespeito (peso máximo)
    'ofendeu': -2.0, 'ofendias': -2.0, 'ofendendo': -2.0, 'ofender': -2.0,
    'desrespeitou': -2.0, 'desrespeitar': -2.0, 'desrespeitosa': -2.0, 'desrespeitoso': -2.0, 'sem respeito': -2.0,
    'desrespeitava': -2.0, 'desrespeitavam': -2.0,
    
    # Insubordinação e indisciplina
    'insubordinação': -2.0, 'insubordinada': -2.0, 'insubordinado': -2.0, 'insubordinados': -2.0,
    
    # Comportamentos inadequados
    'inadequado': -2.0, 'inadequada': -2.0, 'inadequados': -2.0, 'inadequadas': -2.0, 'desleixo': -1.8, 'mau comportamento': -2.0,
    'comportamento ruim': -2.0, 'comportamento inadequado': -2.0, 'tóxico': -1.0,
    
    # Linguagem pesada
    'gritando': -1.8, 'gritou': -1.8, 'gritar': -1.8, 'grita': -1.8,
    'linguajar': -1.5, 'linguagem': -0.5, 'fofoca': -0.5,
    'grosseira': -2.0, 'grosseiro': -2.0, 'grosseiros': -2.0,
    'grosseiras': -2.0,
    
    # Atrito e conflito
    'atrito': -1.8, 'atritos': -1.8, 'acumulo': -0.5, 'preconceito': -2.0,
    'conflito': -1.8, 'conflitos': -1.8, 'autoritário': -0.5, 'autoritária': -0.5,
    'problemas de relacionamento': -2.0, 'relacionamento com a liderança': -1.0,
    'dificuldade de relacionamento': -2.0,
    'dificuldades de relacionamento': -2.0,
    'relação ruim': -2.0, 'relações ruins': -2.0,
    
    # Faltas disciplinares
    'suspensa': -1.8, 'suspenso': -1.8, 'suspensos': -1.8, 'suspensão': -1.8, 'suspensões': -1.8, 'freio': -1.5, 
    'falta': -1.0, 'faltas': -1.0, 'atestados': -1.0, 'atestado': -1.0, 'sem freio': -1.5, 'atrasado': -1.0, 'atrasar': -1.0,
    'ausência': -0.5,
    
    # Ações trabalhistas
    'ação trabalhista': -2.0, 'ações trabalhistas': -2.0,
    'ação judicial': -2.0, 'ações judiciais': -2.0,
    'processo judicial': -2.0, 'processos judiciais': -2.0,
    'processo trabalhista': -2.0,
    'insalubridade': -1.5, 'advertência': -0.5,
    
    # Influência negativa
    'influenciava negativamente': -2.0, 'influência negativa': -2.0, 'influencia negativa': -2.0, 'mentira': -1.0,
    'influenciava': -1.5,
    
    # Problemas gerais
    'problema': -1.0, 'problemas': -1.0, 'humilhado': -1.0, 'humilhada': -1.0, 'humilhou': -1.0, 'intolerância religiosa': -1.0,
    'dificuldade': -1.0, 'dificuldades': -1.0, 'manipulação': -1.0, 'agredido': -2.0, 'agredida': -2.0, 'motivação': -0.5,
    'ruim': -1.5, 'ruins': -1.5, 'desinteresse': -1.0, 'melhorias': -0.5, 'não é compatível': -0.5, 'feedback': -0.5,
    'mau': -1.5, 'má': -1.5, 'maus': -1.5, 'medo': -0.5, 'falta de oportunidade': -1.0,
    'desestimulante': -0.5, 'desestimula': -0.5, 'péssimo': -2.0, 'péssima': -2.0,
}

PALAVRAS_POSITIVAS = {
    'bom': 0.8, 'boa': 0.8, 'bons': 0.8, 'boas': 0.8, 'agradeço': 0.8, 'agradecer': 0.8, 'obrigado': 0.8, 'oportunidade': 0.5,
    'ótimo': 1.0, 'ótima': 1.0, 'ótimos': 1.0, 'ótimas': 1.0, 'maravilha': 1.0, 'maravilhoso': 1.0, 'muito bom': 1.0,
    'excelente': 1.0, 'excelentes': 1.0, 'com prazer': 1.0, 'consolidada': 0.5, 'séria': 0.5, 'gratidão': 1.0,
    'agilidade': 0.8, 'agilidades': 0.8, 'respeitosas': 0.8, 'respeitosa': 0.8,
    'qualidade': 0.8, 'qualidades': 0.8, 'salário': 0.5, 'benefício': 0.5,
    'eficiente': 0.8, 'eficientes': 0.8, 'voltar': 0.8,
    'dedicado': 0.8, 'dedicada': 0.8, 'dedicados': 0.8, 'dedicadas': 0.8,
    'competente': 0.8, 'competentes': 0.8,
    'responsável': 0.8, 'responsáveis': 0.8,
    'pontual': 0.8, 'pontuais': 0.8, 'estrutura': 0.8,
}

def get_polarity_customized(text):
    """Análise de sentimentos customizada."""
    if not isinstance(text, str):
        text = str(text)
    
    text_lower = text.lower()
    
    # Contar negativas e positivas
    negative_score = 0.0
    positive_score = 0.0
    
    for palavra_neg, peso in PALAVRAS_NEGATIVAS.items():
        if palavra_neg in text_lower:
            negative_score += peso
    
    for palavra_pos, peso in PALAVRAS_POSITIVAS.items():
        if palavra_pos in text_lower:
            positive_score += peso
    
    # Se muitas negativas, prevalecer o negativo
    if negative_score <= -3.0:
        return -1.0
    
    if negative_score < -0.5:
        return max(-1.0, negative_score + positive_score * 0.3)
    
    if positive_score > 0.5:
        return min(1.0, positive_score - negative_score * 0.5)
    
    # Combinar com VADER se disponível
    if NLTK_AVAILABLE and sia is not None:
        try:
            vader_score = sia.polarity_scores(text)["compound"]
            combined = (negative_score + positive_score) * 0.6 + vader_score * 0.4
            return max(-1.0, min(1.0, combined))
        except:
            pass
    
    final_score = negative_score + positive_score
    return max(-1.0, min(1.0, final_score / 10))

# ============================================================================
# FUNÇÕES
# ============================================================================

def image_to_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode('utf-8')
    except:
        return None

def filter_words(words_list):
    palavras_genericas = {
        'mesma', 'mesmo', 'mesmos', 'mesmas', 'gostaria', 'gostarias', 'gostariam', 'sobre', 'sobres', 'onde', 'aonde', 'porém',
        'porem', 'sempre', 'semper', 'quando', 'quandos', 'assim', 'assima', 'assimismo', 'então', 'entao', 'entaum', 'também',
        'tambem', 'tamben', 'boa', 'outro', 'melhor',
        'agora', 'agoras',
        'depois', 'depoi',
        'antes', 'antes',
        'nunca', 'nunqua',
        'ainda', 'ainfa',
        'muito', 'muita', 'muitos', 'muitas',
        'pouco', 'pouca', 'poucos', 'poucas',
        'bastante', 'bastantes',
        'apenas', 'apena',
        'somente', 'sómente',
        'só', 'soe', 'começou', 'aprender',
        'já', 'ja', 'damaris',
        'talvez', 'talves',
        'tipo', 'tipos', 'pra', 'pois',
        'coisa', 'coisas','dia',
        'forma', 'formas','afpesp',
        'modo', 'modos','tudo',
        'maneira', 'maneiras',
        'jeito', 'jeitos','bom','todos',
        'lado', 'lados',
        'ponto', 'pontos',
        'parte', 'partes',
        'vez', 'vezes', 'fazer', 'ter',
        'momento', 'momentos',
        'tempo', 'tempos',
    }
    
    stopwords_expandido = STOPWORDS_PT.union(palavras_genericas)
    
    return [
        word.lower() 
        for word in words_list 
        if (word.lower() not in stopwords_expandido and 
            len(word) >= 3 and 
            word.isalpha())
    ]

def extract_wordcloud_base64(texts, background_color='black', colormap='viridis'):
    if not texts:
        return ""
    text = " ".join([str(t) for t in texts if t])
    words = text.lower().translate(str.maketrans("", "", string.punctuation)).split()
    filtered = filter_words(words)
    if not filtered:
        return ""
    try:
        wc = WordCloud(width=900, height=450, background_color=background_color,
                      colormap=colormap, random_state=42).generate_from_frequencies(
                      Counter(filtered))
        buffer = BytesIO()
        wc.to_image().save(buffer, format="PNG")
        buffer.seek(0)
        img_str = base64.b64encode(buffer.getvalue()).decode("utf-8")
        return f"data:image/png;base64,{img_str}"
    except Exception as e:
        logging.error(f"Erro WordCloud: {e}")
        return ""

def create_bar_chart_base64(data_dict, title, color='#0078d4'):
    if not data_dict:
        return ""
    try:
        fig, ax = plt.subplots(figsize=(12, 6))
        fig.patch.set_facecolor('white')
        bars = ax.bar(list(data_dict.keys()), list(data_dict.values()),
                     color=color, edgecolor='#333333', linewidth=2, alpha=0.85)
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height, f'{int(height)}',
                   ha='center', va='bottom', fontsize=11, fontweight='bold')
        ax.set_title(title, fontsize=14, fontweight='bold', pad=20, color='#005a9c')
        ax.set_ylabel('Frequência', fontsize=12, fontweight='bold')
        ax.grid(axis='y', alpha=0.2, linestyle='--')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
        buffer.seek(0)
        img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
        plt.close(fig)
        return f"data:image/png;base64,{img_str}"
    except Exception as e:
        logging.error(f"Erro gráfico: {e}")
        return ""

def create_sentiment_distribution_chart(df):
    if df.empty:
        return "", {}
    try:
        counts = df["sentimento"].value_counts().to_dict()
        fig, ax = plt.subplots(figsize=(12, 7))
        fig.patch.set_facecolor('white')
        colors = {'positivo': '#28a745', 'negativo': '#dc3545', 'neutro': '#6c757d'}
        color_list = [colors.get(k, '#999999') for k in counts.keys()]
        bars = ax.bar(counts.keys(), counts.values(), color=color_list,
                     edgecolor='#333333', linewidth=2.5, alpha=0.85)
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height, f'{int(height)}',
                   ha='center', va='bottom', fontsize=16, fontweight='bold')
        ax.set_title('Distribuição de Sentimentos', fontsize=16, fontweight='bold',
                    pad=25, color='#005a9c')
        ax.set_xlabel('Sentimento', fontsize=12, fontweight='bold')
        ax.set_ylabel('Quantidade', fontsize=12, fontweight='bold')
        ax.grid(axis='y', alpha=0.2)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.tight_layout()
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
        buffer.seek(0)
        img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
        plt.close(fig)
        return f"data:image/png;base64,{img_str}", counts
    except Exception as e:
        logging.error(f"Erro distribuição: {e}")
        return "", {}

# ============================================================================
# ☆☆☆ CARREGAR DADOS (AQUI ESTÁ O PRINCIPAL) ☆☆☆
# ============================================================================

logging.info("📂 Carregando dados...")
df_entrevista = pd.read_excel(PATH_ENTREVISTA, sheet_name='COLABORADOR', engine='pyxlsb')
logging.info(f"   ✓ {len(df_entrevista)} registros")

df_entrevista.dropna(subset=['comentarios', 'registro'], inplace=True)

logging.info("💭 Analisando sentimentos com LÉXICO CUSTOMIZADO...")

# ☆ AQUI USA A FUNÇÃO get_polarity_customized ☆
df_entrevista["polaridade"] = df_entrevista["comentarios"].apply(get_polarity_customized)

df_entrevista["sentimento"] = df_entrevista["polaridade"].apply(
    lambda score: "positivo" if score >= 0.1 else ("negativo" if score <= -0.1 else "neutro")
)

pos = (df_entrevista['sentimento'] == 'positivo').sum()
neu = (df_entrevista['sentimento'] == 'neutro').sum()
neg = (df_entrevista['sentimento'] == 'negativo').sum()
logging.info(f"   ✓ Análise concluída")

# Carregar HC
df_final = df_entrevista.copy()
try:
    logging.info("📂 Carregando HC...")
    df_hc = pd.read_excel(PATH_HC, sheet_name='HC')
    if not df_hc.empty:
        df_entrevista['registro'] = pd.to_numeric(df_entrevista['registro'], errors='coerce')
        df_hc['registro'] = pd.to_numeric(df_hc['registro'], errors='coerce')
        df_entrevista.dropna(subset=['registro'], inplace=True)
        df_hc.dropna(subset=['registro'], inplace=True)
        df_final = pd.merge(df_entrevista, df_hc, on='registro', how='left', 
                           suffixes=('_entrevista', '_hc'))
        logging.info(f"   ✓ Merge concluído: {len(df_final)} registros\n")
except Exception as e:
    logging.warning(f"   ⚠️ HC não disponível\n")

# ============================================================================
# SALVAR XLSX
# ============================================================================

logging.info("💾 Salvando XLSX...")
colunas = ['registro', 'nome', 'sexo', 'data_admissao', 'situacao', 'data_rescisao',
           'descricao_rescisao', 'cargo_completo', 'centro_custo', 'unidade', 
           'empresa_resumo', 'comentarios', 'polaridade', 'sentimento']
cols_existentes = [c for c in colunas if c in df_final.columns]
df_rel = df_final[cols_existentes].copy()

# ===== FUNÇÃO PARA CONVERTER DATAS =====
def converter_data_excel(valor):
    """Converte valor em data, independente do formato de origem"""
    if pd.isna(valor) or valor == '' or valor == 'nan':
        return pd.NaT
    
    # Se já é datetime, retorna
    if isinstance(valor, pd.Timestamp):
        return valor
    
    # Se é número (serial do Excel)
    if isinstance(valor, (int, float)):
        try:
            return pd.Timestamp('1899-12-30') + pd.Timedelta(days=float(valor))
        except:
            return pd.NaT
    
    # Se é string, tenta converter
    if isinstance(valor, str):
        try:
            return pd.to_datetime(valor, format='%d/%m/%Y', errors='coerce')
        except:
            try:
                return pd.to_datetime(valor, errors='coerce')
            except:
                return pd.NaT
    
    return pd.NaT

# ===== APLICAR CONVERSÃO =====
if 'data_admissao' in df_rel.columns:
    logging.info("🔄 Convertendo data_admissao...")
    df_rel['data_admissao'] = df_rel['data_admissao'].apply(converter_data_excel)
    logging.info(f"   ✓ Amostra: {df_rel['data_admissao'].iloc[0]}")

if 'data_rescisao' in df_rel.columns:
    logging.info("🔄 Convertendo data_rescisao...")
    df_rel['data_rescisao'] = df_rel['data_rescisao'].apply(converter_data_excel)
    logging.info(f"   ✓ Amostra: {df_rel['data_rescisao'].iloc[0]}")

try:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    # ===== SALVAR SEM date_format GLOBAL =====
    with pd.ExcelWriter(CAMINHO_RELATORIO_DADOS, engine='openpyxl') as writer:
        df_rel.to_excel(writer, sheet_name='Dados_Sentimentos', index=False)
        ws = writer.sheets['Dados_Sentimentos']
        
        # ===== FORMATAR HEADER =====
        header_fill = PatternFill(start_color='005A9C', end_color='005A9C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_idx, col_name in enumerate(df_rel.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.freeze_panes = 'A2'
        
        # ===== FORMATAR DATAS (CÉLULA POR CÉLULA) =====
        for idx, col_name in enumerate(df_rel.columns, 1):
            if col_name in ['data_admissao', 'data_rescisao']:
                col_letter = get_column_letter(idx)
                
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    # Apenas formata se o valor não é vazio/NaT
                    if cell.value and pd.notna(cell.value):
                        cell.number_format = 'dd/mm/yyyy'
        
        # ===== AJUSTAR LARGURA DAS COLUNAS =====
        for idx, col_name in enumerate(df_rel.columns, 1):
            col_letter = get_column_letter(idx)
            max_length = len(str(col_name)) + 2
            
            for row in range(2, min(ws.max_row + 1, 1000)):  # Limita para performance
                try:
                    cell_value = ws[f'{col_letter}{row}'].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass
            
            adjusted_width = min(max_length, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    logging.info(f"   ✓ XLSX salvo com sucesso: {CAMINHO_RELATORIO_DADOS}\n")

except Exception as e:
    logging.error(f"❌ Erro ao salvar XLSX: {e}\n")
    import traceback
    traceback.print_exc()

# ============================================================================
# GERAR GRÁFICOS
# ============================================================================

logging.info("📈 Gerando gráficos...")
logo = image_to_base64(PATH_LOGO)
wordcloud_img = extract_wordcloud_base64(df_final['comentarios'].astype(str).tolist(),
                                        background_color='white', colormap='viridis')
sentiment_chart, sentiment_dict = create_sentiment_distribution_chart(df_final)

neg_comments = df_final[df_final["sentimento"] == "negativo"]["comentarios"].astype(str).tolist()
wordcloud_neg = extract_wordcloud_base64(neg_comments, background_color='#2d2d2d', colormap='Reds')
neg_chart = ""
if neg_comments:
    neg_words = "".join(neg_comments).lower().translate(str.maketrans("", "", string.punctuation)).split()
    neg_filtered = filter_words(neg_words)
    neg_freq = Counter(neg_filtered).most_common(15)
    if neg_freq:
        neg_chart = create_bar_chart_base64(dict(neg_freq), 'Top 15 Palavras Negativas', color='#dc3545')

pos_comments = df_final[df_final["sentimento"] == "positivo"]["comentarios"].astype(str).tolist()
wordcloud_pos = extract_wordcloud_base64(pos_comments, background_color='white', colormap='Greens')
pos_chart = ""
if pos_comments:
    pos_words = "".join(pos_comments).lower().translate(str.maketrans("", "", string.punctuation)).split()
    pos_filtered = filter_words(pos_words)
    pos_freq = Counter(pos_filtered).most_common(15)
    if pos_freq:
        pos_chart = create_bar_chart_base64(dict(pos_freq), 'Top 15 Palavras Positivas', color='#28a745')

logging.info("   ✓ Gráficos prontos\n")

# ============================================================================
# GERAR HTML
# ============================================================================

logging.info("🎨 Gerando HTML...")

def img_tag(b64, txt="Gráfico"):
    return f'<img src="{b64}" alt="{txt}">' if b64 and b64.startswith("data:image") else f'<p>⚠️ {txt} não disponível</p>'

total = len(df_final)
pos_count = sentiment_dict.get('positivo', 0)
neu_count = sentiment_dict.get('neutro', 0)
neg_count = sentiment_dict.get('negativo', 0)

html = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Análise de Sentimentos - AFPESP</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', sans-serif; background: #f5f5f5; color: #333; }}
        .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
        header {{ background: linear-gradient(135deg, #005a9c, #003d6b); color: white; padding: 40px; text-align: center; border-radius: 8px; margin-bottom: 30px; }}
        header h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
        .section {{ background: white; padding: 30px; margin-bottom: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .section h2 {{ color: #005a9c; border-bottom: 3px solid #28a745; padding-bottom: 15px; margin-bottom: 20px; }}
        .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .metric {{ background: #f5f5f5; padding: 25px; border-radius: 8px; border-left: 5px solid #6c757d; text-align: center; }}
        .metric.pos {{ border-left-color: #28a745; background: #f0fdf4; }}
        .metric.neg {{ border-left-color: #dc3545; background: #fdf0f0; }}
        .metric.neu {{ border-left-color: #ffc107; background: #fffdf0; }}
        .metric h3 {{ margin-bottom: 10px; }}
        .metric .value {{ font-size: 2.5em; font-weight: bold; color: #005a9c; margin: 15px 0; }}
        .chart {{ margin: 20px 0; text-align: center; }}
        .chart img {{ max-width: 100%; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .btn {{ display: inline-block; background: #28a745; color: white; padding: 12px 30px; border-radius: 5px; text-decoration: none; font-weight: bold; margin: 10px 0; }}
        .btn:hover {{ background: #218838; }}
        .footer {{ text-align: center; padding: 30px; color: #666; border-top: 1px solid #ddd; margin-top: 40px; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📊 Análise de Sentimentos - Entrevista de Desligamentos (Colaborador)</h1>
            <p>Relatório Automático - AFPESP</p>
        </header>
        
        <div class="section">
            <h2>📄 Download</h2>
            <a href="file:///{CAMINHO_RELATORIO_DADOS}" class="btn">📥 Baixar XLSX</a>
        </div>
        
        <div class="metrics">
            <div class="metric"><h3>Total</h3><div class="value">{total}</div></div>
            <div class="metric pos"><h3>Positivos</h3><div class="value">{pos_count}</div><p>{pos_count / max(total, 1) * 100:.1f}%</p></div>
            <div class="metric neu"><h3>Neutros</h3><div class="value">{neu_count}</div><p>{neu_count / max(total, 1) * 100:.1f}%</p></div>
            <div class="metric neg"><h3>Negativos</h3><div class="value">{neg_count}</div><p>{neg_count / max(total, 1) * 100:.1f}%</p></div>
        </div>
        
        <div class="section">
            <h2>📈 Distribuição</h2>
            <div class="chart">{img_tag(sentiment_chart)}</div>
        </div>
        
        <div class="section">
            <h2>🔤 Palavras Gerais</h2>
            <div class="chart">{img_tag(wordcloud_img)}</div>
        </div>
        
        <div class="section">
            <h2>😊 Positivos ({len(pos_comments)})</h2>
            <div class="chart">{img_tag(wordcloud_pos)}</div>
            <div class="chart">{img_tag(pos_chart)}</div>
        </div>
        
        <div class="section">
            <h2>😞 Negativos ({len(neg_comments)})</h2>
            <div class="chart">{img_tag(wordcloud_neg)}</div>
            <div class="chart">{img_tag(neg_chart)}</div>
        </div>
        
        <div class="footer">
            <p>✅ Relatório gerado automaticamente - AFPESP © 2025</p>
        </div>
    </div>
</body>
</html>
"""

try:
    with open(CAMINHO_RELATORIO_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    logging.info(f"✅ HTML salvo: {CAMINHO_RELATORIO_HTML}")
    webbrowser.open(f'file:///{os.path.abspath(CAMINHO_RELATORIO_HTML)}')
except Exception as e:
    logging.error(f"❌ Erro HTML: {e}")

logging.info("\n✨ Processo concluído!")